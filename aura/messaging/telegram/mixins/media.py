"""
MediaMixin — photo/document/voice/inline handlers, PDF/DOCX/spreadsheet extraction,
             document context storage, image analysis.
"""
from __future__ import annotations

import asyncio
import base64
import io
import logging
import os
import tempfile
from typing import Optional, Tuple

try:
    from telegram import InlineQueryResultArticle, InputTextMessageContent, Update
    from telegram.constants import ParseMode
    from telegram.ext import ContextTypes
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    Update = None

logger = logging.getLogger(__name__)


class MediaMixin:
    """Photo, document, voice, and inline query handlers."""

    # Image extensions recognized when sent as document attachments
    _IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tiff", ".tif"}

    # Text/code file extensions for document upload Q&A
    _TEXT_EXTENSIONS = {
        ".txt", ".csv", ".json", ".py", ".md", ".js", ".ts", ".jsx", ".tsx",
        ".html", ".css", ".xml", ".yaml", ".yml", ".toml", ".ini", ".cfg",
        ".sh", ".bash", ".sql", ".log", ".env", ".rs", ".go", ".java",
        ".c", ".cpp", ".h", ".hpp", ".rb", ".php", ".swift", ".kt",
    }

    _MAX_DOC_SIZE = 20 * 1024 * 1024  # 20 MB

    # Quick-action prefixes: keyword -> (system_prompt_prefix, result_title_prefix)
    _INLINE_ACTIONS = {
        "translate": (
            "Translate the following text. Auto-detect the source language. "
            "If the text is in English, translate to Russian. Otherwise translate to English. "
            "Only output the translation, nothing else:\n\n",
            "Translate"
        ),
        "explain": (
            "Give a clear, concise explanation of the following topic. "
            "Keep it under 200 words:\n\n",
            "Explain"
        ),
        "summarize": (
            "Summarize the following text in 2-3 sentences. Be concise:\n\n",
            "Summarize"
        ),
    }

    def _analyze_image_sync(self, img_b64: str, prompt: str) -> str:
        """Analyze a base64-encoded image with the vision model (synchronous)."""
        try:
            from aura.tools.vision import VisionTool

            # Reuse brain's Ollama client if available
            brain = getattr(self.aura, 'brain', None)
            if brain is None:
                brain = getattr(getattr(self.aura, 'agent', None), 'brain', None)

            vt = VisionTool(brain=brain)
            content, model_used = vt._analyze_with_fallback(img_b64, prompt)
            logger.info(f"[TelegramBot] Vision analysis done with {model_used}")
            return content
        except Exception as e:
            logger.error(f"[TelegramBot] Vision analysis failed: {e}", exc_info=True)
            return "Sorry, I couldn't analyze that image right now. The vision model may be unavailable."

    async def _handle_photo_upload(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle photos sent as compressed Telegram images -- analyze with vision model."""

        user = update.effective_user
        chat_id = str(update.effective_chat.id)

        if not self._is_user_allowed(user.id):
            return

        max_per_min = self.config.get("max_messages_per_minute", 20)
        from aura.messaging.telegram.bot import _check_rate_limit
        if not _check_rate_limit(str(user.id), max_per_min):
            await update.message.reply_text("You're sending messages too fast. Please wait a moment.")
            return

        # Show typing while we download and process
        await self.send_typing_indicator(chat_id)

        # Get highest resolution photo (last in the list)
        photo = update.message.photo[-1]
        prompt = update.message.caption or "Describe this image in detail."

        # Download photo bytes from Telegram
        try:
            tg_file = await context.bot.get_file(photo.file_id)
            buf = io.BytesIO()
            await tg_file.download_to_memory(buf)
            image_bytes = buf.getvalue()
        except Exception as e:
            logger.error(f"[TelegramBot] Failed to download photo: {e}")
            await update.message.reply_text("Couldn't download the photo. Please try again.")
            return

        img_b64 = base64.b64encode(image_bytes).decode("utf-8")

        # Refresh typing -- vision models can take a while
        await self.send_typing_indicator(chat_id)

        # Run synchronous vision analysis in a thread
        response = await asyncio.to_thread(self._analyze_image_sync, img_b64, prompt)

        await update.message.reply_text(response)
        self._save_state()

    async def _handle_document(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle document uploads -- extract text for PDFs/text files, route images to vision."""
        user = update.effective_user
        chat_id = str(update.effective_chat.id)

        if not self._is_user_allowed(user.id):
            return

        max_per_min = self.config.get("max_messages_per_minute", 20)
        from aura.messaging.telegram.bot import _check_rate_limit
        if not _check_rate_limit(str(user.id), max_per_min):
            await update.message.reply_text(
                "You're sending messages too fast. Please wait a moment."
            )
            return

        doc = update.message.document
        if not doc:
            return

        filename = doc.file_name or "unknown"
        file_size = doc.file_size or 0
        mime = doc.mime_type or ""
        ext = os.path.splitext(filename)[1].lower()

        # Size guard
        if file_size > self._MAX_DOC_SIZE:
            await update.message.reply_text(
                f"That file is too large ({file_size / (1024*1024):.1f} MB). "
                f"I can handle files up to 20 MB."
            )
            return

        await self.send_typing_indicator(chat_id)

        try:
            # --- PDF ---
            if ext == ".pdf" or mime == "application/pdf":
                text, char_count = await self._extract_pdf(doc)
                await self._ingest_document(user.id, filename, text, update)
                await update.message.reply_text(
                    f"\U0001f4c4 Indexed {char_count:,} characters from {filename}.\n"
                    f"Ask me anything about this document!"
                )

            # --- Text / code files ---
            elif ext in self._TEXT_EXTENSIONS or mime.startswith("text/"):
                text, char_count = await self._extract_text_file(doc)
                await self._ingest_document(user.id, filename, text, update)
                await update.message.reply_text(
                    f"\U0001f4c4 Indexed {char_count:,} characters from {filename}.\n"
                    f"Ask me anything about this file!"
                )

            # --- Images sent as documents ---
            elif ext in self._IMAGE_EXTENSIONS or mime.startswith("image/"):
                prompt = update.message.caption or "Describe this image in detail."
                try:
                    tg_file = await context.bot.get_file(doc.file_id)
                    buf = io.BytesIO()
                    await tg_file.download_to_memory(buf)
                    image_bytes = buf.getvalue()
                except Exception as e:
                    logger.error(f"[TelegramBot] Failed to download image document: {e}")
                    await update.message.reply_text(
                        "Couldn't download the image. Please try again."
                    )
                    return

                img_b64 = base64.b64encode(image_bytes).decode("utf-8")
                await self.send_typing_indicator(chat_id)
                response = await asyncio.to_thread(self._analyze_image_sync, img_b64, prompt)
                await update.message.reply_text(response)
                self._save_state()
                return

            # --- DOCX ---
            elif ext == ".docx" or mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                text, char_count = await self._extract_docx(doc)
                await self._ingest_document(user.id, filename, text, update)
                await update.message.reply_text(
                    f"\U0001f4c4 Indexed {char_count:,} characters from {filename}.\n"
                    f"Ask me anything about this document!"
                )

            # --- XLSX / CSV ---
            elif ext in (".xlsx", ".xls", ".csv") or mime in (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel", "text/csv",
            ):
                text, char_count = await self._extract_spreadsheet(doc, ext)
                await self._ingest_document(user.id, filename, text, update)
                await update.message.reply_text(
                    f"\U0001f4ca Indexed {char_count:,} characters from {filename}.\n"
                    f"Ask me anything about this data!"
                )

            # --- Unsupported ---
            else:
                await update.message.reply_text(
                    f"I can't read .{ext or '?'} files yet. "
                    f"I support PDFs, DOCX, XLSX, CSV, text/code files, and images."
                )

        except Exception as e:
            logger.error(f"Document handler error for {filename}: {e}", exc_info=True)
            await update.message.reply_text(
                f"Sorry, I couldn't process {filename}. "
                f"The file might be corrupted or in an unsupported format."
            )

    async def _extract_pdf(self, doc) -> "Tuple[str, int]":
        """Download a Telegram document and extract PDF text."""
        tg_file = await doc.get_file()
        file_bytes = await tg_file.download_as_bytearray()

        def _extract_with_fitz(raw: bytes) -> str:
            import fitz  # PyMuPDF
            text_parts = []
            with fitz.open(stream=raw, filetype="pdf") as pdf_doc:
                for page in pdf_doc:
                    text_parts.append(page.get_text())
            return "\n\n".join(text_parts)

        def _extract_with_pdfplumber(raw: bytes) -> str:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for page in pdf.pages:
                    text_parts.append(page.extract_text() or "")
            return "\n\n".join(text_parts)

        raw = bytes(file_bytes)
        text = ""

        for extractor, name in [
            (_extract_with_fitz, "PyMuPDF"),
            (_extract_with_pdfplumber, "pdfplumber"),
        ]:
            try:
                text = await asyncio.to_thread(extractor, raw)
                if text.strip():
                    logger.info(f"PDF extracted with {name}: {len(text)} chars")
                    break
            except ImportError:
                continue
            except Exception as e:
                logger.warning(f"PDF extraction with {name} failed: {e}")
                continue

        if not text.strip():
            raise ValueError(
                "Could not extract text -- the PDF may be image-based or empty."
            )

        if len(text) > 80_000:
            text = text[:80_000] + "\n\n[... truncated at 80,000 characters ...]"

        return text, len(text)

    async def _extract_text_file(self, doc) -> "Tuple[str, int]":
        """Download a Telegram document and read it as UTF-8 text."""
        tg_file = await doc.get_file()
        file_bytes = await tg_file.download_as_bytearray()

        try:
            text = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")

        if len(text) > 80_000:
            text = text[:80_000] + "\n\n[... truncated at 80,000 characters ...]"

        return text, len(text)

    async def _extract_docx(self, doc) -> "Tuple[str, int]":
        """Download a Telegram document and extract DOCX text."""
        tg_file = await doc.get_file()
        file_bytes = await tg_file.download_as_bytearray()

        def _extract(raw: bytes) -> str:
            try:
                from docx import Document as DocxDocument
                doc_obj = DocxDocument(io.BytesIO(raw))
                return "\n\n".join(p.text for p in doc_obj.paragraphs if p.text.strip())
            except ImportError:
                # Fallback: extract raw XML text
                import re as _re
                import zipfile
                with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                    xml = zf.read("word/document.xml").decode("utf-8")
                    return _re.sub(r"<[^>]+>", " ", xml).strip()

        text = await asyncio.to_thread(_extract, bytes(file_bytes))
        if len(text) > 80_000:
            text = text[:80_000] + "\n\n[... truncated at 80,000 characters ...]"
        return text, len(text)

    async def _extract_spreadsheet(self, doc, ext: str) -> "Tuple[str, int]":
        """Download and extract spreadsheet data as readable text."""
        tg_file = await doc.get_file()
        file_bytes = await tg_file.download_as_bytearray()

        def _extract(raw: bytes, file_ext: str) -> str:
            if file_ext == ".csv":
                import csv
                reader = csv.reader(io.StringIO(raw.decode("utf-8", errors="replace")))
                rows = list(reader)
                if not rows:
                    return "(empty CSV)"
                header = " | ".join(rows[0])
                body = "\n".join(" | ".join(r) for r in rows[1:200])
                return f"{header}\n{'—' * len(header)}\n{body}"
            else:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                    parts = []
                    for sheet_name in wb.sheetnames[:5]:
                        ws = wb[sheet_name]
                        lines = [f"## Sheet: {sheet_name}"]
                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                            if i > 200:
                                lines.append(f"... ({ws.max_row - 200} more rows)")
                                break
                            lines.append(" | ".join(str(c) if c is not None else "" for c in row))
                        parts.append("\n".join(lines))
                    wb.close()
                    return "\n\n".join(parts)
                except ImportError:
                    return "(openpyxl not installed — cannot read Excel files)"

        text = await asyncio.to_thread(_extract, bytes(file_bytes), ext)
        if len(text) > 80_000:
            text = text[:80_000] + "\n\n[... truncated at 80,000 characters ...]"
        return text, len(text)

    def _store_doc_context(self, user_id: int, text: str, filename: str):
        """Store extracted document text for subsequent Q&A (legacy 30-min fallback)."""
        self.store.set_doc_context(str(user_id), text, filename)
        logger.info(
            f"Stored doc context for user {user_id}: {filename} ({len(text)} chars)"
        )

    def _get_doc_index(self):
        """Lazy-load the per-bot DocumentIndex instance."""
        if not hasattr(self, "_doc_index_instance") or self._doc_index_instance is None:
            try:
                from aura.messaging.telegram.doc_rag import DocumentIndex
                self._doc_index_instance = DocumentIndex(self.store)
            except Exception as exc:
                logger.debug(f"[Media] DocumentIndex init failed: {exc}")
                self._doc_index_instance = None
        return self._doc_index_instance

    async def _ingest_document(
        self,
        user_id: int,
        filename: str,
        text: str,
        update: "Update",
    ) -> None:
        """Chunk + embed a document for RAG, push a live DocumentCard to the
        Mini App, and kick off async summarization.

        Also stores the flat text in legacy doc_context for the 30-min fallback
        path, so Telegram-only users (no Mini App) still get Q&A over the doc.
        """
        # Legacy path first — keep the existing inline Q&A flow working
        self._store_doc_context(user_id, text, filename)

        idx = self._get_doc_index()
        if idx is None:
            return

        from aura.messaging.telegram.doc_rag import (
            make_doc_id,
            summarize_document_sync,
        )

        uid = str(user_id)
        doc_id = make_doc_id(uid, filename, len(text))

        # Run chunking + embedding in a worker thread (can take 5-15s)
        try:
            index_result = await asyncio.to_thread(
                idx.index_document, uid, doc_id, filename, text
            )
        except Exception as exc:
            logger.warning(f"[Media] index_document failed: {exc}")
            return

        # Push an immediate tool card to the Mini App with "summarizing..."
        try:
            from api.services.websocket_hub import push_agent_event
            card_payload = {
                "tool_name": "document_index",
                "tool_args": {
                    "filename": filename,
                    "doc_id": doc_id,
                    "size_chars": index_result.get("size_chars", 0),
                    "chunks_count": index_result.get("chunks_count", 0),
                    "summary": "",
                    "facts": [],
                    "questions": [],
                },
                "tool_result": None,
            }
            push_agent_event(
                kind="tool_start",
                run_id=f"doc:{doc_id}",
                iteration=0,
                payload=card_payload,
                user_id=uid,
            )
        except Exception as exc:
            logger.debug(f"[Media] push_agent_event (tool_start) failed: {exc}")

        # Summarize in the background; update the Mini App card when done
        async def _finalize_card():
            brain = self._get_brain()
            if brain is None:
                return
            try:
                summary = await asyncio.to_thread(summarize_document_sync, brain, text)
            except Exception as exc:
                logger.debug(f"[Media] summarize failed: {exc}")
                summary = {"summary": "", "facts": [], "questions": []}

            # Persist so the user sees the same card on reload
            try:
                idx.set_summary(uid, doc_id, filename, {
                    **summary,
                    "chunks_count": index_result.get("chunks_count", 0),
                    "size_chars": index_result.get("size_chars", 0),
                })
            except Exception:
                pass

            try:
                from api.services.websocket_hub import push_agent_event
                push_agent_event(
                    kind="tool_result",
                    run_id=f"doc:{doc_id}",
                    iteration=0,
                    payload={
                        "tool_name": "document_index",
                        "tool_args": {
                            "filename": filename,
                            "doc_id": doc_id,
                            "size_chars": index_result.get("size_chars", 0),
                            "chunks_count": index_result.get("chunks_count", 0),
                        },
                        "tool_result": {
                            "filename": filename,
                            "doc_id": doc_id,
                            "chunks_count": index_result.get("chunks_count", 0),
                            "size_chars": index_result.get("size_chars", 0),
                            "summary": summary.get("summary", ""),
                            "facts": summary.get("facts", []),
                            "questions": summary.get("questions", []),
                        },
                    },
                    user_id=uid,
                )
            except Exception as exc:
                logger.debug(f"[Media] push_agent_event (tool_result) failed: {exc}")

        try:
            asyncio.create_task(_finalize_card())
        except Exception as exc:
            logger.debug(f"[Media] summary task spawn failed: {exc}")

    def _get_doc_context(self, user_id: int):
        """Get active document context, or None if expired/missing."""
        ctx = self.store.get_doc_context(str(user_id), ttl=self._DOC_CONTEXT_TTL)
        if not ctx:
            return None
        return ctx

    def _build_doc_augmented_text(self, user_id: int, text: str) -> str:
        """Inject relevant document chunks into the user's message.

        Prefers per-document RAG search: embeds the query, retrieves top-3
        chunks from the user's most recently indexed doc. Falls back to the
        legacy flat-text prepend if the RAG path returns nothing (e.g. Ollama
        unavailable, or the doc wasn't indexed yet).
        """
        uid = str(user_id)

        # Preferred path: RAG over the most recent indexed doc
        try:
            idx = self._get_doc_index()
            if idx is not None:
                docs = idx.list_user_docs(uid)
                if docs:
                    latest = docs[0]  # newest first
                    hits = idx.search(uid, text, doc_id=latest["doc_id"], k=3)
                    if hits:
                        blocks = "\n\n".join(
                            f"[chunk {h['chunk_idx']}] {h['chunk_text']}"
                            for h in hits
                        )
                        return (
                            f"[DOCUMENT CONTEXT -- file: {latest['filename']}, "
                            f"top {len(hits)} relevant chunks]\n"
                            f"{blocks}\n"
                            f"[END DOCUMENT CONTEXT]\n\n"
                            f"User question: {text}"
                        )
        except Exception as exc:
            logger.debug(f"[Media] RAG augmentation failed: {exc}")

        # Legacy fallback: 30-min TTL blob
        ctx = self._get_doc_context(user_id)
        if not ctx:
            return text

        doc_text = ctx["text"]
        if len(doc_text) > 30_000:
            doc_text = doc_text[:30_000] + "\n[... document truncated ...]"

        return (
            f"[DOCUMENT CONTEXT -- file: {ctx['filename']}]\n"
            f"{doc_text}\n"
            f"[END DOCUMENT CONTEXT]\n\n"
            f"User question: {text}"
        )

    async def _handle_voice(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle voice messages and audio files — transcribe and respond."""

        user = update.effective_user
        chat_id = str(update.effective_chat.id)

        if not self._is_user_allowed(user.id):
            return

        # Rate limit same as text
        max_per_min = self.config.get("max_messages_per_minute", 20)
        from aura.messaging.telegram.bot import _check_rate_limit
        if not _check_rate_limit(str(user.id), max_per_min):
            await update.message.reply_text(
                "You're sending messages too fast. Please wait a moment."
            )
            return

        # Get the voice or audio file object
        voice = update.message.voice
        audio = update.message.audio

        if voice:
            file_id = voice.file_id
            duration = voice.duration or 0
            suffix = ".ogg"
        elif audio:
            file_id = audio.file_id
            duration = audio.duration or 0
            mime = audio.mime_type or ""
            ext_map = {
                "audio/ogg": ".ogg", "audio/mpeg": ".mp3", "audio/mp3": ".mp3",
                "audio/wav": ".wav", "audio/x-wav": ".wav", "audio/flac": ".flac",
                "audio/mp4": ".m4a", "audio/x-m4a": ".m4a", "audio/webm": ".webm",
            }
            suffix = ext_map.get(mime, ".ogg")
        else:
            await update.message.reply_text(
                "Could not read the audio. Try sending as a voice message."
            )
            return

        # Reject very long audio (>5 minutes)
        if duration > 300:
            await update.message.reply_text(
                "That audio is too long (max 5 minutes for voice messages). "
                "Try sending a shorter clip."
            )
            return

        # Show typing while we process
        await self.send_typing_indicator(chat_id)

        # Download the file from Telegram
        tmp_path = None
        try:
            tg_file = await context.bot.get_file(file_id)
            buf = io.BytesIO()
            await tg_file.download_to_memory(buf)
            buf.seek(0)

            # Write to temp file for Whisper
            fd, tmp_path = tempfile.mkstemp(suffix=suffix)
            os.close(fd)
            with open(tmp_path, "wb") as f:
                f.write(buf.read())

            logger.info(f"[Voice] Downloaded {suffix} from user {user.id} ({duration}s)")

        except Exception as e:
            logger.error(f"[Voice] Failed to download voice file: {e}")
            await update.message.reply_text(
                "I couldn't download that voice message. Please try again."
            )
            return

        # Transcribe
        transcription = await self._transcribe_audio_file(tmp_path)

        # Clean up temp file
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)
        except OSError:
            pass

        if not transcription:
            await update.message.reply_text(
                "I couldn't transcribe that voice message. "
                "Transcription isn't available right now — try typing your message instead."
            )
            return

        # Send transcription as italic quote
        try:
            await update.message.reply_text(
                f"_You said: {transcription}_",
                parse_mode=ParseMode.MARKDOWN
            )
        except Exception:
            # Fallback without markdown if special chars break it
            await update.message.reply_text(f"You said: {transcription}")

        # Update active chat info
        from datetime import datetime
        if chat_id in self.active_chats:
            self.active_chats[chat_id]["last_message"] = datetime.now().isoformat()
        else:
            self.active_chats[chat_id] = {
                "user_id": str(user.id),
                "username": user.username,
                "first_name": user.first_name,
                "started_at": datetime.now().isoformat(),
                "last_message": datetime.now().isoformat()
            }

        # Show typing while AURA thinks
        await self.send_typing_indicator(chat_id)

        # Process transcribed text through AURA (same as if user typed it)
        from aura.messaging.base_platform import IncomingMessage, MessageType
        incoming = IncomingMessage(
            platform="telegram",
            user_id=str(user.id),
            chat_id=chat_id,
            username=user.username,
            display_name=user.first_name,
            message_type=MessageType.VOICE,
            text=transcription,
            media_url=None,
            timestamp=datetime.now(),
            raw_message=update.message
        )

        response = await self.handle_incoming(incoming)

        if response:
            await update.message.reply_text(response)

            # Send voice reply if TTS is enabled for this user
            if self.store.get_tts_enabled(str(user.id)):
                try:
                    wav_path = await self._voice_reply.synthesize(response)
                    if wav_path:
                        with open(wav_path, 'rb') as f:
                            await update.message.reply_voice(voice=f)
                        try:
                            os.unlink(wav_path)
                        except OSError:
                            pass
                    else:
                        logger.warning(f"[Voice] TTS returned None for user {user.id} — no TTS engine available")
                        await update.message.reply_text("(Voice reply unavailable — TTS engine not installed. Use /tts off to disable.)")
                except Exception as e:
                    logger.warning(f"[Voice] TTS reply failed: {e}")

        self._save_state()

    async def _transcribe_audio_file(self, file_path: str) -> Optional[str]:
        """Transcribe an audio file. Tries local Whisper, then API endpoint."""
        loop = asyncio.get_running_loop()

        # --- Method 1: Local Whisper via AudioTranscriberTool ---
        try:
            from aura.tools.audio_transcriber import AudioTranscriberTool
            transcriber = AudioTranscriberTool()

            result = await loop.run_in_executor(
                None, lambda: transcriber.transcribe(file_path)
            )

            if result.get("success") and result.get("text", "").strip():
                text = result["text"].strip()
                logger.info(f"[Voice] Whisper transcribed: {text[:80]}...")
                return text
            else:
                logger.warning(
                    f"[Voice] Whisper returned no text: {result.get('error', 'empty')}"
                )
        except ImportError:
            logger.info("[Voice] Whisper not installed, trying API fallback")
        except Exception as e:
            logger.warning(f"[Voice] Whisper transcription failed: {e}")

        # --- Method 2: Backend /api/transcribe endpoint ---
        try:
            import aiohttp

            api_url = os.environ.get("AURA_API_URL", "http://127.0.0.1:8000")
            api_key = os.environ.get("AURA_API_KEY", "")

            headers = {}
            if api_key:
                headers["x-api-key"] = api_key

            async with aiohttp.ClientSession() as session:
                with open(file_path, "rb") as f:
                    form = aiohttp.FormData()
                    form.add_field(
                        "file", f,
                        filename=os.path.basename(file_path),
                        content_type="audio/ogg"
                    )
                    async with session.post(
                        f"{api_url}/api/transcribe",
                        data=form,
                        headers=headers,
                        timeout=aiohttp.ClientTimeout(total=60)
                    ) as resp:
                        if resp.status == 200:
                            data = await resp.json()
                            text = data.get("text", "").strip()
                            if text:
                                logger.info(f"[Voice] API transcribed: {text[:80]}...")
                                return text
                        else:
                            body = await resp.text()
                            logger.warning(
                                f"[Voice] API transcribe returned {resp.status}: {body[:200]}"
                            )
        except ImportError:
            logger.info("[Voice] aiohttp not installed, skipping API fallback")
        except Exception as e:
            logger.warning(f"[Voice] API transcription failed: {e}")

        # All methods exhausted
        logger.warning("[Voice] All transcription methods failed")
        return None

    async def _handle_inline(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle inline queries from any chat via @Aura828Bot <query>."""
        import hashlib

        inline_query = update.inline_query
        if not inline_query:
            return

        user = inline_query.from_user
        query = (inline_query.query or "").strip()
        user_id = user.id
        import time as _time

        # Auth check
        if not self._is_user_allowed(user_id):
            await inline_query.answer([], cache_time=60)
            return

        # Ignore empty or too-short queries
        if len(query) < 3:
            await inline_query.answer([], cache_time=5)
            return

        # Debounce: skip if the user sent another query within the debounce window
        now = _time.time()
        prev = self._inline_last_query.get(user_id)
        self._inline_last_query[user_id] = (query, now)

        if prev and prev[0] != query:
            elapsed = now - prev[1]
            if elapsed < self._INLINE_DEBOUNCE_SEC:
                await asyncio.sleep(self._INLINE_DEBOUNCE_SEC - elapsed)
                current = self._inline_last_query.get(user_id)
                if current and current[0] != query:
                    return

        # Build result list
        results = []

        # Detect quick-action prefix
        query_lower = query.lower()
        detected_action = None
        action_body = query

        for prefix, (prompt_prefix, title_prefix) in self._INLINE_ACTIONS.items():
            if query_lower.startswith(prefix + " ") and len(query) > len(prefix) + 1:
                detected_action = prefix
                action_body = query[len(prefix):].strip()
                break

        if detected_action:
            prompt_prefix, title_prefix = self._INLINE_ACTIONS[detected_action]
            full_prompt = prompt_prefix + action_body

            response = await self._inline_generate(full_prompt, user_id)
            if response:
                result_id = hashlib.md5(
                    f"{detected_action}:{action_body}".encode()
                ).hexdigest()[:16]
                results.append(InlineQueryResultArticle(
                    id=result_id,
                    title=f"{title_prefix}: {action_body[:40]}",
                    description=response[:100],
                    input_message_content=InputTextMessageContent(
                        message_text=response
                    )
                ))

            # Also offer a general AURA answer as second option
            general_response = await self._inline_generate(query, user_id)
            if general_response and general_response != response:
                general_id = hashlib.md5(
                    f"general:{query}".encode()
                ).hexdigest()[:16]
                results.append(InlineQueryResultArticle(
                    id=general_id,
                    title=f"Ask AURA: {query[:40]}",
                    description=general_response[:100],
                    input_message_content=InputTextMessageContent(
                        message_text=general_response
                    )
                ))
        else:
            # General query
            response = await self._inline_generate(query, user_id)
            if response:
                result_id = hashlib.md5(query.encode()).hexdigest()[:16]
                results.append(InlineQueryResultArticle(
                    id=result_id,
                    title=response[:50],
                    description=query,
                    input_message_content=InputTextMessageContent(
                        message_text=response
                    )
                ))

        # Rich inline: add category-specific quick actions
        _INLINE_CATEGORIES = [
            ("code_", "\U0001f4bb Code", f"Write code for: {query}",
             f"Write clean, concise code for the following request:\n\n{query}"),
            ("img_", "\U0001f3a8 Image Prompt", f"Image prompt for: {query}",
             f"Create a detailed image generation prompt for: {query}"),
            ("research_", "\U0001f50d Research", f"Research: {query}",
             f"Give a thorough research summary on: {query}"),
        ]
        for cat_id, cat_title, cat_desc, cat_prompt in _INLINE_CATEGORIES:
            rid = hashlib.md5(f"{cat_id}{query}".encode()).hexdigest()[:16]
            results.append(InlineQueryResultArticle(
                id=rid,
                title=f"{cat_title}: {query[:35]}",
                description=cat_desc[:100],
                input_message_content=InputTextMessageContent(
                    message_text=f"/{cat_title.split(' ')[1].lower()} {query}"
                )
            ))

        # Fallback if nothing generated (keep category shortcuts)
        if len(results) <= len(_INLINE_CATEGORIES):
            results.insert(0, InlineQueryResultArticle(
                id="error",
                title="Could not generate a response",
                description="Try a different query or ask in the direct chat.",
                input_message_content=InputTextMessageContent(
                    message_text="Sorry, I couldn't process that query. Try asking me directly!"
                )
            ))

        try:
            await inline_query.answer(results, cache_time=30)
        except Exception as e:
            logger.error(f"Failed to answer inline query: {e}")

    async def _inline_generate(self, prompt: str, user_id: int) -> Optional[str]:
        """Generate a response for an inline query with caching and timeout."""
        import hashlib

        cache_key = hashlib.md5(prompt.encode()).hexdigest()
        cached = self._inline_cache.get(cache_key)
        if cached:
            return cached

        try:
            response = await asyncio.wait_for(
                self._process_with_aura(prompt, str(user_id)),
                timeout=self._INLINE_TIMEOUT_SEC
            )

            if response and response != "Sorry, I couldn't process that. Try again in a moment.":
                if len(response) > 4000:
                    response = response[:3997] + "..."
                self._inline_cache.put(cache_key, response)
                return response
        except asyncio.TimeoutError:
            logger.warning(f"Inline query timed out for user {user_id}")
        except Exception as e:
            logger.error(f"Inline generation error: {e}")

        return None
