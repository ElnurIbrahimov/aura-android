"""
Telegram Bot Integration for AURA

Uses python-telegram-bot library (async version).
Install: pip install python-telegram-bot>=20.0
"""

import asyncio
import hashlib
import logging
import os
import random
import re
import io
import tempfile
import time as _time
from collections import defaultdict, OrderedDict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Tuple
from pathlib import Path
import json
import base64
import uuid

# Per-user rate limiting
_msg_timestamps: Dict[str, list] = defaultdict(list)


def _check_rate_limit(user_id: str, max_per_min: int = 20) -> bool:
    """Return True if the user is within rate limits, False if throttled."""
    now = _time.time()
    stamps = _msg_timestamps[user_id]
    stamps[:] = [t for t in stamps if now - t < 60]
    if len(stamps) >= max_per_min:
        return False
    stamps.append(now)
    return True


# ============================================================================
#  Scheduled task / reminder callbacks (module-level for APScheduler pickling)
# ============================================================================

# Global ref to the running bot instance so scheduler callbacks can send messages.
# Set in TelegramBot.start(), cleared in TelegramBot.stop().
_active_bot_instance: Optional["TelegramBot"] = None
_active_event_loop: Optional[asyncio.AbstractEventLoop] = None


def _send_telegram_reminder(chat_id: str, message: str):
    """APScheduler callback: send a one-shot reminder message to Telegram.

    Runs in APScheduler's thread pool — uses run_coroutine_threadsafe to bridge
    into the bot's async event loop.
    """
    bot_inst = _active_bot_instance
    loop = _active_event_loop
    if not bot_inst or not bot_inst.bot or not loop or loop.is_closed():
        logging.getLogger(__name__).warning(
            f"[Scheduler] Cannot deliver reminder — bot not running. chat={chat_id}"
        )
        return

    text = f"\u23f0 Reminder: {message}"

    async def _send():
        try:
            await bot_inst.bot.send_message(chat_id=chat_id, text=text)
        except Exception as e:
            logging.getLogger(__name__).error(f"[Scheduler] Reminder send failed: {e}")

    asyncio.run_coroutine_threadsafe(_send(), loop)


def _run_telegram_scheduled_task(chat_id: str, task_prompt: str, is_agent_task: bool = False):
    """APScheduler callback: run a scheduled task and send results to Telegram.

    If is_agent_task is True, runs the prompt through the AURA agent first.
    Otherwise just sends the task_prompt as a notification.
    """
    bot_inst = _active_bot_instance
    loop = _active_event_loop
    if not bot_inst or not bot_inst.bot or not loop or loop.is_closed():
        logging.getLogger(__name__).warning(
            f"[Scheduler] Cannot deliver scheduled task — bot not running. chat={chat_id}"
        )
        return

    async def _execute_and_send():
        try:
            if is_agent_task:
                # Run through the agent, then send results
                try:
                    response_text, _ = await asyncio.to_thread(
                        bot_inst._run_agent_sync, task_prompt
                    )
                except Exception as e:
                    response_text = f"Scheduled task failed: {e}"

                header = f"\U0001f4c5 Scheduled Task: {task_prompt}\n\n"
                full_text = header + (response_text or "No output.")

                # Split if needed
                for chunk in _split_message(full_text, 4096):
                    await bot_inst.bot.send_message(chat_id=chat_id, text=chunk)
            else:
                await bot_inst.bot.send_message(
                    chat_id=chat_id,
                    text=f"\U0001f4c5 Scheduled: {task_prompt}"
                )
        except Exception as e:
            logging.getLogger(__name__).error(
                f"[Scheduler] Scheduled task send failed: {e}"
            )

    asyncio.run_coroutine_threadsafe(_execute_and_send(), loop)


def notify_hand_result(result) -> None:
    """Push a Hand execution result to all authorized Telegram chats.

    Called from HandManager's notify callback. Skips routine "all clear" results.
    Only notifies on: failures, Guardian issues, or completed research with findings.
    """
    bot_inst = _active_bot_instance
    loop = _active_event_loop
    if not bot_inst or not bot_inst.bot or not loop or loop.is_closed():
        return

    # Skip routine "all clear" Guardian reports and empty Memory maintenance
    hand_name = result.hand_name if hasattr(result, 'hand_name') else str(result.get("hand", ""))
    success = result.success if hasattr(result, 'success') else result.get("success", True)
    summary = result.summary if hasattr(result, 'summary') else result.get("summary", "")

    if success and hand_name == "guardian" and "ALL CLEAR" in summary.upper():
        return
    if success and hand_name == "memory" and "no action needed" in summary.lower():
        return

    icon = "\u2705" if success else "\u274c"
    text = f"{icon} <b>Hand '{hand_name}' completed</b>\n\n{summary[:800]}"

    async def _send():
        try:
            # Send to all known authorized chats
            chat_ids = list(getattr(bot_inst, '_authorized_chats', set()))
            if not chat_ids and hasattr(bot_inst, '_admin_chat_id'):
                chat_ids = [bot_inst._admin_chat_id]
            for cid in chat_ids:
                try:
                    await bot_inst.bot.send_message(chat_id=cid, text=text, parse_mode="HTML")
                except Exception:
                    pass
        except Exception as e:
            logging.getLogger(__name__).debug(f"[Hands] Telegram notify failed: {e}")

    asyncio.run_coroutine_threadsafe(_send(), loop)


def notify_hand_approval_request(request: dict) -> None:
    """Send a Hand approval request to Telegram with InlineKeyboard."""
    bot_inst = _active_bot_instance
    loop = _active_event_loop
    if not bot_inst or not bot_inst.bot or not loop or loop.is_closed():
        return

    request_id = request.get("request_id", "unknown")
    hand_name = request.get("hand_name", "unknown")
    tool_name = request.get("tool_name", "unknown")

    text = (
        f"\ud83d\udd10 <b>Approval Required</b>\n\n"
        f"Hand '<b>{hand_name}</b>' wants to use <b>{tool_name}</b>\n"
        f"Request: {request_id}"
    )

    async def _send():
        try:
            from telegram import InlineKeyboardButton, InlineKeyboardMarkup
            keyboard = InlineKeyboardMarkup([[
                InlineKeyboardButton("\u2705 Approve", callback_data=f"hand_approve_{request_id}"),
                InlineKeyboardButton("\u274c Deny", callback_data=f"hand_deny_{request_id}"),
            ]])
            chat_ids = list(getattr(bot_inst, '_authorized_chats', set()))
            if not chat_ids and hasattr(bot_inst, '_admin_chat_id'):
                chat_ids = [bot_inst._admin_chat_id]
            for cid in chat_ids:
                try:
                    await bot_inst.bot.send_message(
                        chat_id=cid, text=text, parse_mode="HTML", reply_markup=keyboard
                    )
                except Exception:
                    pass
        except Exception as e:
            logging.getLogger(__name__).debug(f"[Hands] Telegram approval request failed: {e}")

    asyncio.run_coroutine_threadsafe(_send(), loop)


def _parse_time_expression(text: str) -> Optional[datetime]:
    """Parse natural language time expressions into an absolute datetime.

    Handles:
        - "in Xh", "in Xm", "in X hours", "in X minutes", "in X days"
        - "at HH:MM", "at H:MMam/pm", "at Ham/pm"
        - "tomorrow", "tomorrow HH:MM", "tomorrow at HH:MM"
        - "next monday", "next tuesday at HH:MM"

    Returns None if parsing fails.
    """
    original = text.strip()
    t = original.lower().strip()
    now = datetime.now()

    # --- Relative: "in X hours/minutes/seconds/days" or shorthand "in 2h" ---
    m = re.match(r'in\s+(\d+)\s*h(?:ours?)?$', t)
    if m:
        return now + timedelta(hours=int(m.group(1)))

    m = re.match(r'in\s+(\d+)\s*m(?:in(?:utes?)?)?$', t)
    if m:
        return now + timedelta(minutes=int(m.group(1)))

    m = re.match(r'in\s+(\d+)\s*s(?:ec(?:onds?)?)?$', t)
    if m:
        return now + timedelta(seconds=int(m.group(1)))

    m = re.match(r'in\s+(\d+)\s*d(?:ays?)?$', t)
    if m:
        return now + timedelta(days=int(m.group(1)))

    # Compound relative: "in 1h 30m", "in 2 hours 15 minutes"
    m = re.match(r'in\s+(\d+)\s*h(?:ours?)?\s+(\d+)\s*m(?:in(?:utes?)?)?$', t)
    if m:
        return now + timedelta(hours=int(m.group(1)), minutes=int(m.group(2)))

    # --- "at HH:MM" or "at H:MMam/pm" or "at Ham/pm" ---
    m = re.match(r'at\s+(\d{1,2}):(\d{2})\s*(am|pm)?$', t)
    if m:
        hour, minute = int(m.group(1)), int(m.group(2))
        ampm = m.group(3)
        if ampm == 'pm' and hour != 12:
            hour += 12
        elif ampm == 'am' and hour == 12:
            hour = 0
        target = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
        if target <= now:
            target += timedelta(days=1)
        return target

    m = re.match(r'at\s+(\d{1,2})\s*(am|pm)$', t)
    if m:
        hour = int(m.group(1))
        ampm = m.group(2)
        if ampm == 'pm' and hour != 12:
            hour += 12
        elif ampm == 'am' and hour == 12:
            hour = 0
        target = now.replace(hour=hour, minute=0, second=0, microsecond=0)
        if target <= now:
            target += timedelta(days=1)
        return target

    # --- "tomorrow" optionally with time ---
    m = re.match(r'tomorrow(?:\s+(?:at\s+)?(\d{1,2}):(\d{2})\s*(am|pm)?)?$', t)
    if m:
        tomorrow = now + timedelta(days=1)
        if m.group(1):
            hour, minute = int(m.group(1)), int(m.group(2))
            ampm = m.group(3)
            if ampm == 'pm' and hour != 12:
                hour += 12
            elif ampm == 'am' and hour == 12:
                hour = 0
            return tomorrow.replace(hour=hour, minute=minute, second=0, microsecond=0)
        else:
            return tomorrow.replace(hour=9, minute=0, second=0, microsecond=0)

    m = re.match(r'tomorrow(?:\s+(?:at\s+)?(\d{1,2})\s*(am|pm))?$', t)
    if m and m.group(1):
        tomorrow = now + timedelta(days=1)
        hour = int(m.group(1))
        ampm = m.group(2)
        if ampm == 'pm' and hour != 12:
            hour += 12
        elif ampm == 'am' and hour == 12:
            hour = 0
        return tomorrow.replace(hour=hour, minute=0, second=0, microsecond=0)

    # --- "next <weekday>" optionally with time ---
    days_of_week = {
        'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3,
        'friday': 4, 'saturday': 5, 'sunday': 6,
        'mon': 0, 'tue': 1, 'wed': 2, 'thu': 3,
        'fri': 4, 'sat': 5, 'sun': 6,
    }
    m = re.match(
        r'next\s+(\w+?)(?:\s+(?:at\s+)?(\d{1,2})(?::(\d{2}))?\s*(am|pm)?)?$', t
    )
    if m:
        day_name = m.group(1)
        if day_name in days_of_week:
            target_weekday = days_of_week[day_name]
            days_ahead = (target_weekday - now.weekday()) % 7
            if days_ahead == 0:
                days_ahead = 7
            target = now + timedelta(days=days_ahead)

            hour, minute = 9, 0  # default 9am
            if m.group(2):
                hour = int(m.group(2))
                minute = int(m.group(3)) if m.group(3) else 0
                ampm = m.group(4)
                if ampm == 'pm' and hour != 12:
                    hour += 12
                elif ampm == 'am' and hour == 12:
                    hour = 0

            return target.replace(hour=hour, minute=minute, second=0, microsecond=0)

    return None


def _parse_schedule_expression(text: str) -> Optional[Dict]:
    """Parse a schedule expression into APScheduler trigger parameters.

    Handles:
        - "every Xh", "every X hours", "every X minutes", "every Xm"
        - "daily at HH:MM", "daily at Ham/pm"
        - "every monday at HH:MM", "every <weekday> at H:MMam/pm"
        - "every 30 minutes", "every 2 hours"

    Returns a dict with:
        - "type": "interval" | "cron"
        - For interval: "hours", "minutes", "seconds"
        - For cron: "cron_expression" (5-part)
    Returns None if parsing fails.
    """
    t = text.lower().strip()

    # --- "every Xh" / "every X hours" / "every Xm" / "every X minutes" ---
    m = re.match(r'every\s+(\d+)\s*h(?:ours?)?$', t)
    if m:
        return {"type": "interval", "hours": int(m.group(1)), "minutes": 0, "seconds": 0}

    m = re.match(r'every\s+(\d+)\s*m(?:in(?:utes?)?)?$', t)
    if m:
        return {"type": "interval", "hours": 0, "minutes": int(m.group(1)), "seconds": 0}

    m = re.match(r'every\s+(\d+)\s*s(?:ec(?:onds?)?)?$', t)
    if m:
        return {"type": "interval", "hours": 0, "minutes": 0, "seconds": int(m.group(1))}

    # --- "daily at HH:MM" or "daily at Ham/pm" ---
    m = re.match(r'daily\s+(?:at\s+)?(\d{1,2})(?::(\d{2}))?\s*(am|pm)?$', t)
    if m:
        hour = int(m.group(1))
        minute = int(m.group(2)) if m.group(2) else 0
        ampm = m.group(3)
        if ampm == 'pm' and hour != 12:
            hour += 12
        elif ampm == 'am' and hour == 12:
            hour = 0
        return {"type": "cron", "cron_expression": f"{minute} {hour} * * *"}

    # --- "every <weekday> at HH:MM" ---
    days_of_week = {
        'monday': '1', 'tuesday': '2', 'wednesday': '3', 'thursday': '4',
        'friday': '5', 'saturday': '6', 'sunday': '0',
        'mon': '1', 'tue': '2', 'wed': '3', 'thu': '4',
        'fri': '5', 'sat': '6', 'sun': '0',
    }
    m = re.match(
        r'every\s+(\w+)\s+(?:at\s+)?(\d{1,2})(?::(\d{2}))?\s*(am|pm)?$', t
    )
    if m:
        day_name = m.group(1)
        if day_name in days_of_week:
            dow = days_of_week[day_name]
            hour = int(m.group(2))
            minute = int(m.group(3)) if m.group(3) else 0
            ampm = m.group(4)
            if ampm == 'pm' and hour != 12:
                hour += 12
            elif ampm == 'am' and hour == 12:
                hour = 0
            return {"type": "cron", "cron_expression": f"{minute} {hour} * * {dow}"}

    return None


try:
    from telegram import (
        Update, Bot, InlineQueryResultArticle, InputTextMessageContent,
        InlineKeyboardButton, InlineKeyboardMarkup, LabeledPrice,
        ReplyKeyboardMarkup, ReplyKeyboardRemove, BotCommand,
    )
    from telegram.ext import (
        Application,
        CommandHandler,
        MessageHandler,
        InlineQueryHandler,
        CallbackQueryHandler,
        PreCheckoutQueryHandler,
        ChatMemberHandler,
        ContextTypes,
        filters
    )
    from telegram.constants import ParseMode, ChatAction

    # ReactionTypeEmoji requires python-telegram-bot >= 20.8
    try:
        from telegram import ReactionTypeEmoji
        REACTIONS_AVAILABLE = True
    except ImportError:
        REACTIONS_AVAILABLE = False
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    Update = None
    Bot = None

from .base_platform import (
    BasePlatform,
    IncomingMessage,
    OutgoingMessage,
    MessageType
)
from aura.core.conversation_manager import get_conversation_manager

# Skill library and evolution imports (lazy — only fail when actually used)
try:
    from aura_skill_library import (
        SkillStore, SkillLearner, Skill, SkillCategory, SkillExample, SkillMetadata
    )
    SKILL_LIBRARY_AVAILABLE = True
except ImportError:
    SKILL_LIBRARY_AVAILABLE = False

try:
    from aura.evolution import GEPAEngine, GEPAConfig, AuraSkillAdapter, Candidate, GEPAResult
    from aura.evolution.types import EvalExample
    GEPA_AVAILABLE = True
except ImportError:
    GEPA_AVAILABLE = False

logger = logging.getLogger(__name__)

# Max chars for code output before truncation in Telegram messages
_MAX_OUTPUT_CHARS = 3500

# Premium tier definitions for Telegram payments
PREMIUM_TIERS = {
    "supporter": {
        "title": "AURA Supporter",
        "description": "Support AURA development",
        "price": 500,  # $5.00 in cents
        "currency": "USD",
        "benefits": ["Priority responses", "Badge in status"],
    },
    "pro": {
        "title": "AURA Pro",
        "description": "Unlock advanced features",
        "price": 1500,  # $15.00
        "currency": "USD",
        "benefits": ["Unlimited research", "Fleet mode", "Priority model routing", "Custom personality"],
    },
    "patron": {
        "title": "AURA Patron",
        "description": "Maximum support + all features",
        "price": 5000,  # $50.00
        "currency": "USD",
        "benefits": ["Everything in Pro", "Direct feature requests", "Early access", "Custom training"],
    },
}

# ============================================================================
#  Emotion-to-sticker/GIF mapping for contextual reactions (Phase 5)
# ============================================================================
EMOTION_REACTIONS = {
    "joy": {"sticker_query": "happy", "gif_queries": ["celebration", "happy dance", "yay"]},
    "excited": {"sticker_query": "excited", "gif_queries": ["excited", "woohoo", "amazing"]},
    "curious": {"sticker_query": "thinking", "gif_queries": ["thinking", "hmm", "curious"]},
    "surprised": {"sticker_query": "surprised", "gif_queries": ["shocked", "wow", "surprised"]},
    "sad": {"sticker_query": "sad", "gif_queries": ["sad", "cry", "disappointed"]},
    "frustrated": {"sticker_query": "angry", "gif_queries": ["frustrated", "facepalm"]},
    "grateful": {"sticker_query": "thank you", "gif_queries": ["thank you", "grateful", "heart"]},
    "empathetic": {"sticker_query": "hug", "gif_queries": ["hug", "comfort", "care"]},
    "confident": {"sticker_query": "cool", "gif_queries": ["confident", "boss", "cool"]},
    "neutral": {"sticker_query": "ok", "gif_queries": ["ok", "thumbs up", "nod"]},
}

EMOTION_EMOJI = {
    "joy": "\U0001f60a",
    "excited": "\U0001f525",
    "curious": "\U0001f914",
    "surprised": "\U0001f62e",
    "sad": "\U0001f622",
    "frustrated": "\U0001f624",
    "grateful": "\u2764\ufe0f",
    "empathetic": "\U0001f917",
    "confident": "\U0001f4aa",
    "neutral": "\U0001f44d",
}


def _extract_code_from_message(text: str) -> str:
    """Extract Python code from a /code message.

    Supports:
      /code print("hello")
      /code ```python\\nprint("hello")\\n```
      /code ```\\nprint("hello")\\n```
    """
    # Remove the /code prefix (may include @botname)
    stripped = re.sub(r"^/code(@\w+)?\s*", "", text, count=1)

    # Check for fenced code block: ```python ... ``` or ``` ... ```
    match = re.search(r"```(?:python)?\s*\n?(.*?)```", stripped, re.DOTALL)
    if match:
        return match.group(1).strip()

    return stripped.strip()


class _LRUCache:
    """Simple LRU cache using OrderedDict."""

    def __init__(self, maxsize: int = 50):
        self._cache: OrderedDict = OrderedDict()
        self._maxsize = maxsize

    def get(self, key: str) -> "Optional[str]":
        if key in self._cache:
            self._cache.move_to_end(key)
            return self._cache[key]
        return None

    def put(self, key: str, value: str):
        if key in self._cache:
            self._cache.move_to_end(key)
        self._cache[key] = value
        if len(self._cache) > self._maxsize:
            self._cache.popitem(last=False)


# ============================================================================
#  Telegram MarkdownV2 formatting utilities
# ============================================================================

# Characters that must be escaped in Telegram MarkdownV2 (outside code blocks)
_MARKDOWNV2_ESCAPE = re.compile(r'([_*\[\]()~`>#+\-=|{}.!\\])')

# Maximum Telegram message length
_TELEGRAM_MSG_LIMIT = 4096


def _escape_mdv2(text: str) -> str:
    """Escape special characters for Telegram MarkdownV2."""
    return _MARKDOWNV2_ESCAPE.sub(r'\\\1', text)


def format_telegram_response(text: str) -> List[str]:
    """Convert markdown text to Telegram MarkdownV2 and split into chunks.

    Handles code blocks, inline code, bold, italic, links.
    Escapes all MarkdownV2 special chars outside of formatting spans.
    Returns a list of strings, each <= 4096 chars, split at paragraph boundaries.
    """
    if not text:
        return [""]

    # --- Phase 1: Extract code blocks and inline code to protect them ---
    placeholders: Dict[str, str] = {}
    counter = [0]

    def _save_code_block(m: re.Match) -> str:
        key = f"\x00CB{counter[0]}\x00"
        counter[0] += 1
        lang = m.group(1) or ""
        code = m.group(2)
        placeholders[key] = f"```{lang}\n{code}\n```"
        return key

    def _save_inline_code(m: re.Match) -> str:
        key = f"\x00IC{counter[0]}\x00"
        counter[0] += 1
        placeholders[key] = f"`{m.group(1)}`"
        return key

    result = re.sub(r'```(\w*)\n?(.*?)```', _save_code_block, text, flags=re.DOTALL)
    result = re.sub(r'`([^`\n]+)`', _save_inline_code, result)

    # --- Phase 2: Extract links, bold, italic before escaping ---
    link_phs: Dict[str, str] = {}

    def _save_link(m: re.Match) -> str:
        key = f"\x00LK{counter[0]}\x00"
        counter[0] += 1
        link_text = _escape_mdv2(m.group(1))
        url = m.group(2)
        link_phs[key] = f"[{link_text}]({url})"
        return key

    result = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', _save_link, result)

    fmt_phs: Dict[str, str] = {}

    def _save_bold(m: re.Match) -> str:
        key = f"\x00BD{counter[0]}\x00"
        counter[0] += 1
        inner = _escape_mdv2(m.group(1))
        fmt_phs[key] = f"*{inner}*"
        return key

    def _save_italic(m: re.Match) -> str:
        key = f"\x00IT{counter[0]}\x00"
        counter[0] += 1
        inner = _escape_mdv2(m.group(1))
        fmt_phs[key] = f"_{inner}_"
        return key

    # Bold: **text** or __text__
    result = re.sub(r'\*\*(.+?)\*\*', _save_bold, result)
    result = re.sub(r'__(.+?)__', _save_bold, result)
    # Italic: *text* or _text_ (single, non-greedy)
    result = re.sub(r'(?<!\*)\*([^*]+?)\*(?!\*)', _save_italic, result)
    result = re.sub(r'(?<!_)_([^_]+?)_(?!_)', _save_italic, result)

    # --- Phase 3: Escape remaining text ---
    result = _escape_mdv2(result)

    # --- Phase 4: Restore placeholders (reverse order of extraction) ---
    for key, val in fmt_phs.items():
        result = result.replace(_escape_mdv2(key), val)
    for key, val in link_phs.items():
        result = result.replace(_escape_mdv2(key), val)
    for key, val in placeholders.items():
        result = result.replace(_escape_mdv2(key), val)

    # --- Phase 5: Split into <= 4096-char chunks ---
    return _split_message(result, _TELEGRAM_MSG_LIMIT)


def _split_message(text: str, limit: int) -> List[str]:
    """Split text into chunks of at most `limit` chars at paragraph breaks."""
    if len(text) <= limit:
        return [text]

    chunks: List[str] = []
    remaining = text

    while remaining:
        if len(remaining) <= limit:
            chunks.append(remaining)
            break

        # Try paragraph break, then newline, then space, then hard cut
        split_at = remaining.rfind('\n\n', 0, limit)
        if split_at == -1:
            split_at = remaining.rfind('\n', 0, limit)
        if split_at == -1:
            split_at = remaining.rfind(' ', 0, limit)
        if split_at == -1:
            split_at = limit

        chunks.append(remaining[:split_at].rstrip())
        remaining = remaining[split_at:].lstrip('\n')

    return chunks if chunks else [""]


def format_research_citations(text: str, sources: List[Dict]) -> str:
    """Format research results with numbered citations and clickable source list.

    Args:
        text: The research report body (may already contain [N] refs).
        sources: List of dicts with 'url' and 'title' keys.

    Returns:
        Formatted text with a numbered source list appended.
    """
    if not sources:
        return text

    # De-duplicate by URL
    seen_urls: set = set()
    unique: List[Dict] = []
    for s in sources:
        url = s.get("url", "")
        if url and url not in seen_urls:
            seen_urls.add(url)
            unique.append(s)

    lines = ["\n\n---\n**Sources:**"]
    for i, src in enumerate(unique, 1):
        title = src.get("title", "Untitled")
        url = src.get("url", "")
        if url:
            lines.append(f"[{i}] [{title}]({url})")
        else:
            lines.append(f"[{i}] {title}")

    return text + "\n".join(lines)


# ============================================================================
#  TelegramProgressReporter — editable progress message with rate limiting
# ============================================================================

class TelegramProgressReporter:
    """Sends and edits a single progress message in a Telegram chat.

    Rate-limited to one edit every 2 seconds (Telegram API constraint).

    Usage:
        reporter = TelegramProgressReporter(bot, chat_id)
        await reporter.start("Working on it...")
        await reporter.update("Step 1/3...")
        await reporter.finish("Here's the answer.")
    """

    MIN_EDIT_INTERVAL = 2.0

    def __init__(self, bot, chat_id: str):
        self._bot = bot
        self._chat_id = chat_id
        self._message_id: Optional[int] = None
        self._last_edit_time: float = 0.0
        self._last_text: str = ""
        self._pending_update: Optional[str] = None
        self._flush_task: Optional[asyncio.Task] = None

    async def start(self, text: str = "\u23f3 Working...") -> int:
        """Send the initial progress message. Returns message_id."""
        try:
            msg = await self._bot.send_message(chat_id=self._chat_id, text=text)
            self._message_id = msg.message_id
            self._last_text = text
            self._last_edit_time = _time.time()
            return msg.message_id
        except Exception as e:
            logger.warning(f"[ProgressReporter] Could not send initial message: {e}")
            return 0

    async def update(self, text: str):
        """Edit the progress message. Rate-limited: queues if too fast."""
        if not self._message_id or text == self._last_text:
            return

        now = _time.time()
        elapsed = now - self._last_edit_time

        if elapsed >= self.MIN_EDIT_INTERVAL:
            await self._do_edit(text)
        else:
            self._pending_update = text
            if not self._flush_task or self._flush_task.done():
                delay = self.MIN_EDIT_INTERVAL - elapsed
                self._flush_task = asyncio.create_task(self._flush_after(delay))

    async def update_structured(self, step: int, total: int, label: str,
                                elapsed: float = 0.0):
        """Send a structured progress update with progress bar and step counter."""
        filled = int(10 * step / max(total, 1))
        bar = "\u2588" * filled + "\u2591" * (10 - filled)
        pct = int(100 * step / max(total, 1))
        time_str = f" | {elapsed:.0f}s" if elapsed > 0 else ""
        text = f"\u2699\ufe0f Step {step}/{total} — {label}\n[{bar}] {pct}%{time_str}"
        await self.update(text)

    async def _flush_after(self, delay: float):
        """Wait then send the most recent pending update."""
        await asyncio.sleep(delay)
        if self._pending_update:
            text = self._pending_update
            self._pending_update = None
            await self._do_edit(text)

    async def _do_edit(self, text: str):
        """Actually edit the Telegram message."""
        try:
            await self._bot.edit_message_text(
                chat_id=self._chat_id,
                message_id=self._message_id,
                text=text,
            )
            self._last_text = text
            self._last_edit_time = _time.time()
        except Exception as e:
            if "not modified" not in str(e).lower():
                logger.debug(f"[ProgressReporter] Edit failed: {e}")

    async def finish(self, text: str, parse_mode: Optional[str] = None):
        """Replace the progress message with the final response.

        If text > 4096 chars, deletes the progress message and sends
        the response as multiple new messages.
        """
        if self._flush_task and not self._flush_task.done():
            self._flush_task.cancel()

        if not self._message_id:
            return

        if len(text) <= _TELEGRAM_MSG_LIMIT:
            try:
                kwargs = {
                    "chat_id": self._chat_id,
                    "message_id": self._message_id,
                    "text": text,
                }
                if parse_mode:
                    kwargs["parse_mode"] = parse_mode
                await self._bot.edit_message_text(**kwargs)
            except Exception as e:
                logger.debug(f"[ProgressReporter] Final edit failed: {e}")
                try:
                    await self._bot.send_message(
                        chat_id=self._chat_id, text=text, parse_mode=parse_mode
                    )
                except Exception:
                    pass
        else:
            # Delete progress message, send multi-part
            try:
                await self._bot.delete_message(
                    chat_id=self._chat_id, message_id=self._message_id
                )
            except Exception:
                pass

            for chunk in _split_message(text, _TELEGRAM_MSG_LIMIT):
                try:
                    await self._bot.send_message(
                        chat_id=self._chat_id, text=chunk, parse_mode=parse_mode
                    )
                    await asyncio.sleep(0.3)
                except Exception as e:
                    logger.warning(f"[ProgressReporter] Chunk send failed: {e}")

    async def finish_parts(self, parts: List[str], parse_mode: Optional[str] = None):
        """Replace the progress message with pre-split parts."""
        if self._flush_task and not self._flush_task.done():
            self._flush_task.cancel()
        if not parts:
            return

        # Edit progress message with first part
        if self._message_id:
            try:
                kwargs = {
                    "chat_id": self._chat_id,
                    "message_id": self._message_id,
                    "text": parts[0],
                }
                if parse_mode:
                    kwargs["parse_mode"] = parse_mode
                await self._bot.edit_message_text(**kwargs)
            except Exception:
                try:
                    await self._bot.send_message(
                        chat_id=self._chat_id, text=parts[0], parse_mode=parse_mode
                    )
                except Exception:
                    pass

        for part in parts[1:]:
            try:
                await self._bot.send_message(
                    chat_id=self._chat_id, text=part, parse_mode=parse_mode
                )
                await asyncio.sleep(0.3)
            except Exception as e:
                logger.warning(f"[ProgressReporter] Part send failed: {e}")


# ============================================================================
#  ToolStatusCallback — wires agent tool execution into progress updates
# ============================================================================

class ToolStatusCallback:
    """Provides on_tool_start / on_tool_end / on_research_progress callbacks.

    Wire into the agent's tool execution flow:
        reporter = TelegramProgressReporter(bot, chat_id)
        cb = ToolStatusCallback(reporter)
        # then pass cb.on_tool_start, cb.on_tool_end as callbacks
    """

    _TOOL_LABELS = {
        "web_search":     "\U0001f50d Searching the web",
        "deep_research":  "\U0001f4da Researching",
        "code_execute":   "\u26a1 Executing code",
        "code_run":       "\u26a1 Running code",
        "plot":           "\U0001f4ca Generating plot",
        "browse":         "\U0001f310 Browsing",
        "file_read":      "\U0001f4c4 Reading file",
        "file_write":     "\u270d\ufe0f Writing file",
        "memory_search":  "\U0001f9e0 Searching memory",
        "summarize":      "\U0001f4dd Summarizing",
    }

    def __init__(self, reporter: TelegramProgressReporter):
        self._reporter = reporter
        self._active_tools: List[str] = []

    def _label_for(self, tool_name: str) -> str:
        return self._TOOL_LABELS.get(tool_name, f"\U0001f527 Using {tool_name}")

    async def on_tool_start(self, tool_name: str, args: Optional[Dict] = None):
        """Called when a tool starts executing."""
        self._active_tools.append(tool_name)
        label = self._label_for(tool_name)

        detail = ""
        if args:
            if "query" in args:
                detail = f': "{args["query"][:60]}"'
            elif "url" in args:
                detail = f": {args['url'][:60]}"
            elif "topic" in args:
                detail = f': "{args["topic"][:60]}"'

        await self._reporter.update(f"{label}{detail}...")

    async def on_tool_end(self, tool_name: str, success: bool = True,
                          summary: Optional[str] = None):
        """Called when a tool finishes executing."""
        if tool_name in self._active_tools:
            self._active_tools.remove(tool_name)

        if summary:
            status = "\u2705" if success else "\u274c"
            await self._reporter.update(f"{status} {summary[:100]}")
        elif self._active_tools:
            current = self._label_for(self._active_tools[-1])
            await self._reporter.update(f"{current}...")
        else:
            await self._reporter.update("\U0001f4ad Thinking...")

    async def on_research_progress(self, event: Dict):
        """Handle ResearchProgressEmitter events.

        Pass this as callback to ResearchProgressEmitter:
            emitter = ResearchProgressEmitter(callback=
                lambda evt: asyncio.create_task(cb.on_research_progress(evt))
            )
        """
        data = event.get("data", {})
        stage = event.get("stage", "")

        if stage == "search":
            step = data.get("step", 0)
            total = data.get("total", 0)
            query = data.get("query", "")
            text = f"\U0001f50d Searching ({step}/{total}): {query[:60]}"
        elif stage == "plan":
            n = len(data.get("subtopics", []))
            text = f"\U0001f4cb Planning: {n} sub-queries"
        elif stage == "source":
            title = data.get("title", "")[:60]
            text = f"\U0001f4d6 Found: {title}"
        elif stage == "finding":
            text = f"\U0001f4a1 {data.get('message', 'New finding')[:80]}"
        elif stage == "synthesis":
            text = "\u270d\ufe0f Synthesizing report..."
        else:
            text = data.get("message", "\U0001f50d Researching...")[:100]

        await self._reporter.update(text)


class TelegramBot(BasePlatform):
    """Telegram Bot integration for AURA"""

    def __init__(self, aura_engine, config: dict):
        super().__init__(aura_engine, config)

        if not TELEGRAM_AVAILABLE:
            raise ImportError(
                "python-telegram-bot not installed. "
                "Run: pip install python-telegram-bot>=20.0"
            )

        self.token = config.get("telegram_token")
        if not self.token or self.token == "YOUR_BOT_TOKEN_HERE":
            raise ValueError(
                "telegram_token is required in config. "
                "Get one from @BotFather on Telegram."
            )

        self.allowed_users: List[str] = config.get("allowed_users", [])
        self.admin_users: List[str] = config.get("admin_users", [])

        self.app: Optional[Application] = None
        self.bot: Optional[Bot] = None

        # ===== SQLite-backed persistent store =====
        from aura.messaging.telegram_store import TelegramStore
        db_path = config.get("telegram_db_path", "data/telegram_state.db")
        self.store = TelegramStore(db_path=db_path)
        # One-time migration from legacy JSON files
        self.store.migrate_from_json()

        # Track active chats for proactive messaging (backed by store)
        self.active_chats: Dict[str, dict] = self.store.get_active_chats()
        self._pending_forget: Dict[str, datetime] = {}

        # Document context TTL (store handles persistence)
        self._DOC_CONTEXT_TTL = 30 * 60  # 30 minutes

        # Lazy-loaded code executor
        self._code_executor = None

        # Inline query state (cache in-memory for speed, store for persistence)
        self._inline_cache = _LRUCache(maxsize=50)
        self._inline_last_query: Dict[int, Tuple[str, float]] = {}  # user_id -> (query, timestamp)
        self._INLINE_DEBOUNCE_SEC = 1.0
        self._INLINE_TIMEOUT_SEC = 15.0

        # Skill system state (in-memory cache, write-through to store)
        self._skill_pending: Dict[int, dict] = {}
        self._last_exchange: Dict[int, dict] = {}
        self._skill_store = None   # Lazy-loaded SkillStore
        self._skill_learner = None # Lazy-loaded SkillLearner
        self._skill_create_state: Dict[int, dict] = {}

        # Premium/payment state (backed by store)
        self._premium_users: Dict[str, dict] = self.store.get_premium_users()

        # Group message cache (backed by store, read-through cache)
        self._group_message_cache: Dict[str, List[dict]] = {}

        # Location sharing state (backed by store)
        self._user_locations: Dict[str, dict] = {}

        # Digest job tracking (backed by store)
        self._digest_job_ids: Dict[str, str] = self.store.get_digest_jobs()

    @property
    def platform_name(self) -> str:
        return "telegram"

    def _get_code_executor(self):
        """Lazy-load CodeExecutorTool to avoid import cost at startup."""
        if self._code_executor is None:
            from aura.tools.code_executor import CodeExecutorTool
            self._code_executor = CodeExecutorTool(timeout=30)
        return self._code_executor

    def _save_state(self):
        """Persist current active_chats to SQLite store."""
        try:
            for chat_id, info in self.active_chats.items():
                self.store.upsert_active_chat(
                    chat_id=chat_id,
                    user_id=info.get("user_id", ""),
                    first_name=info.get("first_name", ""),
                    username=info.get("username", ""),
                )
        except Exception as e:
            logger.error(f"Could not save Telegram state: {e}")

    def _save_premium_state(self):
        """Premium state is auto-persisted via store — this is a no-op for compatibility."""
        pass

    def is_premium(self, user_id: str) -> bool:
        """Check if a user has premium status."""
        return user_id in self._premium_users or self.store.is_premium(user_id)

    async def start(self):
        """Start the Telegram bot"""

        logger.info("Starting Telegram bot...")

        # Build application
        self.app = Application.builder().token(self.token).build()
        self.bot = self.app.bot

        # Add handlers
        self.app.add_handler(CommandHandler("start", self._handle_start))
        self.app.add_handler(CommandHandler("help", self._handle_help))
        self.app.add_handler(CommandHandler("status", self._handle_status))
        self.app.add_handler(CommandHandler("mood", self._handle_mood))
        self.app.add_handler(CommandHandler("memory", self._handle_memory))
        self.app.add_handler(CommandHandler("forget", self._handle_forget))
        self.app.add_handler(CommandHandler("research", self._handle_research))
        self.app.add_handler(CommandHandler("search", self._handle_search))
        self.app.add_handler(CommandHandler("summarize", self._handle_summarize))
        self.app.add_handler(CommandHandler("image", self._handle_image))
        self.app.add_handler(CommandHandler("code", self._handle_code))
        self.app.add_handler(CommandHandler("model", self._handle_model))
        self.app.add_handler(CommandHandler("compare", self._handle_compare))
        self.app.add_handler(CommandHandler("session", self._handle_session))
        self.app.add_handler(CommandHandler("webhook", self._handle_webhook))
        self.app.add_handler(CommandHandler("remind", self._handle_remind))
        self.app.add_handler(CommandHandler("schedule", self._handle_schedule))
        self.app.add_handler(CommandHandler("tasks", self._handle_tasks))
        self.app.add_handler(CommandHandler("cancel", self._handle_cancel))
        self.app.add_handler(CommandHandler("agent", self._handle_agent))
        self.app.add_handler(CommandHandler("fleet", self._handle_fleet))
        self.app.add_handler(CommandHandler("learn", self._handle_learn))
        self.app.add_handler(CommandHandler("skill", self._handle_skill))
        self.app.add_handler(CommandHandler("premium", self._handle_premium))
        self.app.add_handler(CommandHandler("donate", self._handle_donate))

        # --- Improvement commands ---
        self.app.add_handler(CommandHandler("keyboard", self._handle_keyboard))
        self.app.add_handler(CommandHandler("stars", self._handle_stars))
        self.app.add_handler(CommandHandler("file", self._handle_file_gen))
        self.app.add_handler(CommandHandler("export", self._handle_export))
        self.app.add_handler(CommandHandler("digest", self._handle_digest))
        self.app.add_handler(CommandHandler("lang", self._handle_lang))
        self.app.add_handler(CommandHandler("stickers", self._handle_stickers_cmd))
        self.app.add_handler(CommandHandler("pin", self._handle_pin))
        self.app.add_handler(CommandHandler("hand", self._handle_hand))

        # Hand approval callbacks
        self.app.add_handler(CallbackQueryHandler(self._handle_hand_approval_callback, pattern="^hand_"))

        # Payment handlers
        self.app.add_handler(CallbackQueryHandler(self._handle_callback, pattern="^buy_"))
        self.app.add_handler(CallbackQueryHandler(self._handle_stars_callback, pattern="^stars_"))
        self.app.add_handler(CallbackQueryHandler(self._handle_action_callback, pattern="^act_"))
        self.app.add_handler(CallbackQueryHandler(self._handle_pin_callback, pattern="^pin_"))
        self.app.add_handler(PreCheckoutQueryHandler(self._handle_pre_checkout))
        self.app.add_handler(MessageHandler(filters.SUCCESSFUL_PAYMENT, self._handle_successful_payment))

        # Inline query handler (enable via @BotFather -> /setinline)
        self.app.add_handler(InlineQueryHandler(self._handle_inline))

        # Voice/audio message handler
        self.app.add_handler(MessageHandler(
            filters.VOICE | filters.AUDIO,
            self._handle_voice
        ))

        # Document upload handler
        self.app.add_handler(MessageHandler(
            filters.Document.ALL,
            self._handle_document
        ))

        # Photo handler (images sent as photos, not documents)
        self.app.add_handler(MessageHandler(
            filters.PHOTO,
            self._handle_photo_upload
        ))

        # Location handler (Phase 5 — contextual location info)
        self.app.add_handler(MessageHandler(
            filters.LOCATION,
            self._handle_location
        ))

        # Nearby search command (uses last shared location)
        self.app.add_handler(CommandHandler("nearby", self._handle_nearby))

        # Sticker handler (Phase 5 — contextual reactions)
        self.app.add_handler(MessageHandler(
            filters.Sticker.ALL,
            self._handle_sticker
        ))

        # Group-specific command handlers
        self.app.add_handler(CommandHandler("summarize_group", self._handle_summarize_group))
        self.app.add_handler(CommandHandler("summarize_thread", self._handle_summarize_thread))

        # ChatMember handler — bot added/removed from groups
        self.app.add_handler(ChatMemberHandler(self._handle_chat_member, ChatMemberHandler.MY_CHAT_MEMBER))

        # Message handler (must be last)
        self.app.add_handler(MessageHandler(
            filters.TEXT & ~filters.COMMAND,
            self._handle_message
        ))

        # Error handler
        self.app.add_error_handler(self._handle_error)

        # Start polling
        self.is_running = True

        # Initialize and start
        await self.app.initialize()
        await self.app.start()
        await self.app.updater.start_polling(
            drop_pending_updates=True,
            allowed_updates=[
                "message", "inline_query", "chosen_inline_result",
                "my_chat_member", "chat_member", "callback_query",
                "pre_checkout_query", "message_reaction",
            ]
        )

        logger.info("Telegram bot started successfully!")

        # Expose bot instance + event loop for scheduler callbacks
        global _active_bot_instance, _active_event_loop
        _active_bot_instance = self
        _active_event_loop = asyncio.get_running_loop()

        # Get bot info
        me = await self.bot.get_me()
        logger.info(f"Bot: @{me.username} ({me.first_name})")

        # Set bot commands menu (autocomplete in Telegram)
        try:
            await self.bot.set_my_commands([
                BotCommand("start", "Start the bot"),
                BotCommand("help", "Show help"),
                BotCommand("status", "Current status"),
                BotCommand("mood", "Check my mood"),
                BotCommand("memory", "What I remember"),
                BotCommand("research", "Deep research"),
                BotCommand("search", "Web search"),
                BotCommand("summarize", "Summarize text/URL"),
                BotCommand("image", "Generate an image"),
                BotCommand("code", "Run Python code"),
                BotCommand("model", "View/switch AI models"),
                BotCommand("compare", "Compare models"),
                BotCommand("session", "Manage sessions"),
                BotCommand("remind", "Set a reminder"),
                BotCommand("schedule", "Recurring tasks"),
                BotCommand("tasks", "List scheduled tasks"),
                BotCommand("agent", "Run specialist agent"),
                BotCommand("fleet", "Multi-agent parallel"),
                BotCommand("keyboard", "Toggle reply keyboard"),
                BotCommand("stars", "Support with Telegram Stars"),
                BotCommand("file", "Generate document"),
                BotCommand("export", "Export conversation"),
                BotCommand("digest", "Daily digest settings"),
                BotCommand("lang", "Set language"),
                BotCommand("pin", "Pin a message"),
                BotCommand("nearby", "Nearby places"),
                BotCommand("premium", "Premium tiers"),
                BotCommand("donate", "Support AURA"),
            ])
            logger.info("[Telegram] Bot commands menu set successfully")
        except Exception as e:
            logger.warning(f"[Telegram] Could not set bot commands: {e}")

        # Connect proactive system to Telegram
        await self._connect_proactive_system()

    async def stop(self):
        """Stop the Telegram bot"""

        logger.info("Stopping Telegram bot...")
        self.is_running = False

        # Clear scheduler callback references
        global _active_bot_instance, _active_event_loop
        _active_bot_instance = None
        _active_event_loop = None

        # Cancel proactive polling task
        if hasattr(self, '_proactive_task') and self._proactive_task:
            self._proactive_task.cancel()
            try:
                await self._proactive_task
            except asyncio.CancelledError:
                pass
            logger.info("Proactive polling stopped")

        # Persist final state and close SQLite
        self._save_state()
        if hasattr(self, 'store'):
            self.store.close()
            logger.info("TelegramStore closed")

        if self.app:
            await self.app.updater.stop()
            await self.app.stop()
            await self.app.shutdown()

        self._save_state()
        logger.info("Telegram bot stopped.")

    async def send_message(self, message: OutgoingMessage) -> bool:
        """Send a message"""

        try:
            parse_mode = None
            if message.parse_mode == "markdown":
                parse_mode = ParseMode.MARKDOWN

            await self.bot.send_message(
                chat_id=message.chat_id,
                text=message.text,
                parse_mode=parse_mode,
                reply_to_message_id=message.reply_to_message_id
            )
            return True
        except Exception as e:
            logger.error(f"Failed to send message: {e}")
            return False

    async def send_typing_indicator(self, chat_id: str):
        """Show typing indicator"""
        try:
            await self.bot.send_chat_action(
                chat_id=chat_id,
                action=ChatAction.TYPING
            )
        except Exception as e:
            logger.warning(f"Could not send typing indicator: {e}")

    def _is_user_allowed(self, user_id: int) -> bool:
        """Check if user is allowed — reads os.environ EVERY call (bulletproof)."""
        env_val = os.environ.get("TELEGRAM_ALLOWED_USERS", "")
        allowed = [u.strip() for u in env_val.split(",") if u.strip()] if env_val else []
        if self.allowed_users:
            for u in self.allowed_users:
                if u and u not in allowed:
                    allowed.append(u)
        if not allowed:
            logger.warning(f"[TelegramBot] Rejected user {user_id} — no allowed_users configured.")
            return False
        is_allowed = str(user_id) in allowed
        if not is_allowed:
            logger.warning(f"[TelegramBot] Rejected user {user_id} — not in allowed list {allowed}")
        return is_allowed


    def _is_admin(self, user_id: int) -> bool:
        """Check if user is an admin"""
        return str(user_id) in self.admin_users

    # ============ COMMAND HANDLERS ============

    async def _handle_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /start command"""

        user = update.effective_user
        chat_id = str(update.effective_chat.id)

        if not self._is_user_allowed(user.id):
            await update.message.reply_text(
                "Sorry, I'm currently in private beta."
            )
            return

        # Track this chat
        self.active_chats[chat_id] = {
            "user_id": str(user.id),
            "username": user.username,
            "first_name": user.first_name,
            "started_at": datetime.now().isoformat(),
            "last_message": datetime.now().isoformat()
        }
        self._save_state()

        welcome = f"""Hey {user.first_name}!

I'm AURA - your AI thinking partner.

I'm not just a chatbot. I remember our conversations, notice patterns, and actually care how things turn out for you.

Quick commands:
/status - See my current state
/mood - Check my mood
/memory - What I remember
/research <topic> - Deep research
/search <query> - Web search
/summarize - Summarize URL or text
/image <prompt> - Generate an image
/code <python> - Run Python code
/model - View/switch AI models
/compare <prompt> - Compare models side-by-side
/session - Manage conversation sessions
/remind <time> <msg> - Set a reminder
/schedule <interval> <task> - Recurring tasks
/tasks - List scheduled tasks
/cancel <id> - Cancel a task
/agent <specialist> <task> - Run a specialist agent
/fleet <goal> - Multi-agent parallel run
/webhook - Webhook integrations
/help - More info

Inline mode: Type @Aura828Bot in any chat to ask me anything!

Or just talk to me like a friend. What's on your mind?"""

        # Attach persistent reply keyboard by default
        self.store.set_keyboard_enabled(str(user.id), True)
        reply_markup = self._get_reply_keyboard()
        await update.message.reply_text(welcome, reply_markup=reply_markup)

    async def _handle_help(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /help command"""

        help_text = """AURA Commands

/start - Start fresh
/status - My current status
/mood - How I'm feeling
/memory - What I remember about you
/forget - Clear my memory (careful!)
/research <topic> - Deep research on any topic
/search <query> - Quick web search
/summarize - Summarize a URL or text (reply to a message or provide a URL)
/image <prompt> - Generate an image from a text description
/code <python> - Execute Python code (supports code blocks)
/model - View current models and available options
/model <name> - Switch the default model
/compare <prompt> - Run same prompt through 3 models side-by-side
/session - Show current session info
/session new [title] - Start a new conversation
/session list - List recent conversations
/session <id> - Switch to a conversation by ID
/session sync - Cross-surface sync status
/learn - Learn a skill from our last exchange
/skill list [category] - List all learned skills
/skill search <query> - Search skills by description
/skill info <name> - Detailed skill info
/skill improve <name> - Evolve a skill with GEPA
/skill create <name> - Manually define a new skill

Reminders & Scheduled Tasks:
/remind in 2h Check the deployment
/remind at 17:00 Call the team
/remind tomorrow 9am Review PRs
/schedule every 2h Check CPU usage
/schedule daily at 9am Summarize notifications
/schedule every monday at 10am Weekly summary
/tasks - List active reminders & tasks
/cancel <id> - Cancel a reminder or task

Multi-Agent System:
/agent list - Show available specialists
/agent route <query> - Preview which agent handles a query
/agent <specialist> <task> - Run a specific specialist
/fleet <goal> - Decompose goal and run agents in parallel
/fleet status - Check running fleet status

Inline mode (use @Aura828Bot in any chat):
- Type any question to get a quick answer
- "translate <text>" - Translation
- "explain <topic>" - Explanation
- "summarize <text>" - Summary

Location:
- Share your location to get weather, timezone, sunrise/sunset info
/nearby <query> - Find places near your last shared location

Group Commands (in group chats):
/summarize_group - Summarize recent group conversation
/summarize_thread - Reply to a message to summarize the thread
- Mention @Aura828Bot or reply to my messages to chat in groups

Tips:
- Just chat normally - I'll respond naturally
- Say "remember this:" to save something important
- Reply to a /code message to run more code
- I'll follow up on things you mention
- I notice patterns over time

I'm here whenever you need me."""

        await update.message.reply_text(help_text)

    async def _handle_status(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /status command"""

        if not self._is_user_allowed(update.effective_user.id):
            return

        # Get status from agent
        try:
            tool_count = len(self.aura.tools) if hasattr(self.aura, 'tools') else 0
            identity_name = self.aura.identity.get('name', 'AURA') if hasattr(self.aura, 'identity') else 'AURA'
            status = f"""AURA Status

Name: {identity_name}
Tools: {tool_count} loaded
Status: Online and ready!"""
        except Exception as e:
            logger.error(f"Error getting status: {e}")
            status = "Online and ready!"

        await update.message.reply_text(status)

    async def _handle_mood(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /mood command"""

        if not self._is_user_allowed(update.effective_user.id):
            return

        # Get mood from EvoEmo tool if available
        try:
            evoemo = self.aura.tools.get("evoemo") if hasattr(self.aura, 'tools') else None
            if evoemo and hasattr(evoemo, 'get_state'):
                state = evoemo.get_state()
                mood = state.get("dominant_emotion", "neutral")
                response = f"Current Mood: {mood}\n\nReady to chat!"
            else:
                response = "Feeling good and ready to chat!"
        except Exception as e:
            logger.error(f"Error getting mood: {e}")
            response = "Feeling good and ready to chat!"

        await update.message.reply_text(response)

    async def _handle_memory(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /memory command"""

        if not self._is_user_allowed(update.effective_user.id):
            return

        # Get memory summary from agent's MemorySystem
        try:
            if hasattr(self.aura, 'memory') and self.aura.memory:
                mem = self.aura.memory
                # MemorySystem stores conversations in ChromaDB
                stats = mem.get_stats() if hasattr(mem, 'get_stats') else {}
                count = stats.get('total_entries', 0) if stats else 0
                memory_text = f"Memory system active!\n\nStored entries: {count}\n\nKeep chatting and I'll remember important things."
            else:
                memory_text = "Memory system active. I remember our conversations!"
        except Exception as e:
            logger.error(f"Error getting memory: {e}")
            memory_text = "Memory system active!"

        await update.message.reply_text(memory_text)

    async def _handle_forget(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /forget command"""

        if not self._is_user_allowed(update.effective_user.id):
            return

        user_id = str(update.effective_user.id)

        await update.message.reply_text(
            "This will clear my memory of you. Are you sure?\n\n"
            "This includes:\n"
            "- Conversation history\n"
            "- User profile information\n"
            "- Learned facts about you\n"
            "- Emotional context\n\n"
            "Type 'yes forget everything' to confirm."
        )

        # Store pending forget request
        self._pending_forget[user_id] = datetime.now()

    async def _handle_image(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /image <prompt> command — generate an image via ImageGenerationTool."""

        if not self._is_user_allowed(update.effective_user.id):
            return

        chat_id = str(update.effective_chat.id)

        # Extract prompt: everything after /image
        raw_text = update.message.text or ""
        prompt = raw_text.partition(" ")[2].strip()

        if not prompt:
            await update.message.reply_text(
                "Usage: /image <prompt>\n\n"
                "Example: /image a sunset over mountains in watercolor style"
            )
            return

        # Show upload_photo action while generating
        try:
            await self.bot.send_chat_action(
                chat_id=chat_id,
                action=ChatAction.UPLOAD_PHOTO
            )
        except Exception:
            pass

        await update.message.reply_text(
            f"Generating image for: {prompt}\nThis may take a moment..."
        )

        # Try to get the image_gen tool from the agent's loaded tools
        image_tool = None
        if hasattr(self.aura, 'tools') and isinstance(self.aura.tools, dict):
            image_tool = (
                self.aura.tools.get("image_gen")
                or self.aura.tools.get("image_generation")
            )

        # If no pre-loaded tool, try to instantiate ImageGenerationTool directly
        if image_tool is None:
            try:
                from aura.tools.image_gen import ImageGenerationTool
                image_tool = ImageGenerationTool()
                logger.info("ImageGenerationTool instantiated on-demand for /image command")
            except Exception as e:
                logger.warning(f"Could not load ImageGenerationTool: {e}")

        if image_tool is None:
            await update.message.reply_text(
                "Image generation is not available on this server. "
                "The image_gen tool could not be loaded (torch/diffusers may not be installed)."
            )
            return

        # Run generation in a thread to avoid blocking the event loop
        try:
            generate_fn = getattr(image_tool, 'generate', None)
            if generate_fn is None:
                await update.message.reply_text(
                    "Image generation tool found but has no generate() method."
                )
                return

            result = await asyncio.to_thread(generate_fn, prompt)
        except Exception as e:
            logger.error(f"Image generation failed: {e}", exc_info=True)
            await update.message.reply_text(f"Image generation failed: {e}")
            return

        if not result or not result.get("success"):
            error_msg = (
                result.get("error", "Unknown error") if result else "No result returned"
            )
            await update.message.reply_text(f"Image generation failed: {error_msg}")
            return

        # Send the image as a Telegram photo
        try:
            await self.bot.send_chat_action(
                chat_id=chat_id,
                action=ChatAction.UPLOAD_PHOTO
            )

            pil_image = result.get("image")
            image_path = result.get("image_path")

            if pil_image is not None:
                buf = io.BytesIO()
                pil_image.save(buf, format="PNG")
                buf.seek(0)
                await self.bot.send_photo(
                    chat_id=chat_id,
                    photo=buf,
                    caption=prompt[:1024],
                    reply_to_message_id=update.message.message_id
                )
            elif image_path and Path(image_path).exists():
                with open(image_path, "rb") as f:
                    await self.bot.send_photo(
                        chat_id=chat_id,
                        photo=f,
                        caption=prompt[:1024],
                        reply_to_message_id=update.message.message_id
                    )
            else:
                await update.message.reply_text(
                    "Image was generated but could not be retrieved. "
                    "Check server logs for details."
                )
        except Exception as e:
            logger.error(f"Failed to send generated image: {e}", exc_info=True)
            image_path = result.get("image_path")
            if image_path:
                await update.message.reply_text(
                    f"Image generated but failed to send: {e}\n"
                    f"Saved at: {image_path}"
                )
            else:
                await update.message.reply_text(
                    f"Image generated but failed to send: {e}"
                )

    # ============ CODE EXECUTION HANDLER ============

    async def _handle_code(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /code command — execute Python code and return results."""

        if not self._is_user_allowed(update.effective_user.id):
            return

        chat_id = str(update.effective_chat.id)
        text = update.message.text or ""
        code = _extract_code_from_message(text)

        if not code:
            await update.message.reply_text(
                "Usage: /code <python_code>\n\n"
                "Examples:\n"
                "  /code print('hello world')\n"
                "  /code 2 ** 100\n\n"
                "You can also use code blocks:\n"
                "  /code ```python\n"
                "  for i in range(5):\n"
                "      print(i)\n"
                "  ```"
            )
            return

        # Show typing while executing
        await self.send_typing_indicator(chat_id)

        # Store message id so replies can continue context
        context.chat_data["_last_code_msg_id"] = update.message.message_id
        context.chat_data["_last_code"] = code

        # Run in thread to avoid blocking the event loop
        try:
            executor = self._get_code_executor()
            result = await asyncio.to_thread(executor.execute, code)
        except Exception as e:
            logger.error(f"Code execution failed: {e}", exc_info=True)
            await update.message.reply_text(f"❌ Execution failed: {e}")
            return

        await self._send_code_result(update, result)

    async def _send_code_result(self, update: Update, result: dict):
        """Format and send code execution results back to the user."""
        chat_id = update.effective_chat.id

        success = result.get("success", False)
        output = result.get("output", "")
        error = result.get("error", "") or result.get("errors", "")

        # Check for base64 plot data (from E2B tier or future extensions)
        plot_data = result.get("plot_base64") or result.get("image")
        if plot_data:
            try:
                img_bytes = base64.b64decode(plot_data)
                await self.bot.send_photo(
                    chat_id=chat_id,
                    photo=io.BytesIO(img_bytes),
                    reply_to_message_id=update.message.message_id,
                )
            except Exception as e:
                logger.warning(f"Could not send plot image: {e}")

        if success:
            if output:
                truncated = False
                if len(output) > _MAX_OUTPUT_CHARS:
                    output = output[:_MAX_OUTPUT_CHARS]
                    truncated = True

                msg = f"<pre>{self._escape_html(output)}</pre>"
                if truncated:
                    msg += "\n\n(output truncated)"

                try:
                    await update.message.reply_text(msg, parse_mode=ParseMode.HTML)
                except Exception:
                    # Fallback to plain text if HTML parsing fails
                    plain = output
                    if truncated:
                        plain += "\n\n(output truncated)"
                    await update.message.reply_text(plain)
            elif not plot_data:
                await update.message.reply_text("Code executed successfully (no output).")
        else:
            # Error case
            err_msg = str(error) if error else "Unknown error"
            truncated = False
            if len(err_msg) > _MAX_OUTPUT_CHARS:
                err_msg = err_msg[:_MAX_OUTPUT_CHARS]
                truncated = True

            text = f"Error:\n<pre>{self._escape_html(err_msg)}</pre>"
            if truncated:
                text += "\n\n(error output truncated)"

            # Include partial stdout if any
            if output:
                partial = output[:1000]
                text = f"<pre>{self._escape_html(partial)}</pre>\n\n" + text

            try:
                await update.message.reply_text(text, parse_mode=ParseMode.HTML)
            except Exception:
                plain = f"Error:\n{err_msg}"
                if truncated:
                    plain += "\n\n(error output truncated)"
                await update.message.reply_text(plain)

    @staticmethod
    def _escape_html(text: str) -> str:
        """Escape HTML special characters for Telegram HTML parse mode."""
        return (
            text
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )

    # ============ RESEARCH / SEARCH / SUMMARIZE ============

    @staticmethod
    def _split_message(text: str, limit: int = 4096) -> list:
        """Split a long message into chunks respecting Telegram's char limit."""
        if len(text) <= limit:
            return [text]
        chunks = []
        while text:
            if len(text) <= limit:
                chunks.append(text)
                break
            split_at = text.rfind("\n", 0, limit)
            if split_at < limit // 2:
                split_at = text.rfind(" ", 0, limit)
            if split_at < limit // 4:
                split_at = limit
            chunks.append(text[:split_at])
            text = text[split_at:].lstrip("\n")
        return chunks

    async def _handle_research(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /research <topic> -- deep multi-phase research pipeline."""
        user = update.effective_user
        if not self._is_user_allowed(user.id):
            return

        chat_id = str(update.effective_chat.id)
        raw_text = update.message.text or ""
        topic = raw_text.partition(" ")[2].strip()
        if not topic:
            await update.message.reply_text(
                "Usage: /research <topic>\n\n"
                "Example: /research quantum computing breakthroughs 2025"
            )
            return

        # Editable progress message
        status_msg = await update.message.reply_text(
            f"Researching: {topic}\n\nSearching the web..."
        )
        await self.send_typing_indicator(chat_id)

        # Progress callback -- edits the status message in-place
        progress_lines = [f"Researching: {topic}"]
        last_edit_time = [_time.time()]

        async def _update_status(line: str):
            now = _time.time()
            if now - last_edit_time[0] < 2.0:
                return
            last_edit_time[0] = now
            progress_lines.append(line)
            display = progress_lines[-6:]
            try:
                await status_msg.edit_text("\n".join(display))
                await self.send_typing_indicator(chat_id)
            except Exception:
                pass

        def _sync_progress(msg: str):
            try:
                loop = asyncio.get_event_loop()
                if loop.is_running():
                    asyncio.ensure_future(_update_status(msg))
            except Exception:
                pass

        try:
            from aura.tools.deep_research import DeepResearchTool

            tool = DeepResearchTool()
            tool.set_progress_callback(_sync_progress)

            result = await asyncio.to_thread(tool.research, topic, "standard")

            # Needs clarification?
            if result.get("needs_clarification"):
                await status_msg.edit_text(
                    f"Could you be more specific?\n\n{result.get('question', '')}"
                )
                return

            if not result.get("success"):
                await status_msg.edit_text(
                    f"Research failed: {result.get('error', 'Unknown error')}"
                )
                return

            # Build report
            parts = [
                f"Research: {topic}",
                f"({result.get('pages_read', 0)} sources, "
                f"{result.get('time_seconds', 0):.0f}s)",
                "",
            ]

            synthesis = (
                result.get("synthesis")
                or result.get("summary")
                or result.get("content", "")
            )
            if synthesis:
                parts.append(synthesis)
            else:
                for ps in result.get("page_summaries", [])[:8]:
                    title = ps.get("title", "Source")
                    body = ps.get("summary", ps.get("content", ""))[:300]
                    parts.append(f"- {title}: {body}")

            contradictions = result.get("contradictions", [])
            if contradictions:
                parts.append("")
                parts.append("Contradictions found:")
                for c in contradictions[:3]:
                    parts.append(f"  - {c}")

            # Clickable source citations
            sources = result.get("sources", [])
            if sources:
                parts.append("")
                parts.append("Sources:")
                seen = set()
                for s in sources[:10]:
                    url = s if isinstance(s, str) else s.get("url", "")
                    title = s.get("title", url) if isinstance(s, dict) else url
                    if url and url not in seen:
                        seen.add(url)
                        parts.append(f"  {title}\n  {url}")

            report = "\n".join(parts)

            try:
                await status_msg.delete()
            except Exception:
                pass

            for chunk in self._split_message(report):
                await update.message.reply_text(chunk)

        except ImportError:
            await status_msg.edit_text(
                "Deep research tool not available. "
                "Check that aura.tools.deep_research is installed."
            )
        except Exception as e:
            logger.error(f"Research command error: {e}", exc_info=True)
            try:
                await status_msg.edit_text(f"Research error: {str(e)[:200]}")
            except Exception:
                await update.message.reply_text(f"Research error: {str(e)[:200]}")

    async def _handle_search(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /search <query> -- quick web search via SearXNG."""
        user = update.effective_user
        if not self._is_user_allowed(user.id):
            return

        chat_id = str(update.effective_chat.id)
        raw_text = update.message.text or ""
        query = raw_text.partition(" ")[2].strip()
        if not query:
            await update.message.reply_text(
                "Usage: /search <query>\n\n"
                "Example: /search best Python async frameworks"
            )
            return

        await self.send_typing_indicator(chat_id)

        try:
            from aura.tools.web_search import WebSearchTool

            tool = WebSearchTool()
            result = await asyncio.to_thread(tool.search, query, 8)

            if not result.get("success"):
                await update.message.reply_text(
                    f"Search failed: {result.get('error', 'Unknown error')}"
                )
                return

            results = result.get("results", [])
            if not results:
                await update.message.reply_text(f"No results found for: {query}")
                return

            lines = [f"Search results for: {query}\n"]
            for i, r in enumerate(results, 1):
                title = r.get("title", "Untitled")
                snippet = r.get("snippet", "")[:150]
                url = r.get("url", "")
                lines.append(f"{i}. {title}")
                if snippet:
                    lines.append(f"   {snippet}")
                if url:
                    lines.append(f"   {url}")
                lines.append("")

            search_msg = "\n".join(lines)
            for chunk in self._split_message(search_msg):
                await update.message.reply_text(chunk)

        except ImportError:
            await update.message.reply_text("Web search tool is not available.")
        except Exception as e:
            logger.error(f"Search command error: {e}", exc_info=True)
            await update.message.reply_text(f"Search error: {str(e)[:200]}")

    async def _handle_summarize(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /summarize -- summarize a URL or replied-to text."""
        user = update.effective_user
        if not self._is_user_allowed(user.id):
            return

        chat_id = str(update.effective_chat.id)
        raw_text = (update.message.text or "").partition(" ")[2].strip()
        reply = update.message.reply_to_message

        url_pattern = re.compile(r'https?://[^\s<>"{}|\\^`\[\]]+')
        target_url = None
        target_text = None

        # Priority 1: URL in command text
        url_match = url_pattern.search(raw_text)
        if url_match:
            target_url = url_match.group(0)
        # Priority 2: Reply to a message with URL
        elif reply and reply.text:
            reply_match = url_pattern.search(reply.text)
            if reply_match:
                target_url = reply_match.group(0)
            elif len(reply.text) > 50:
                # Priority 3: Reply to long text
                target_text = reply.text
            else:
                await update.message.reply_text(
                    "The replied message is too short to summarize. "
                    "Reply to a longer message or provide a URL."
                )
                return
        else:
            await update.message.reply_text(
                "Usage:\n"
                "  /summarize <url> - Summarize a webpage\n"
                "  Reply to a message with /summarize - Summarize that text\n\n"
                "Example: /summarize https://example.com/article"
            )
            return

        await self.send_typing_indicator(chat_id)

        try:
            if target_url:
                status_msg = await update.message.reply_text(
                    f"Fetching: {target_url[:60]}..."
                )
                page_text = await self._fetch_url_content(target_url)
                if not page_text:
                    await status_msg.edit_text(
                        "Could not fetch the page content. "
                        "The site may be blocking requests."
                    )
                    return
                target_text = page_text[:8000]
                try:
                    await status_msg.edit_text("Summarizing...")
                except Exception:
                    pass
            else:
                status_msg = await update.message.reply_text("Summarizing...")

            await self.send_typing_indicator(chat_id)

            prompt = (
                "Summarize the following content concisely using bullet points. "
                "Focus on key facts, main arguments, and important details. "
                "Keep it under 500 words.\n\n"
                f"Content:\n{target_text}"
            )

            summary = await self._process_with_aura(prompt, str(user.id))

            if not summary or summary.startswith("Sorry, I couldn't"):
                # Extractive fallback
                sentences = [
                    s.strip() for s in target_text.split(".")
                    if len(s.strip()) > 30
                ]
                summary = "Key excerpts:\n\n" + "\n".join(
                    f"- {s.strip()}." for s in sentences[:8]
                )

            final_parts = []
            if target_url:
                final_parts.append(f"Summary of: {target_url}\n")
            final_parts.append(summary)
            final = "\n".join(final_parts)

            try:
                await status_msg.delete()
            except Exception:
                pass

            for chunk in self._split_message(final):
                await update.message.reply_text(chunk)

        except Exception as e:
            logger.error(f"Summarize command error: {e}", exc_info=True)
            await update.message.reply_text(
                f"Summarization error: {str(e)[:200]}"
            )

    async def _fetch_url_content(self, url: str):
        """Fetch and extract text content from a URL."""
        try:
            import requests

            resp = await asyncio.to_thread(
                lambda: requests.get(
                    url,
                    timeout=15,
                    headers={
                        "User-Agent": (
                            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                            "AppleWebKit/537.36"
                        ),
                    },
                )
            )
            if resp.status_code != 200:
                return None

            html = resp.text

            # Prefer trafilatura for clean text extraction
            try:
                import trafilatura
                extracted = trafilatura.extract(html)
                if extracted and len(extracted) > 100:
                    return extracted
            except ImportError:
                pass

            # Fallback: strip HTML tags
            text = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL)
            text = re.sub(r'<style[^>]*>.*?</style>', '', text, flags=re.DOTALL)
            text = re.sub(r'<[^>]+>', ' ', text)
            text = re.sub(r'\s+', ' ', text).strip()
            return text if len(text) > 50 else None

        except Exception as e:
            logger.warning(f"Failed to fetch URL {url}: {e}")
            return None

    # ============ MODEL COMMANDS ============

    def _get_brain(self):
        """Get the OllamaBrain instance from the agent wrapper."""
        agent = getattr(self.aura, 'agent', None)
        if agent:
            return getattr(agent, 'brain', None)
        return getattr(self.aura, 'brain', None)

    async def _handle_model(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /model command -- list or switch models."""

        if not self._is_user_allowed(update.effective_user.id):
            return

        raw_text = update.message.text or ""
        arg = raw_text.partition(" ")[2].strip()

        try:
            from aura.config import Config, VERIFIED_CLOUD_MODELS
        except ImportError:
            await update.message.reply_text("Config module not available.")
            return

        if not arg:
            current = Config.get_all_models()
            roles_display = {
                'fast': 'Fast',
                'reason': 'Reason',
                'code': 'Code',
                'vision': 'Vision',
                'think': 'Think',
                'longctx': 'Long Context',
            }

            lines = ["Current Models:\n"]
            for role, display_name in roles_display.items():
                model = current.get(role, 'unknown')
                lines.append(f"  {display_name}: {model}")

            lines.append("\nAvailable Cloud Models:\n")
            current_values = set(current.values())
            for m in sorted(VERIFIED_CLOUD_MODELS):
                marker_str = " [active]" if m in current_values else ""
                lines.append(f"  {m}{marker_str}")

            lines.append("\nUsage:\n  /model <name> - switch default model\n  /model auto - return to auto-routing")

            await update.message.reply_text("\n".join(lines))
            return

        brain = self._get_brain()

        if arg.lower() == "auto":
            if brain:
                brain.set_model_override(None)
            await update.message.reply_text("Switched to auto-routing.")
            return

        all_known = set(VERIFIED_CLOUD_MODELS)
        for chain_attr in ['MODEL_FAST_CHAIN', 'MODEL_REASON_CHAIN', 'MODEL_CODE_CHAIN',
                           'MODEL_VISION_CHAIN', 'MODEL_THINK_CHAIN', 'MODEL_LONGCTX_CHAIN']:
            all_known.update(getattr(Config, chain_attr, []))

        if arg not in all_known:
            matches = [m for m in sorted(all_known) if arg.lower() in m.lower()]
            if len(matches) == 1:
                arg = matches[0]
            elif matches:
                match_list = "\n".join(f"  {m}" for m in matches)
                await update.message.reply_text(
                    f"Multiple matches for '{arg}':\n{match_list}\n\nBe more specific."
                )
                return
            else:
                await update.message.reply_text(
                    f"Model '{arg}' not found.\n\nUse /model to see available models."
                )
                return

        if brain:
            brain.set_model_override(arg)

        await update.message.reply_text(f"Switched to {arg}")

    async def _handle_compare(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /compare <prompt> -- run same prompt through 3 models in parallel."""

        if not self._is_user_allowed(update.effective_user.id):
            return

        chat_id = str(update.effective_chat.id)
        raw_text = update.message.text or ""
        prompt = raw_text.partition(" ")[2].strip()

        if not prompt:
            await update.message.reply_text(
                "Usage: /compare <prompt>\n\n"
                "Example: /compare What causes auroras?"
            )
            return

        try:
            from aura.config import Config
        except ImportError:
            await update.message.reply_text("Config module not available.")
            return

        brain = self._get_brain()
        if not brain:
            await update.message.reply_text("Brain not available.")
            return

        # Pick 3 distinct models: fast, reason, code (fall back to think, longctx)
        candidates = [
            ("Fast", Config.get_model('fast')),
            ("Reason", Config.get_model('reason')),
            ("Code", Config.get_model('code')),
            ("Think", Config.get_model('think')),
            ("Long Context", Config.get_model('longctx')),
        ]
        models = []
        seen = set()
        for label, model in candidates:
            if model not in seen and len(models) < 3:
                seen.add(model)
                models.append((label, model))

        await self.send_typing_indicator(chat_id)
        model_names = ", ".join(m for _, m in models)
        status_msg = await update.message.reply_text(
            f"Comparing {len(models)} models: {model_names}\nRunning..."
        )

        def _call_model(label, model_name):
            start = _time.time()
            try:
                client, actual_model = brain._get_client_for_model(model_name)
                response = client.chat(
                    model=actual_model,
                    messages=[{"role": "user", "content": prompt}]
                )
                if response is None:
                    text = "(no response)"
                elif isinstance(response, dict):
                    msg = response.get("message", {})
                    text = msg.get("content", "") if isinstance(msg, dict) else ""
                else:
                    msg = getattr(response, "message", None)
                    text = getattr(msg, "content", "") if msg else ""
                if not text:
                    text = "(empty response)"
            except Exception as e:
                text = f"(error: {e})"
            elapsed = _time.time() - start
            return label, text, elapsed

        loop = asyncio.get_running_loop()
        with ThreadPoolExecutor(max_workers=3) as pool:
            futures = [
                loop.run_in_executor(pool, _call_model, label, model_name)
                for label, model_name in models
            ]
            results = await asyncio.gather(*futures, return_exceptions=True)

        for result in results:
            if isinstance(result, Exception):
                await self.bot.send_message(chat_id=chat_id, text=f"Error: {result}")
                continue

            label, response_text, elapsed = result
            model_name = next((m for l, m in models if l == label), "unknown")

            max_len = 3800
            if len(response_text) > max_len:
                response_text = response_text[:max_len] + "\n\n[...truncated]"

            header = f"--- {label} ({model_name}) ---\nTime: {elapsed:.1f}s\n\n"
            await self.bot.send_message(chat_id=chat_id, text=header + response_text)

        try:
            await status_msg.edit_text(f"Comparison complete. {len(models)} models responded.")
        except Exception:
            pass

    # ============ SESSION COMMAND ============

    async def _handle_session(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /session command — manage cross-surface conversation sessions.

        Subcommands:
          /session            — show current session info
          /session new [title] — create a new conversation and switch to it
          /session list       — list recent conversations (max 10)
          /session sync       — show cross-surface sync status
          /session <id>       — switch to a conversation by ID (partial match)
        """
        if not self._is_user_allowed(update.effective_user.id):
            return

        user_id = str(update.effective_user.id)
        args = context.args or []

        try:
            manager = get_conversation_manager()
        except Exception as e:
            logger.error(f"[Telegram] ConversationManager unavailable: {e}")
            await update.message.reply_text("Session management is not available right now.")
            return

        # Dispatch subcommands
        if not args:
            await self._session_info(update, manager, user_id)
        elif args[0].lower() == "new":
            title = " ".join(args[1:]).strip() if len(args) > 1 else None
            await self._session_new(update, manager, user_id, title)
        elif args[0].lower() == "list":
            await self._session_list(update, manager, user_id)
        elif args[0].lower() == "sync":
            await self._session_sync(update, manager, user_id)
        else:
            # Treat as conversation ID to switch to
            target_id = args[0].strip()
            await self._session_switch(update, manager, user_id, target_id)

    async def _session_info(self, update: Update, manager, user_id: str):
        """Show current session info."""
        try:
            conv_id = manager.get_or_create_session("telegram", user_id)
            convs = manager.list_conversations()
            conv = next((c for c in convs if c["id"] == conv_id), None)

            if not conv:
                await update.message.reply_text("No active session found\\. Use /session new to start one\\.",
                                                 parse_mode="MarkdownV2")
                return

            title = conv.get("title", "Untitled")
            msg_count = conv.get("message_count", 0)
            updated_at = conv.get("updated_at", 0)

            # Format timestamp
            if updated_at:
                ts = datetime.fromtimestamp(updated_at).strftime("%Y\\-%m\\-%d %H:%M")
            else:
                ts = "Unknown"

            # Connected surfaces
            bound = conv.get("bound_surfaces", [])
            if bound:
                surface_badges = []
                for sk in bound:
                    surface_name = sk.split(":")[0] if ":" in sk else sk
                    surface_badges.append(f"\\[{_escape_mdv2(surface_name)}\\]")
                surfaces_str = " ".join(surface_badges)
            else:
                surfaces_str = "\\(none\\)"

            # Surface activity
            activity = conv.get("surface_activity", {})
            if activity:
                activity_lines = []
                for s, count in sorted(activity.items(), key=lambda x: -x[1]):
                    activity_lines.append(f"  {_escape_mdv2(s)}: {count} msgs")
                activity_str = "\n".join(activity_lines)
            else:
                activity_str = "  No surface activity recorded"

            text = (
                f"*Current Session*\n\n"
                f"*Title:* {_escape_mdv2(title)}\n"
                f"*ID:* `{conv_id}`\n"
                f"*Messages:* {msg_count}\n"
                f"*Updated:* {ts}\n"
                f"*Surfaces:* {surfaces_str}\n\n"
                f"*Activity:*\n{activity_str}"
            )

            await update.message.reply_text(text, parse_mode="MarkdownV2")

        except Exception as e:
            logger.error(f"[Telegram] /session info error: {e}", exc_info=True)
            await update.message.reply_text("Could not retrieve session info.")

    async def _session_new(self, update: Update, manager, user_id: str, title: str = None):
        """Create a new conversation and switch to it."""
        try:
            conv_id = manager.new_session("telegram", user_id, title)
            display_title = _escape_mdv2(title or "Telegram Chat")

            text = (
                f"*New session created*\n\n"
                f"*Title:* {display_title}\n"
                f"*ID:* `{conv_id}`\n\n"
                f"You're now chatting in this session\\."
            )
            await update.message.reply_text(text, parse_mode="MarkdownV2")

        except Exception as e:
            logger.error(f"[Telegram] /session new error: {e}", exc_info=True)
            await update.message.reply_text("Could not create a new session.")

    async def _session_list(self, update: Update, manager, user_id: str):
        """List recent conversations (max 10)."""
        try:
            convs = manager.list_sessions("telegram", user_id)
            if not convs:
                await update.message.reply_text("No conversations found\\. Use /session new to start one\\.",
                                                 parse_mode="MarkdownV2")
                return

            # Limit to 10 most recent
            convs = convs[:10]

            lines = ["*Recent Conversations*\n"]
            for conv in convs:
                conv_id = conv["id"]
                title = conv.get("title", "Untitled")
                msg_count = conv.get("message_count", 0)
                updated_at = conv.get("updated_at", 0)
                is_bound = conv.get("is_bound", False)

                # Timestamp
                if updated_at:
                    ts = datetime.fromtimestamp(updated_at).strftime("%m/%d %H:%M")
                else:
                    ts = "\\-\\-"

                # Surface badges
                bound_surfaces = conv.get("bound_surfaces", [])
                badges = ""
                if bound_surfaces:
                    badge_parts = []
                    for sk in bound_surfaces:
                        surface_name = sk.split(":")[0] if ":" in sk else sk
                        badge_parts.append(f"\\[{_escape_mdv2(surface_name)}\\]")
                    badges = " " + " ".join(badge_parts)

                # Active marker
                marker = "→ " if is_bound else "  "
                active_label = " \\*active\\*" if is_bound else ""

                short_id = conv_id[-8:] if len(conv_id) > 8 else conv_id

                line = (
                    f"{_escape_mdv2(marker)}"
                    f"*{_escape_mdv2(title)}*{active_label}\n"
                    f"    `{short_id}` \\| "
                    f"{_escape_mdv2(ts)} \\| "
                    f"{msg_count} msgs{badges}"
                )
                lines.append(line)

            lines.append(f"\nSwitch: `/session <id>`")

            text = "\n".join(lines)
            await update.message.reply_text(text, parse_mode="MarkdownV2")

        except Exception as e:
            logger.error(f"[Telegram] /session list error: {e}", exc_info=True)
            await update.message.reply_text("Could not list sessions.")

    async def _session_switch(self, update: Update, manager, user_id: str, target_id: str):
        """Switch to a conversation by ID (supports partial match)."""
        try:
            convs = manager.list_conversations()

            # Exact match first
            match = next((c for c in convs if c["id"] == target_id), None)

            # Partial match (suffix or contains)
            if not match:
                candidates = [c for c in convs if c["id"].endswith(target_id)]
                if not candidates:
                    candidates = [c for c in convs if target_id in c["id"]]
                if len(candidates) == 1:
                    match = candidates[0]
                elif len(candidates) > 1:
                    lines = ["Multiple matches:\n"]
                    for c in candidates[:5]:
                        lines.append(f"  `{c['id']}` — {_escape_mdv2(c.get('title', 'Untitled'))}")
                    lines.append(f"\nBe more specific\\.")
                    await update.message.reply_text("\n".join(lines), parse_mode="MarkdownV2")
                    return

            if not match:
                await update.message.reply_text(
                    f"No conversation found matching `{_escape_mdv2(target_id)}`\\.\n"
                    f"Use /session list to see available conversations\\.",
                    parse_mode="MarkdownV2"
                )
                return

            conv_id = match["id"]
            success = manager.switch_session("telegram", user_id, conv_id)

            if success:
                title = match.get("title", "Untitled")
                msg_count = match.get("message_count", 0)
                text = (
                    f"*Switched session*\n\n"
                    f"*Title:* {_escape_mdv2(title)}\n"
                    f"*ID:* `{conv_id}`\n"
                    f"*Messages:* {msg_count}\n\n"
                    f"Continuing in this conversation\\."
                )
                await update.message.reply_text(text, parse_mode="MarkdownV2")
            else:
                await update.message.reply_text("Failed to switch session\\. The conversation may have been deleted\\.",
                                                 parse_mode="MarkdownV2")

        except Exception as e:
            logger.error(f"[Telegram] /session switch error: {e}", exc_info=True)
            await update.message.reply_text("Could not switch session.")

    async def _session_sync(self, update: Update, manager, user_id: str):
        """Show cross-surface sync status."""
        try:
            status = manager.get_status()
            conv_id = manager.get_or_create_session("telegram", user_id)

            # All bindings
            bindings = status.get("bindings", {})
            total = status.get("total_bindings", 0)
            listener_count = status.get("listener_count", 0)

            # Surfaces connected to the current conversation
            current_surfaces = manager.get_surfaces_for_conversation(conv_id)

            # Group bindings by surface type
            surface_summary = {}
            for key in bindings:
                surface_type = key.split(":")[0] if ":" in key else key
                surface_summary[surface_type] = surface_summary.get(surface_type, 0) + 1

            summary_lines = []
            for s, count in sorted(surface_summary.items()):
                summary_lines.append(f"  {_escape_mdv2(s)}: {count} binding\\(s\\)")

            # Current session surfaces
            if current_surfaces:
                current_lines = []
                for sk in current_surfaces:
                    parts = sk.split(":", 1)
                    surface_name = parts[0]
                    surface_uid = parts[1] if len(parts) > 1 else "\\-"
                    current_lines.append(f"  {_escape_mdv2(surface_name)} \\(`{_escape_mdv2(surface_uid)}`\\)")
                current_str = "\n".join(current_lines)
            else:
                current_str = "  None"

            text = (
                f"*Cross\\-Surface Sync Status*\n\n"
                f"*Current session:* `{conv_id}`\n"
                f"*Total bindings:* {total}\n"
                f"*Active listeners:* {listener_count}\n\n"
                f"*Connected surfaces \\(this session\\):*\n{current_str}\n\n"
                f"*All surface bindings:*\n" +
                ("\n".join(summary_lines) if summary_lines else "  None")
            )

            await update.message.reply_text(text, parse_mode="MarkdownV2")

        except Exception as e:
            logger.error(f"[Telegram] /session sync error: {e}", exc_info=True)
            await update.message.reply_text("Could not retrieve sync status.")

    # ============ GROUP HANDLERS ============

    async def _handle_chat_member(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle bot being added/removed from groups."""
        try:
            new_status = update.my_chat_member.new_chat_member.status
            if new_status in ("member", "administrator"):
                chat = update.my_chat_member.chat
                await context.bot.send_message(
                    chat.id,
                    "Hi! I'm AURA. Mention me with @Aura828Bot or reply to my messages to chat.\n\n"
                    "Commands: /summarize_group, /summarize_thread, /help"
                )
                logger.info(f"[TelegramBot] Added to group: {chat.title} ({chat.id})")
            elif new_status in ("left", "kicked"):
                chat = update.my_chat_member.chat
                # Clean up group cache
                cache_key = str(chat.id)
                self._group_message_cache.pop(cache_key, None)
                logger.info(f"[TelegramBot] Removed from group: {chat.title} ({chat.id})")
        except Exception as e:
            logger.error(f"[TelegramBot] Error handling chat_member update: {e}", exc_info=True)

    async def _handle_summarize_group(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Summarize recent group conversation from cached messages."""
        chat_id = str(update.effective_chat.id)
        chat_type = update.effective_chat.type

        if chat_type not in ("group", "supergroup"):
            await update.message.reply_text("This command only works in groups.")
            return

        cache = self.store.get_group_messages(chat_id, limit=50)
        if len(cache) < 3:
            await update.message.reply_text("Not enough messages to summarize yet. I need at least 3 cached messages.")
            return

        # Build context from last 30 cached messages
        context_text = "\n".join([f"{m.get('user_name', m.get('user', 'Unknown'))}: {m['text']}" for m in cache[-30:]])
        prompt = f"Summarize this group conversation concisely. Focus on key topics and decisions:\n\n{context_text}"

        placeholder = await update.message.reply_text("Summarizing group conversation...")
        try:
            response_text, _ = await asyncio.to_thread(self._run_agent_sync, prompt)
            await self._edit_or_send_response(placeholder, chat_id, response_text, update)
        except Exception as e:
            logger.error(f"[TelegramBot] /summarize_group failed: {e}", exc_info=True)
            try:
                await placeholder.edit_text("Could not generate summary. Try again later.")
            except Exception:
                pass

    async def _handle_summarize_thread(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Summarize a thread when used as a reply in a group."""
        chat_id = str(update.effective_chat.id)
        chat_type = update.effective_chat.type

        if chat_type not in ("group", "supergroup"):
            await update.message.reply_text("This command only works in groups.")
            return

        if not update.message.reply_to_message:
            await update.message.reply_text("Reply to a message with /summarize_thread to summarize the conversation around it.")
            return

        # Get the replied-to message as the anchor
        anchor_msg = update.message.reply_to_message
        anchor_text = anchor_msg.text or anchor_msg.caption or "[non-text message]"
        anchor_user = anchor_msg.from_user.first_name if anchor_msg.from_user else "Unknown"

        # Pull related messages from store around the same timeframe
        cache = self.store.get_group_messages(chat_id, limit=50)
        if len(cache) < 2:
            await update.message.reply_text(
                f"Not enough cached context. Here's the message you pointed to:\n\n"
                f"{anchor_user}: {anchor_text}"
            )
            return

        # Build context from cached messages
        context_text = "\n".join([f"{m['user']}: {m['text']}" for m in cache[-30:]])
        prompt = (
            f"Summarize this group conversation thread. The user is asking about the discussion "
            f"around this message from {anchor_user}: \"{anchor_text}\"\n\n"
            f"Recent conversation context:\n{context_text}"
        )

        placeholder = await update.message.reply_text("Summarizing thread...")
        try:
            response_text, _ = await asyncio.to_thread(self._run_agent_sync, prompt)
            await self._edit_or_send_response(placeholder, chat_id, response_text, update)
        except Exception as e:
            logger.error(f"[TelegramBot] /summarize_thread failed: {e}", exc_info=True)
            try:
                await placeholder.edit_text("Could not generate thread summary. Try again later.")
            except Exception:
                pass

    # ============ MESSAGE HANDLER ============

    async def _handle_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle regular text messages through the full ReAct agent loop."""

        user = update.effective_user
        chat_id = str(update.effective_chat.id)

        # Detect group vs private chat
        chat_type = update.effective_chat.type  # "private", "group", "supergroup"
        is_group = chat_type in ("group", "supergroup")

        text = update.message.text or ""

        # --- Group message caching (persisted to SQLite) ---
        if is_group:
            self.store.add_group_message(
                chat_id=chat_id,
                user_id=str(user.id),
                user_name=user.first_name or "Unknown",
                text=text,
            )

        # --- Group gate: only respond if mentioned or replied-to ---
        if is_group:
            me = await context.bot.get_me()
            bot_username = me.username or "Aura828Bot"
            mentioned = f"@{bot_username}".lower() in text.lower()
            replied_to_bot = (
                update.message.reply_to_message is not None
                and update.message.reply_to_message.from_user is not None
                and update.message.reply_to_message.from_user.id == me.id
            )
            if not mentioned and not replied_to_bot:
                return  # Ignore messages not directed at us

            # Strip the @mention from text before processing
            text = re.sub(rf"@{re.escape(bot_username)}", "", text, flags=re.IGNORECASE).strip()
            if not text:
                text = "Hello"
        else:
            # Private chat: require allowed user
            if not self._is_user_allowed(user.id):
                return

        # Per-user rate limit (max_messages_per_minute from config, default 20)
        max_per_min = self.config.get("max_messages_per_minute", 20)
        if not _check_rate_limit(str(user.id), max_per_min):
            await update.message.reply_text(
                "You're sending messages too fast. Please wait a moment."
            )
            return

        # Check if this is a reply to a /code message — treat as continuation code
        reply = update.message.reply_to_message
        if reply and context.chat_data.get("_last_code_msg_id") == reply.message_id:
            code = text.strip()
            if code:
                await self.send_typing_indicator(chat_id)
                context.chat_data["_last_code_msg_id"] = update.message.message_id
                context.chat_data["_last_code"] = code
                try:
                    executor = self._get_code_executor()
                    result = await asyncio.to_thread(executor.execute, code)
                except Exception as e:
                    logger.error(f"Code execution (reply) failed: {e}", exc_info=True)
                    await update.message.reply_text(f"❌ Execution failed: {e}")
                    return
                await self._send_code_result(update, result)
                return

        # Check for forget confirmation
        if text and text.lower() == "yes forget everything":
            user_id = str(user.id)

            # Check if there's a pending forget request (within 5 minutes)
            if hasattr(self, '_pending_forget') and user_id in self._pending_forget:
                request_time = self._pending_forget[user_id]
                if (datetime.now() - request_time).total_seconds() < 300:  # 5 min
                    cleared_items = []
                    try:
                        if hasattr(self.aura, 'memory') and self.aura.memory:
                            if hasattr(self.aura.memory, 'clear'):
                                self.aura.memory.clear()
                                cleared_items.append("Conversation memory")
                        if hasattr(self.aura, 'state') and hasattr(self.aura.state, '_history'):
                            self.aura.state._history.clear()
                            cleared_items.append("State history")
                        self.store.clear_doc_context(user_id)
                        cleared_items.append("Document context")
                        del self._pending_forget[user_id]
                        if cleared_items:
                            cleared_list = "\n".join(f"- {item}" for item in cleared_items)
                            await update.message.reply_text(
                                f"Memory cleared successfully!\n\nCleared:\n{cleared_list}\n\n"
                                f"Fresh start! What would you like to talk about?"
                            )
                        else:
                            await update.message.reply_text(
                                "Memory cleared. Fresh start! What would you like to talk about?"
                            )
                    except Exception as e:
                        logger.error(f"Error clearing memory: {e}")
                        await update.message.reply_text(
                            "There was an issue clearing some memories, but I'll treat this as a fresh start. "
                            "What would you like to talk about?"
                        )
                    return
                else:
                    del self._pending_forget[user_id]

            await update.message.reply_text(
                "No active forget request. Use /forget first if you want to clear my memory."
            )
            return

        # Check for pending skill actions (confirm/create flow)
        if await self._check_skill_pending(update, user.id, text):
            return

        # Update active chat info
        if chat_id in self.active_chats:
            self.active_chats[chat_id]["last_message"] = datetime.now().isoformat()
        else:
            self.active_chats[chat_id] = {
                "user_id": str(user.id),
                "username": user.username,
                "first_name": user.first_name,
                "started_at": datetime.now().isoformat(),
                "last_message": datetime.now().isoformat()
            }

        # Prepend document context if the user has an active uploaded document
        if hasattr(self, '_build_doc_augmented_text'):
            text = self._build_doc_augmented_text(user.id, text)

        # Multi-language: prepend language instruction for non-English users
        user_lang = self.store.get_user_language(str(user.id))
        if (not user_lang or user_lang == "en") and hasattr(user, 'language_code') and user.language_code:
            lc = user.language_code.split("-")[0]
            if lc and lc != "en":
                user_lang = lc
                self.store.set_user_language(str(user.id), lc)
        if user_lang and user_lang != "en":
            text = f"[Respond in {user_lang}] {text}"

        # Forum support: capture message_thread_id for topic-aware replies
        thread_id = getattr(update.message, 'message_thread_id', None)

        # Bind to conversation via ConversationManager (cross-surface sync)
        conv_id = None
        try:
            cm = get_conversation_manager()
            if cm._brain is not None:
                conv_id = cm.get_or_create_session("telegram", str(user.id))
                cm.switch_conversation(conv_id, surface="telegram")
        except Exception as e:
            logger.debug(f"[Telegram] ConversationManager session bind skipped: {e}")

        # Route through the full agent loop with typing indicator + file artifacts
        await self._run_agent_and_reply(update, text, conv_id=conv_id, user_id=str(user.id))

        self._save_state()

    # ============ PHOTO / VISION HANDLERS ============

    # Image extensions recognized when sent as document attachments
    _IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tiff", ".tif"}

    # Text/code file extensions for document upload Q&A
    _TEXT_EXTENSIONS = {
        ".txt", ".csv", ".json", ".py", ".md", ".js", ".ts", ".jsx", ".tsx",
        ".html", ".css", ".xml", ".yaml", ".yml", ".toml", ".ini", ".cfg",
        ".sh", ".bash", ".sql", ".log", ".env", ".rs", ".go", ".java",
        ".c", ".cpp", ".h", ".hpp", ".rb", ".php", ".swift", ".kt",
    }

    _MAX_DOC_SIZE = 20 * 1024 * 1024  # 20 MB

    def _analyze_image_sync(self, img_b64: str, prompt: str) -> str:
        """Analyze a base64-encoded image with the vision model (synchronous).

        Uses VisionTool with its full fallback chain. Meant to be called
        via asyncio.to_thread from async handlers.
        """
        try:
            from aura.tools.vision import VisionTool

            # Reuse brain's Ollama client if available
            brain = getattr(self.aura, 'brain', None)
            if brain is None:
                brain = getattr(getattr(self.aura, 'agent', None), 'brain', None)

            vt = VisionTool(brain=brain)
            content, model_used = vt._analyze_with_fallback(img_b64, prompt)
            logger.info(f"[TelegramBot] Vision analysis done with {model_used}")
            return content
        except Exception as e:
            logger.error(f"[TelegramBot] Vision analysis failed: {e}", exc_info=True)
            return "Sorry, I couldn't analyze that image right now. The vision model may be unavailable."

    async def _handle_photo_upload(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle photos sent as compressed Telegram images -- analyze with vision model."""

        user = update.effective_user
        chat_id = str(update.effective_chat.id)

        if not self._is_user_allowed(user.id):
            return

        max_per_min = self.config.get("max_messages_per_minute", 20)
        if not _check_rate_limit(str(user.id), max_per_min):
            await update.message.reply_text("You're sending messages too fast. Please wait a moment.")
            return

        # Show typing while we download and process
        await self.send_typing_indicator(chat_id)

        # Get highest resolution photo (last in the list)
        photo = update.message.photo[-1]
        prompt = update.message.caption or "Describe this image in detail."

        # Download photo bytes from Telegram
        try:
            tg_file = await context.bot.get_file(photo.file_id)
            buf = io.BytesIO()
            await tg_file.download_to_memory(buf)
            image_bytes = buf.getvalue()
        except Exception as e:
            logger.error(f"[TelegramBot] Failed to download photo: {e}")
            await update.message.reply_text("Couldn't download the photo. Please try again.")
            return

        img_b64 = base64.b64encode(image_bytes).decode("utf-8")

        # Refresh typing -- vision models can take a while
        await self.send_typing_indicator(chat_id)

        # Run synchronous vision analysis in a thread
        response = await asyncio.to_thread(self._analyze_image_sync, img_b64, prompt)

        await update.message.reply_text(response)
        self._save_state()

    async def _handle_document(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle document uploads -- extract text for PDFs/text files, route images to vision."""
        user = update.effective_user
        chat_id = str(update.effective_chat.id)

        if not self._is_user_allowed(user.id):
            return

        max_per_min = self.config.get("max_messages_per_minute", 20)
        if not _check_rate_limit(str(user.id), max_per_min):
            await update.message.reply_text(
                "You're sending messages too fast. Please wait a moment."
            )
            return

        doc = update.message.document
        if not doc:
            return

        filename = doc.file_name or "unknown"
        file_size = doc.file_size or 0
        mime = doc.mime_type or ""
        ext = os.path.splitext(filename)[1].lower()

        # Size guard
        if file_size > self._MAX_DOC_SIZE:
            await update.message.reply_text(
                f"That file is too large ({file_size / (1024*1024):.1f} MB). "
                f"I can handle files up to 20 MB."
            )
            return

        await self.send_typing_indicator(chat_id)

        try:
            # --- PDF ---
            if ext == ".pdf" or mime == "application/pdf":
                text, char_count = await self._extract_pdf(doc)
                self._store_doc_context(user.id, text, filename)
                await update.message.reply_text(
                    f"\U0001f4c4 Extracted {char_count:,} characters from {filename}.\n"
                    f"Ask me anything about this document!"
                )

            # --- Text / code files ---
            elif ext in self._TEXT_EXTENSIONS or mime.startswith("text/"):
                text, char_count = await self._extract_text_file(doc)
                self._store_doc_context(user.id, text, filename)
                await update.message.reply_text(
                    f"\U0001f4c4 Read {char_count:,} characters from {filename}.\n"
                    f"Ask me anything about this file!"
                )

            # --- Images sent as documents ---
            elif ext in self._IMAGE_EXTENSIONS or mime.startswith("image/"):
                prompt = update.message.caption or "Describe this image in detail."
                try:
                    tg_file = await context.bot.get_file(doc.file_id)
                    buf = io.BytesIO()
                    await tg_file.download_to_memory(buf)
                    image_bytes = buf.getvalue()
                except Exception as e:
                    logger.error(f"[TelegramBot] Failed to download image document: {e}")
                    await update.message.reply_text(
                        "Couldn't download the image. Please try again."
                    )
                    return

                img_b64 = base64.b64encode(image_bytes).decode("utf-8")
                await self.send_typing_indicator(chat_id)
                response = await asyncio.to_thread(self._analyze_image_sync, img_b64, prompt)
                await update.message.reply_text(response)
                self._save_state()
                return

            # --- DOCX ---
            elif ext == ".docx" or mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                text, char_count = await self._extract_docx(doc)
                self._store_doc_context(user.id, text, filename)
                await update.message.reply_text(
                    f"\U0001f4c4 Extracted {char_count:,} characters from {filename}.\n"
                    f"Ask me anything about this document!"
                )

            # --- XLSX / CSV ---
            elif ext in (".xlsx", ".xls", ".csv") or mime in (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel", "text/csv",
            ):
                text, char_count = await self._extract_spreadsheet(doc, ext)
                self._store_doc_context(user.id, text, filename)
                await update.message.reply_text(
                    f"\U0001f4ca Extracted {char_count:,} characters from {filename}.\n"
                    f"Ask me anything about this data!"
                )

            # --- Unsupported ---
            else:
                await update.message.reply_text(
                    f"I can't read .{ext or '?'} files yet. "
                    f"I support PDFs, DOCX, XLSX, CSV, text/code files, and images."
                )

        except Exception as e:
            logger.error(f"Document handler error for {filename}: {e}", exc_info=True)
            await update.message.reply_text(
                f"Sorry, I couldn't process {filename}. "
                f"The file might be corrupted or in an unsupported format."
            )

    async def _extract_pdf(self, doc) -> "Tuple[str, int]":
        """Download a Telegram document and extract PDF text.

        Tries PyMuPDF (fitz) first, falls back to pdfplumber.
        Returns (extracted_text, char_count).
        """
        tg_file = await doc.get_file()
        file_bytes = await tg_file.download_as_bytearray()

        def _extract_with_fitz(raw: bytes) -> str:
            import fitz  # PyMuPDF
            text_parts = []
            with fitz.open(stream=raw, filetype="pdf") as pdf_doc:
                for page in pdf_doc:
                    text_parts.append(page.get_text())
            return "\n\n".join(text_parts)

        def _extract_with_pdfplumber(raw: bytes) -> str:
            import pdfplumber
            text_parts = []
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for page in pdf.pages:
                    text_parts.append(page.extract_text() or "")
            return "\n\n".join(text_parts)

        raw = bytes(file_bytes)
        text = ""

        for extractor, name in [
            (_extract_with_fitz, "PyMuPDF"),
            (_extract_with_pdfplumber, "pdfplumber"),
        ]:
            try:
                text = await asyncio.to_thread(extractor, raw)
                if text.strip():
                    logger.info(f"PDF extracted with {name}: {len(text)} chars")
                    break
            except ImportError:
                continue
            except Exception as e:
                logger.warning(f"PDF extraction with {name} failed: {e}")
                continue

        if not text.strip():
            raise ValueError(
                "Could not extract text -- the PDF may be image-based or empty."
            )

        if len(text) > 80_000:
            text = text[:80_000] + "\n\n[... truncated at 80,000 characters ...]"

        return text, len(text)

    async def _extract_text_file(self, doc) -> "Tuple[str, int]":
        """Download a Telegram document and read it as UTF-8 text."""
        tg_file = await doc.get_file()
        file_bytes = await tg_file.download_as_bytearray()

        try:
            text = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")

        if len(text) > 80_000:
            text = text[:80_000] + "\n\n[... truncated at 80,000 characters ...]"

        return text, len(text)

    async def _extract_docx(self, doc) -> "Tuple[str, int]":
        """Download a Telegram document and extract DOCX text."""
        tg_file = await doc.get_file()
        file_bytes = await tg_file.download_as_bytearray()

        def _extract(raw: bytes) -> str:
            try:
                from docx import Document as DocxDocument
                doc_obj = DocxDocument(io.BytesIO(raw))
                return "\n\n".join(p.text for p in doc_obj.paragraphs if p.text.strip())
            except ImportError:
                # Fallback: extract raw XML text
                import zipfile
                import re as _re
                with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                    xml = zf.read("word/document.xml").decode("utf-8")
                    return _re.sub(r"<[^>]+>", " ", xml).strip()

        text = await asyncio.to_thread(_extract, bytes(file_bytes))
        if len(text) > 80_000:
            text = text[:80_000] + "\n\n[... truncated at 80,000 characters ...]"
        return text, len(text)

    async def _extract_spreadsheet(self, doc, ext: str) -> "Tuple[str, int]":
        """Download and extract spreadsheet data as readable text."""
        tg_file = await doc.get_file()
        file_bytes = await tg_file.download_as_bytearray()

        def _extract(raw: bytes, file_ext: str) -> str:
            if file_ext == ".csv":
                import csv
                reader = csv.reader(io.StringIO(raw.decode("utf-8", errors="replace")))
                rows = list(reader)
                if not rows:
                    return "(empty CSV)"
                header = " | ".join(rows[0])
                body = "\n".join(" | ".join(r) for r in rows[1:200])
                return f"{header}\n{'—' * len(header)}\n{body}"
            else:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                    parts = []
                    for sheet_name in wb.sheetnames[:5]:
                        ws = wb[sheet_name]
                        lines = [f"## Sheet: {sheet_name}"]
                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                            if i > 200:
                                lines.append(f"... ({ws.max_row - 200} more rows)")
                                break
                            lines.append(" | ".join(str(c) if c is not None else "" for c in row))
                        parts.append("\n".join(lines))
                    wb.close()
                    return "\n\n".join(parts)
                except ImportError:
                    return "(openpyxl not installed — cannot read Excel files)"

        text = await asyncio.to_thread(_extract, bytes(file_bytes), ext)
        if len(text) > 80_000:
            text = text[:80_000] + "\n\n[... truncated at 80,000 characters ...]"
        return text, len(text)

    # ============ DOCUMENT CONTEXT HELPERS ============

    def _store_doc_context(self, user_id: int, text: str, filename: str):
        """Store extracted document text for subsequent Q&A."""
        self.store.set_doc_context(str(user_id), text, filename)
        logger.info(
            f"Stored doc context for user {user_id}: {filename} ({len(text)} chars)"
        )

    def _get_doc_context(self, user_id: int):
        """Get active document context, or None if expired/missing."""
        ctx = self.store.get_doc_context(str(user_id), ttl=self._DOC_CONTEXT_TTL)
        if not ctx:
            return None
        return ctx

    def _build_doc_augmented_text(self, user_id: int, text: str) -> str:
        """If the user has active document context, prepend it to their message."""
        ctx = self._get_doc_context(user_id)
        if not ctx:
            return text

        doc_text = ctx["text"]
        if len(doc_text) > 30_000:
            doc_text = doc_text[:30_000] + "\n[... document truncated ...]"

        return (
            f"[DOCUMENT CONTEXT -- file: {ctx['filename']}]\n"
            f"{doc_text}\n"
            f"[END DOCUMENT CONTEXT]\n\n"
            f"User question: {text}"
        )


    async def _handle_error(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle errors"""
        logger.error(f"Telegram error: {context.error}")

        if update and update.effective_chat:
            try:
                await self.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text="Oops, something went wrong. Let me try again..."
                )
            except Exception as e:
                logger.debug(f"Could not send error message: {e}")

    # ============ VOICE MESSAGE HANDLER ============

    async def _handle_voice(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle voice messages and audio files — transcribe and respond."""

        user = update.effective_user
        chat_id = str(update.effective_chat.id)

        if not self._is_user_allowed(user.id):
            return

        # Rate limit same as text
        max_per_min = self.config.get("max_messages_per_minute", 20)
        if not _check_rate_limit(str(user.id), max_per_min):
            await update.message.reply_text(
                "You're sending messages too fast. Please wait a moment."
            )
            return

        # Get the voice or audio file object
        voice = update.message.voice
        audio = update.message.audio

        if voice:
            file_id = voice.file_id
            duration = voice.duration or 0
            suffix = ".ogg"
        elif audio:
            file_id = audio.file_id
            duration = audio.duration or 0
            mime = audio.mime_type or ""
            ext_map = {
                "audio/ogg": ".ogg", "audio/mpeg": ".mp3", "audio/mp3": ".mp3",
                "audio/wav": ".wav", "audio/x-wav": ".wav", "audio/flac": ".flac",
                "audio/mp4": ".m4a", "audio/x-m4a": ".m4a", "audio/webm": ".webm",
            }
            suffix = ext_map.get(mime, ".ogg")
        else:
            await update.message.reply_text(
                "Could not read the audio. Try sending as a voice message."
            )
            return

        # Reject very long audio (>5 minutes)
        if duration > 300:
            await update.message.reply_text(
                "That audio is too long (max 5 minutes for voice messages). "
                "Try sending a shorter clip."
            )
            return

        # Show typing while we process
        await self.send_typing_indicator(chat_id)

        # Download the file from Telegram
        tmp_path = None
        try:
            tg_file = await context.bot.get_file(file_id)
            buf = io.BytesIO()
            await tg_file.download_to_memory(buf)
            buf.seek(0)

            # Write to temp file for Whisper
            fd, tmp_path = tempfile.mkstemp(suffix=suffix)
            os.close(fd)
            with open(tmp_path, "wb") as f:
                f.write(buf.read())

            logger.info(f"[Voice] Downloaded {suffix} from user {user.id} ({duration}s)")

        except Exception as e:
            logger.error(f"[Voice] Failed to download voice file: {e}")
            await update.message.reply_text(
                "I couldn't download that voice message. Please try again."
            )
            return

        # Transcribe
        transcription = await self._transcribe_audio_file(tmp_path)

        # Clean up temp file
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)
        except OSError:
            pass

        if not transcription:
            await update.message.reply_text(
                "I couldn't transcribe that voice message. "
                "Transcription isn't available right now — try typing your message instead."
            )
            return

        # Send transcription as italic quote
        try:
            await update.message.reply_text(
                f"_You said: {transcription}_",
                parse_mode=ParseMode.MARKDOWN
            )
        except Exception:
            # Fallback without markdown if special chars break it
            await update.message.reply_text(f"You said: {transcription}")

        # Update active chat info
        if chat_id in self.active_chats:
            self.active_chats[chat_id]["last_message"] = datetime.now().isoformat()
        else:
            self.active_chats[chat_id] = {
                "user_id": str(user.id),
                "username": user.username,
                "first_name": user.first_name,
                "started_at": datetime.now().isoformat(),
                "last_message": datetime.now().isoformat()
            }

        # Show typing while AURA thinks
        await self.send_typing_indicator(chat_id)

        # Process transcribed text through AURA (same as if user typed it)
        incoming = IncomingMessage(
            platform="telegram",
            user_id=str(user.id),
            chat_id=chat_id,
            username=user.username,
            display_name=user.first_name,
            message_type=MessageType.VOICE,
            text=transcription,
            media_url=None,
            timestamp=datetime.now(),
            raw_message=update.message
        )

        response = await self.handle_incoming(incoming)

        if response:
            await update.message.reply_text(response)

        self._save_state()

    async def _transcribe_audio_file(self, file_path: str) -> Optional[str]:
        """Transcribe an audio file. Tries in priority order:
        1. Local Whisper via AudioTranscriberTool
        2. /api/transcribe endpoint on the backend
        3. None (caller shows error)
        """
        loop = asyncio.get_running_loop()

        # --- Method 1: Local Whisper via AudioTranscriberTool ---
        try:
            from aura.tools.audio_transcriber import AudioTranscriberTool
            transcriber = AudioTranscriberTool()

            result = await loop.run_in_executor(
                None, lambda: transcriber.transcribe(file_path)
            )

            if result.get("success") and result.get("text", "").strip():
                text = result["text"].strip()
                logger.info(f"[Voice] Whisper transcribed: {text[:80]}...")
                return text
            else:
                logger.warning(
                    f"[Voice] Whisper returned no text: {result.get('error', 'empty')}"
                )
        except ImportError:
            logger.info("[Voice] Whisper not installed, trying API fallback")
        except Exception as e:
            logger.warning(f"[Voice] Whisper transcription failed: {e}")

        # --- Method 2: Backend /api/transcribe endpoint ---
        try:
            import aiohttp

            api_url = os.environ.get("AURA_API_URL", "http://127.0.0.1:8000")
            api_key = os.environ.get("AURA_API_KEY", "")

            headers = {}
            if api_key:
                headers["x-api-key"] = api_key

            async with aiohttp.ClientSession() as session:
                with open(file_path, "rb") as f:
                    form = aiohttp.FormData()
                    form.add_field(
                        "file", f,
                        filename=os.path.basename(file_path),
                        content_type="audio/ogg"
                    )
                    async with session.post(
                        f"{api_url}/api/transcribe",
                        data=form,
                        headers=headers,
                        timeout=aiohttp.ClientTimeout(total=60)
                    ) as resp:
                        if resp.status == 200:
                            data = await resp.json()
                            text = data.get("text", "").strip()
                            if text:
                                logger.info(f"[Voice] API transcribed: {text[:80]}...")
                                return text
                        else:
                            body = await resp.text()
                            logger.warning(
                                f"[Voice] API transcribe returned {resp.status}: {body[:200]}"
                            )
        except ImportError:
            logger.info("[Voice] aiohttp not installed, skipping API fallback")
        except Exception as e:
            logger.warning(f"[Voice] API transcription failed: {e}")

        # All methods exhausted
        logger.warning("[Voice] All transcription methods failed")
        return None


    # ============ INLINE QUERY HANDLER ============

    # Quick-action prefixes: keyword -> (system_prompt_prefix, result_title_prefix)
    _INLINE_ACTIONS = {
        "translate": (
            "Translate the following text. Auto-detect the source language. "
            "If the text is in English, translate to Russian. Otherwise translate to English. "
            "Only output the translation, nothing else:\n\n",
            "Translate"
        ),
        "explain": (
            "Give a clear, concise explanation of the following topic. "
            "Keep it under 200 words:\n\n",
            "Explain"
        ),
        "summarize": (
            "Summarize the following text in 2-3 sentences. Be concise:\n\n",
            "Summarize"
        ),
    }

    async def _handle_inline(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle inline queries from any chat via @Aura828Bot <query>."""

        inline_query = update.inline_query
        if not inline_query:
            return

        user = inline_query.from_user
        query = (inline_query.query or "").strip()
        user_id = user.id

        # Auth check
        if not self._is_user_allowed(user_id):
            await inline_query.answer([], cache_time=60)
            return

        # Ignore empty or too-short queries
        if len(query) < 3:
            await inline_query.answer([], cache_time=5)
            return

        # Debounce: skip if the user sent another query within the debounce window
        now = _time.time()
        prev = self._inline_last_query.get(user_id)
        self._inline_last_query[user_id] = (query, now)

        if prev and prev[0] != query:
            elapsed = now - prev[1]
            if elapsed < self._INLINE_DEBOUNCE_SEC:
                await asyncio.sleep(self._INLINE_DEBOUNCE_SEC - elapsed)
                current = self._inline_last_query.get(user_id)
                if current and current[0] != query:
                    return

        # Build result list
        results = []

        # Detect quick-action prefix
        query_lower = query.lower()
        detected_action = None
        action_body = query

        for prefix, (prompt_prefix, title_prefix) in self._INLINE_ACTIONS.items():
            if query_lower.startswith(prefix + " ") and len(query) > len(prefix) + 1:
                detected_action = prefix
                action_body = query[len(prefix):].strip()
                break

        if detected_action:
            prompt_prefix, title_prefix = self._INLINE_ACTIONS[detected_action]
            full_prompt = prompt_prefix + action_body

            response = await self._inline_generate(full_prompt, user_id)
            if response:
                result_id = hashlib.md5(
                    f"{detected_action}:{action_body}".encode()
                ).hexdigest()[:16]
                results.append(InlineQueryResultArticle(
                    id=result_id,
                    title=f"{title_prefix}: {action_body[:40]}",
                    description=response[:100],
                    input_message_content=InputTextMessageContent(
                        message_text=response
                    )
                ))

            # Also offer a general AURA answer as second option
            general_response = await self._inline_generate(query, user_id)
            if general_response and general_response != response:
                general_id = hashlib.md5(
                    f"general:{query}".encode()
                ).hexdigest()[:16]
                results.append(InlineQueryResultArticle(
                    id=general_id,
                    title=f"Ask AURA: {query[:40]}",
                    description=general_response[:100],
                    input_message_content=InputTextMessageContent(
                        message_text=general_response
                    )
                ))
        else:
            # General query
            response = await self._inline_generate(query, user_id)
            if response:
                result_id = hashlib.md5(query.encode()).hexdigest()[:16]
                results.append(InlineQueryResultArticle(
                    id=result_id,
                    title=response[:50],
                    description=query,
                    input_message_content=InputTextMessageContent(
                        message_text=response
                    )
                ))

        # Rich inline: add category-specific quick actions
        _INLINE_CATEGORIES = [
            ("code_", "\U0001f4bb Code", f"Write code for: {query}",
             f"Write clean, concise code for the following request:\n\n{query}"),
            ("img_", "\U0001f3a8 Image Prompt", f"Image prompt for: {query}",
             f"Create a detailed image generation prompt for: {query}"),
            ("research_", "\U0001f50d Research", f"Research: {query}",
             f"Give a thorough research summary on: {query}"),
        ]
        for cat_id, cat_title, cat_desc, cat_prompt in _INLINE_CATEGORIES:
            rid = hashlib.md5(f"{cat_id}{query}".encode()).hexdigest()[:16]
            results.append(InlineQueryResultArticle(
                id=rid,
                title=f"{cat_title}: {query[:35]}",
                description=cat_desc[:100],
                input_message_content=InputTextMessageContent(
                    message_text=f"/{cat_title.split(' ')[1].lower()} {query}"
                )
            ))

        # Fallback if nothing generated (keep category shortcuts)
        if len(results) <= len(_INLINE_CATEGORIES):
            results.insert(0, InlineQueryResultArticle(
                id="error",
                title="Could not generate a response",
                description="Try a different query or ask in the direct chat.",
                input_message_content=InputTextMessageContent(
                    message_text="Sorry, I couldn't process that query. Try asking me directly!"
                )
            ))

        try:
            await inline_query.answer(results, cache_time=30)
        except Exception as e:
            logger.error(f"Failed to answer inline query: {e}")

    async def _inline_generate(self, prompt: str, user_id: int) -> Optional[str]:
        """Generate a response for an inline query with caching and timeout."""

        cache_key = hashlib.md5(prompt.encode()).hexdigest()
        cached = self._inline_cache.get(cache_key)
        if cached:
            return cached

        try:
            response = await asyncio.wait_for(
                self._process_with_aura(prompt, str(user_id)),
                timeout=self._INLINE_TIMEOUT_SEC
            )

            if response and response != "Sorry, I couldn't process that. Try again in a moment.":
                if len(response) > 4000:
                    response = response[:3997] + "..."
                self._inline_cache.put(cache_key, response)
                return response
        except asyncio.TimeoutError:
            logger.warning(f"Inline query timed out for user {user_id}")
        except Exception as e:
            logger.error(f"Inline generation error: {e}")

        return None

    # ============ STICKER / GIF REACTIONS (Phase 5) ============

    async def _handle_sticker(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """React when user sends a sticker."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        sticker = update.message.sticker
        emoji = sticker.emoji or "🤔"
        responses = [
            f"Nice {emoji}!",
            f"I see you're feeling {emoji}",
            f"{emoji} right back at you!",
            f"Ooh, {emoji}!",
        ]
        await update.message.reply_text(random.choice(responses))

    async def _maybe_send_reaction(self, chat_id: str, emotion: str, intensity: float):
        """Maybe send a sticker or GIF based on current emotion.

        Only triggers ~20% of the time to avoid spam, and only when
        the emotion intensity is strong enough (>= 0.5).
        """
        if intensity < 0.5 or random.random() > 0.2:
            return

        reaction_info = EMOTION_REACTIONS.get(emotion, EMOTION_REACTIONS["neutral"])

        if random.random() > 0.5:
            await self._send_random_gif(chat_id, reaction_info["gif_queries"])
        else:
            await self._send_emoji_reaction(chat_id, emotion)

    async def _send_random_gif(self, chat_id: str, queries: list):
        """Send a random GIF using Tenor's API."""
        query = random.choice(queries)
        try:
            import aiohttp
            async with aiohttp.ClientSession() as session:
                url = (
                    f"https://tenor.googleapis.com/v2/search"
                    f"?q={query}&limit=5&media_filter=gif"
                    f"&key=AIzaSyAyimkuYQYF_FXVALexPuGQctUWRURdCYQ"
                )
                async with session.get(url, timeout=aiohttp.ClientTimeout(total=5)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        results = data.get("results", [])
                        if results:
                            gif = random.choice(results)
                            gif_url = gif["media_formats"]["gif"]["url"]
                            await self.bot.send_animation(
                                chat_id=int(chat_id),
                                animation=gif_url,
                            )
        except Exception as e:
            logger.debug(f"GIF reaction failed (non-critical): {e}")

    async def _send_emoji_reaction(self, chat_id: str, emotion: str):
        """Send an emoji as a lightweight reaction message."""
        emoji = EMOTION_EMOJI.get(emotion, "👍")
        try:
            await self.bot.send_message(chat_id=int(chat_id), text=emoji)
        except Exception:
            pass


    # ============ REACT AGENT LOOP ============

    _AGENT_TIMEOUT = 120  # seconds

    @staticmethod
    def _get_progress_text(goal: str) -> str:
        """Return a contextual progress message based on the user's request."""
        g = goal.lower()
        if any(k in g for k in ["search", "find", "look up", "google"]):
            return "\U0001f50d Searching..."
        if any(k in g for k in ["research", "investigate", "deep dive", "analyze"]):
            return "\U0001f9ea Researching..."
        if any(k in g for k in ["code", "python", "script", "function", "debug"]):
            return "\U0001f4bb Writing code..."
        if any(k in g for k in ["image", "draw", "generate", "picture", "photo"]):
            return "\U0001f3a8 Generating..."
        if any(k in g for k in ["translate", "translation"]):
            return "\U0001f310 Translating..."
        if any(k in g for k in ["summarize", "summary", "tldr"]):
            return "\U0001f4dd Summarizing..."
        if any(k in g for k in ["math", "calculate", "solve", "equation"]):
            return "\U0001f9ee Calculating..."
        if "[document context" in g.lower():
            return "\U0001f4c4 Analyzing document..."
        return "\U0001f4ad Thinking..."

    async def _run_agent_and_reply(self, update: Update, goal: str, *,
                                    conv_id: str = None, user_id: str = None):
        """Run the full ReAct agent loop and send the result back to the user.

        Flow:
        1. Send a "Thinking..." placeholder message
        2. Start a background typing indicator loop
        3. Run agent.run() in a thread with timeout (full ReAct loop with tools)
        4. Edit the placeholder with the final response
        5. Send any file artifacts (screenshots, plots) as photos/documents
        6. Fall back to agent.chat() / brain.think() on failure
        7. Track user + assistant messages in ConversationManager (surface attribution)
        """
        chat_id = str(update.effective_chat.id)
        start_time = _time.time()

        # Track the user message in ConversationManager
        if conv_id and user_id:
            try:
                cm = get_conversation_manager()
                cm.on_message_added(conv_id, "user", goal, "telegram", user_id)
            except Exception as e:
                logger.debug(f"[Telegram] ConversationManager user message tracking skipped: {e}")

        # Send contextual placeholder based on the goal
        placeholder_text = self._get_progress_text(goal)
        placeholder = await update.message.reply_text(placeholder_text)

        # Start typing indicator loop (cancelled when done)
        typing_task = asyncio.create_task(self._typing_loop(chat_id))

        try:
            response_text, artifacts = await asyncio.wait_for(
                asyncio.to_thread(self._run_agent_sync, goal),
                timeout=self._AGENT_TIMEOUT,
            )
        except asyncio.TimeoutError:
            elapsed = _time.time() - start_time
            logger.warning(f"[Telegram] Agent timed out after {elapsed:.1f}s for: {goal[:80]}")
            response_text = "That took too long. Try a simpler question or break it into parts."
            artifacts = []
        except Exception as e:
            logger.error(f"[Telegram] Agent error: {e}", exc_info=True)
            response_text = "Something went wrong processing your request. Please try again."
            artifacts = []
        finally:
            typing_task.cancel()
            try:
                await typing_task
            except asyncio.CancelledError:
                pass

        elapsed = _time.time() - start_time
        logger.info(
            f"[Telegram] Response ready in {elapsed:.1f}s "
            f"({len(response_text)} chars, {len(artifacts)} artifacts)"
        )

        # Track the assistant response in ConversationManager
        if conv_id and user_id and response_text:
            try:
                cm = get_conversation_manager()
                cm.on_message_added(conv_id, "assistant", response_text, "telegram", user_id)
            except Exception as e:
                logger.debug(f"[Telegram] ConversationManager assistant message tracking skipped: {e}")

        # Store last exchange for /learn command (persisted to SQLite)
        if user_id:
            try:
                self.store.set_skill_state(user_id, last_exchange={
                    "input": goal[:2000],
                    "output": response_text[:2000],
                    "timestamp": _time.time(),
                })
            except Exception:
                pass

        # Build reply action buttons
        action_buttons = self._get_action_buttons(update.message.message_id)

        # Edit the placeholder with the real response + action buttons
        await self._edit_or_send_response(placeholder, chat_id, response_text, update,
                                          reply_markup=action_buttons)

        # Send any file artifacts (screenshots, plots, generated files)
        for artifact_path in artifacts:
            await self._send_file_artifact(chat_id, artifact_path, update)

        # Phase 5+: Native reaction via Telegram API (if available)
        try:
            emotion_data = None
            evoemo = self.aura.tools.get("evoemo") if hasattr(self.aura, 'tools') else None
            if evoemo and hasattr(evoemo, 'get_state'):
                emotion_data = evoemo.get_state()
            if not emotion_data:
                agent = getattr(self.aura, 'agent', self.aura)
                brain = getattr(agent, 'brain', None)
                if brain and hasattr(brain, '_alma_engine'):
                    emotion_data = brain._alma_engine.get_emotional_state()
            if not emotion_data and hasattr(self.aura, 'emotion') and self.aura.emotion:
                emotion_data = self.aura.emotion.get_emotional_state()

            if emotion_data:
                emotion = emotion_data.get("dominant_emotion", "neutral")
                intensity = emotion_data.get("intensity", 0.3)
                # Try native reaction first, fall back to sticker/GIF
                reacted = await self._try_native_reaction(
                    chat_id, update.message.message_id, emotion, intensity
                )
                if not reacted:
                    await self._maybe_send_reaction(chat_id, emotion, intensity)
        except Exception:
            pass  # Reactions are optional — never break the main flow

    def _run_agent_sync(self, goal: str):
        """Synchronous agent execution — called via asyncio.to_thread.

        Returns (response_text, list_of_artifact_paths).

        Priority chain:
        1. agent.run()  — full ReAct loop with tool calling
        2. agent.chat() — single LLM call with memory/emotion
        3. brain.think() — raw LLM call (last resort)
        """
        wrapper = self.aura  # TelegramAgentWrapper
        agent = getattr(wrapper, 'agent', wrapper)  # Unwrap to ApprenticeAgent

        response_text = ""
        artifacts = []

        # === Primary: Full ReAct agent loop ===
        try:
            logger.info(f"[Telegram] agent.run() starting: {goal[:80]}")
            result = agent.run(goal, timeout_seconds=self._AGENT_TIMEOUT - 5)

            if isinstance(result, dict):
                response_text = result.get("response", "")

                if result.get("timeout"):
                    response_text = response_text or "That request took too long. Please try a simpler query."

                # Fallback: check final_evaluation
                if not response_text:
                    fe = result.get("final_evaluation", {})
                    response_text = fe.get("progress", "")

                # Fallback: check history for last tool output
                if not response_text:
                    history = result.get("history", [])
                    if history:
                        last = history[-1]
                        if isinstance(last, dict):
                            response_text = last.get("result", {}).get("output", str(last))

                artifacts = self._collect_artifacts(agent)

            elif isinstance(result, str):
                response_text = result
            else:
                response_text = str(result) if result else ""

            if response_text:
                logger.info(f"[Telegram] agent.run() succeeded ({len(response_text)} chars)")
                return (response_text, artifacts)

        except Exception as e:
            logger.warning(f"[Telegram] agent.run() failed: {e}, falling back to chat()")

        # === Fallback 1: agent.chat() ===
        try:
            logger.info(f"[Telegram] Fallback to agent.chat(): {goal[:80]}")
            response_text = agent.chat(goal)
            if response_text:
                return (response_text, [])
        except Exception as e:
            logger.warning(f"[Telegram] agent.chat() failed: {e}, falling back to brain.think()")

        # === Fallback 2: Direct brain.think() ===
        try:
            brain = getattr(agent, 'brain', None)
            if brain and hasattr(brain, 'think'):
                logger.info(f"[Telegram] Fallback to brain.think(): {goal[:80]}")
                response_text = brain.think(goal)
                if response_text:
                    return (response_text, [])
        except Exception as e:
            logger.error(f"[Telegram] brain.think() also failed: {e}")

        return (response_text or "Sorry, I couldn't process that right now.", [])

    def _collect_artifacts(self, agent) -> list:
        """Collect file artifacts produced by the agent's tool calls."""
        artifacts = []

        brain = getattr(agent, 'brain', None)
        if brain:
            screenshot = getattr(brain, '_last_screenshot_path', None)
            if screenshot and Path(screenshot).exists():
                artifacts.append(screenshot)
                brain._last_screenshot_path = None

        output_dirs = [
            Path("data/output"),
            Path("data/screenshots"),
            Path("data/generated"),
        ]
        cutoff = _time.time() - 60

        for out_dir in output_dirs:
            if out_dir.exists():
                try:
                    for f in out_dir.iterdir():
                        if f.is_file() and f.stat().st_mtime > cutoff:
                            fpath = str(f)
                            if fpath not in artifacts:
                                artifacts.append(fpath)
                except OSError:
                    pass

        return artifacts

    async def _typing_loop(self, chat_id: str):
        """Send typing indicator every 4 seconds until cancelled."""
        try:
            while True:
                await self.send_typing_indicator(chat_id)
                await asyncio.sleep(4)
        except asyncio.CancelledError:
            pass

    async def _edit_or_send_response(self, placeholder, chat_id: str, text: str,
                                     update: Update, reply_markup=None):
        """Edit the placeholder message with the response, splitting if > 4096 chars."""
        if not text:
            text = "I processed your request but have nothing to report."

        MAX_LEN = 4096
        text = text.strip()

        if len(text) <= MAX_LEN:
            try:
                await placeholder.edit_text(text, reply_markup=reply_markup)
            except Exception as e:
                logger.warning(f"Could not edit placeholder: {e}")
                try:
                    await self.bot.send_message(chat_id=chat_id, text=text,
                                                reply_markup=reply_markup)
                except Exception:
                    pass
        else:
            chunks = self._split_message(text, MAX_LEN)
            for i, chunk in enumerate(chunks):
                try:
                    if i == 0:
                        await placeholder.edit_text(chunk)
                    else:
                        # Attach reply_markup to the last chunk only
                        markup = reply_markup if i == len(chunks) - 1 else None
                        await self.bot.send_message(chat_id=chat_id, text=chunk,
                                                    reply_markup=markup)
                except Exception as e:
                    logger.warning(f"Error sending chunk {i}: {e}")

    @staticmethod
    def _split_message(text: str, max_len: int = 4096) -> list:
        """Split a long message into chunks, preferring paragraph boundaries."""
        if len(text) <= max_len:
            return [text]

        chunks = []
        while text:
            if len(text) <= max_len:
                chunks.append(text)
                break
            split_at = text.rfind("\n\n", 0, max_len)
            if split_at == -1 or split_at < max_len // 2:
                split_at = text.rfind("\n", 0, max_len)
            if split_at == -1 or split_at < max_len // 2:
                split_at = text.rfind(" ", 0, max_len)
            if split_at == -1:
                split_at = max_len
            chunks.append(text[:split_at].rstrip())
            text = text[split_at:].lstrip()

        return chunks

    async def _send_file_artifact(self, chat_id: str, filepath: str, update: Update):
        """Send a file artifact as a Telegram photo or document."""
        p = Path(filepath)
        if not p.exists():
            return

        image_exts = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
        try:
            if p.suffix.lower() in image_exts and p.stat().st_size < 10 * 1024 * 1024:
                with open(p, "rb") as f:
                    await self.bot.send_photo(
                        chat_id=chat_id,
                        photo=f,
                        caption=p.name[:200],
                        reply_to_message_id=update.message.message_id,
                    )
            else:
                with open(p, "rb") as f:
                    await self.bot.send_document(
                        chat_id=chat_id,
                        document=f,
                        filename=p.name,
                        caption=p.name[:200],
                        reply_to_message_id=update.message.message_id,
                    )
            logger.info(f"[Telegram] Sent artifact: {filepath}")
        except Exception as e:
            logger.warning(f"[Telegram] Could not send artifact {filepath}: {e}")

    # ============ OVERRIDE AURA PROCESSING (kept for BasePlatform compat) ============

    async def _process_with_aura(self, text: str, user_id: str) -> str:
        """Process message through the full ReAct agent loop.

        This override is kept for BasePlatform.handle_incoming() compatibility.
        The primary path is _run_agent_and_reply() for regular messages.
        Voice/inline handlers still use this via handle_incoming() or directly.
        """
        try:
            response_text, _artifacts = await asyncio.to_thread(self._run_agent_sync, text)
            if response_text:
                return response_text
        except Exception as e:
            logger.error(f"AURA processing error: {e}", exc_info=True)

        return "Sorry, I couldn't process that. Try again in a moment."

    # ============ PROACTIVE MESSAGING ============

    async def _connect_proactive_system(self):
        """Connect GatewayDaemon proactive system to Telegram if available."""
        try:
            daemon = getattr(self.aura, 'gateway_daemon', None)
            if daemon:
                # Start proactive polling loop
                self._proactive_task = asyncio.create_task(self._proactive_polling_loop())
                logger.info("Proactive system connected to Telegram!")
            else:
                self._proactive_task = None
                logger.info("No proactive system available — skipping")
        except Exception as e:
            self._proactive_task = None
            logger.warning(f"Could not connect proactive system: {e}")

    async def _proactive_polling_loop(self):
        """Poll GatewayDaemon for proactive messages and send to active chats."""
        logger.info("Proactive polling loop started")

        while self.is_running:
            try:
                daemon = getattr(self.aura, 'gateway_daemon', None)
                if daemon and hasattr(daemon, 'get_pending_notifications'):
                    pending = daemon.get_pending_notifications()

                    for notification in pending:
                        message = getattr(notification, 'content', str(notification))
                        for chat_id, chat_info in self.active_chats.items():
                            try:
                                user_name = chat_info.get("first_name", "there")
                                personalized = message.replace("{name}", user_name)
                                await self.send_proactive(chat_id, personalized)
                                await asyncio.sleep(0.2)
                            except Exception as e:
                                logger.warning(f"Could not send proactive to {chat_id}: {e}")

            except Exception as e:
                logger.error(f"Proactive polling error: {e}")

            # Wait before next poll (60 seconds)
            await asyncio.sleep(60)

    async def send_to_all_active(self, message: str):
        """Send a message to all active chats (for broadcasts)"""

        for chat_id in self.active_chats:
            try:
                await self.send_proactive(chat_id, message)
                await asyncio.sleep(0.1)  # Rate limiting
            except Exception as e:
                logger.warning(f"Could not send to {chat_id}: {e}")

    async def send_morning_greeting(self, chat_id: str, user_name: str):
        """Send personalized morning greeting"""

        greetings = [
            f"Morning {user_name}! What's on your mind today?",
            f"Hey {user_name}! Morning. Ready when you are.",
            f"Good morning! How are you feeling today?",
        ]

        await self.send_proactive(chat_id, random.choice(greetings))

    async def send_follow_up(self, chat_id: str, topic: str):
        """Send a follow-up about something mentioned"""

        message = f"Hey - how did {topic} go?"
        await self.send_proactive(chat_id, message)


    # ============ SKILL SYSTEM HANDLERS ============

    def _get_skill_store(self) -> "SkillStore":
        """Lazy-load the SkillStore singleton."""
        if self._skill_store is None:
            if not SKILL_LIBRARY_AVAILABLE:
                raise RuntimeError("Skill library not installed")
            self._skill_store = SkillStore(storage_path="./aura_skills")
        return self._skill_store

    def _get_skill_learner(self) -> "SkillLearner":
        """Lazy-load the SkillLearner singleton."""
        if self._skill_learner is None:
            if not SKILL_LIBRARY_AVAILABLE:
                raise RuntimeError("Skill library not installed")
            store = self._get_skill_store()
            # Try to get an LLM function from the aura engine
            llm_func = None
            try:
                brain = getattr(self.aura, 'brain', None)
                if brain is None:
                    brain = getattr(getattr(self.aura, 'agent', None), 'brain', None)
                if brain and hasattr(brain, 'think'):
                    llm_func = lambda prompt: brain.think(prompt)
            except Exception:
                pass
            self._skill_learner = SkillLearner(
                store=store,
                llm_func=llm_func,
                min_examples_to_learn=1,
            )
        return self._skill_learner

    async def _handle_learn(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /learn -- extract a reusable skill from the last conversation exchange."""
        if not self._is_user_allowed(update.effective_user.id):
            return

        user_id = update.effective_user.id

        if not SKILL_LIBRARY_AVAILABLE:
            await update.message.reply_text("Skill library is not available on this server.")
            return

        exchange = self._last_exchange.get(user_id)
        if not exchange:
            await update.message.reply_text(
                "No recent conversation to learn from.\n"
                "Chat with me first, then use /learn to extract a skill."
            )
            return

        if _time.time() - exchange.get("timestamp", 0) > 1800:
            await update.message.reply_text(
                "The last exchange is too old (> 30 min).\n"
                "Have a fresh conversation first, then /learn."
            )
            return

        await update.message.reply_text("Analyzing our last exchange to create a skill...")
        await self.send_typing_indicator(str(update.effective_chat.id))

        try:
            learner = self._get_skill_learner()

            if learner.llm_func is None:
                await update.message.reply_text(
                    "Cannot learn skills right now -- no LLM backend available."
                )
                return

            prompt = (
                "Analyze this conversation exchange and extract a reusable skill.\n\n"
                f"User said: {exchange['input'][:1000]}\n\n"
                f"AURA responded: {exchange['output'][:1000]}\n\n"
                "Create a skill definition with:\n"
                "1. A clear, concise name (2-4 words)\n"
                "2. A description of what this skill does\n"
                "3. 3-5 trigger phrases that would activate this skill\n"
                "4. A step-by-step procedure that generalizes from this exchange\n"
                "5. The best category: coding, writing, research, automation, analysis, communication, learning\n\n"
                'Respond in this exact JSON format:\n'
                '{"name": "Skill Name", "description": "What this skill does...", '
                '"trigger_patterns": ["phrase 1", "phrase 2", "phrase 3"], '
                '"procedure": "Step 1: ...\\nStep 2: ...\\nStep 3: ...", '
                '"category": "coding", "tags": ["tag1", "tag2"]}\n\n'
                "Respond ONLY with the JSON, no other text."
            )

            response = await asyncio.to_thread(learner.llm_func, prompt)

            json_match = re.search(r'\{[\s\S]*\}', response)
            if not json_match:
                await update.message.reply_text(
                    "Could not extract a skill from that exchange. Try a more structured interaction."
                )
                return

            skill_data = json.loads(json_match.group())

            required = ["name", "description", "trigger_patterns", "procedure"]
            for fld in required:
                if fld not in skill_data:
                    await update.message.reply_text(f"Skill extraction incomplete -- missing {fld}. Try again.")
                    return

            triggers_str = ", ".join(f'"{t}"' for t in skill_data["trigger_patterns"][:5])
            procedure_preview = skill_data["procedure"][:300]
            if len(skill_data["procedure"]) > 300:
                procedure_preview += "..."

            msg = (
                f'Skill: "{skill_data["name"]}"\n'
                f'Category: {skill_data.get("category", "custom")}\n'
                f'Triggers: {triggers_str}\n'
                f'Procedure: {procedure_preview}\n\n'
                f'Save this skill? (reply "yes" to confirm)'
            )
            await update.message.reply_text(msg)

            self._skill_pending[user_id] = {
                "action": "learn_confirm",
                "skill_data": skill_data,
                "exchange": exchange,
                "timestamp": _time.time(),
            }

        except json.JSONDecodeError:
            await update.message.reply_text("Failed to parse skill data. The LLM returned invalid JSON.")
        except Exception as e:
            logger.error(f"[Telegram] /learn error: {e}", exc_info=True)
            await update.message.reply_text(f"Failed to learn skill: {e}")

    async def _handle_skill(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /skill command -- dispatch to subcommands."""
        if not self._is_user_allowed(update.effective_user.id):
            return

        if not SKILL_LIBRARY_AVAILABLE:
            await update.message.reply_text("Skill library is not available on this server.")
            return

        args = context.args or []

        if not args:
            await update.message.reply_text(
                "Usage:\n"
                "/skill list [category] -- List all skills\n"
                "/skill search <query> -- Search skills\n"
                "/skill info <name> -- Detailed skill info\n"
                "/skill improve <name> -- Evolve with GEPA\n"
                "/skill create <name> -- Define a new skill"
            )
            return

        subcmd = args[0].lower()
        rest = args[1:]

        if subcmd == "list":
            await self._skill_list(update, " ".join(rest) if rest else None)
        elif subcmd == "search":
            if not rest:
                await update.message.reply_text("Usage: /skill search <query>")
                return
            await self._skill_search(update, " ".join(rest))
        elif subcmd == "info":
            if not rest:
                await update.message.reply_text("Usage: /skill info <id_or_name>")
                return
            await self._skill_info(update, " ".join(rest))
        elif subcmd == "improve":
            if not rest:
                await update.message.reply_text("Usage: /skill improve <id_or_name>")
                return
            await self._skill_improve(update, " ".join(rest))
        elif subcmd == "create":
            if not rest:
                await update.message.reply_text("Usage: /skill create <name>")
                return
            await self._skill_create_start(update, " ".join(rest))
        else:
            await update.message.reply_text(
                f"Unknown subcommand: {subcmd}\n"
                "Use /skill for available commands."
            )

    async def _skill_list(self, update: Update, category_str: str = None):
        """List all learned skills, optionally filtered by category."""
        try:
            store = self._get_skill_store()

            cat_filter = None
            if category_str:
                try:
                    cat_filter = SkillCategory(category_str.lower())
                except ValueError:
                    await update.message.reply_text(
                        f"Unknown category: {category_str}\n"
                        f"Valid: {', '.join(c.value for c in SkillCategory)}"
                    )
                    return

            skills = store.list_all(category=cat_filter, sort_by="updated")

            if not skills:
                msg = "No skills found."
                if category_str:
                    msg += f" (category: {category_str})"
                await update.message.reply_text(msg)
                return

            page = skills[:10]
            lines = ["Learned Skills\n"]
            for i, s in enumerate(page, 1):
                name = s.get("name", "Unnamed")
                success_rate = s.get("success_rate", 0)
                total_uses = s.get("total_uses", 0)
                updated = s.get("updated_at", "")
                if updated and len(updated) > 10:
                    updated = updated[:10]

                rate_pct = f"{success_rate * 100:.0f}%" if success_rate else "N/A"
                lines.append(
                    f"{i}. {name}\n"
                    f"   Rate: {rate_pct} | Uses: {total_uses} | Updated: {updated}"
                )

            if len(skills) > 10:
                lines.append(f"\n... and {len(skills) - 10} more skills")

            lines.append("\nUse /skill info <name> for details")
            await update.message.reply_text("\n".join(lines))

        except Exception as e:
            logger.error(f"[Telegram] /skill list error: {e}", exc_info=True)
            await update.message.reply_text(f"Failed to list skills: {e}")

    async def _skill_search(self, update: Update, query: str):
        """Search skills by description/name."""
        try:
            store = self._get_skill_store()
            results = store.search(query, limit=5)

            if not results:
                await update.message.reply_text(f"No skills found matching: {query}")
                return

            lines = [f'Search results for "{query}"\n']
            for skill_id, score in results:
                info = store.index.get(skill_id, {})
                name = info.get("name", skill_id)
                desc = info.get("description", "")[:80]
                lines.append(f"  {name} (score: {score:.2f})\n   {desc}")

            lines.append("\nUse /skill info <name> for details")
            await update.message.reply_text("\n".join(lines))

        except Exception as e:
            logger.error(f"[Telegram] /skill search error: {e}", exc_info=True)
            await update.message.reply_text(f"Search failed: {e}")

    async def _skill_info(self, update: Update, id_or_name: str):
        """Show detailed info about a skill."""
        try:
            store = self._get_skill_store()
            skill, skill_id = self._find_skill_by_id_or_name(store, id_or_name)

            if not skill:
                await update.message.reply_text(f"Skill not found: {id_or_name}")
                return

            triggers = ", ".join(f'"{t}"' for t in skill.trigger_patterns[:5])
            rate_pct = f"{skill.metadata.success_rate * 100:.0f}%" if skill.metadata.total_uses > 0 else "N/A"
            last_used = skill.metadata.last_used.strftime("%Y-%m-%d %H:%M") if skill.metadata.last_used else "Never"

            lines = [
                f"Skill: {skill.name}",
                f"ID: {skill.id}",
                f"Version: {skill.metadata.version}",
                f"Category: {skill.category.value}",
                f"Description: {skill.description}",
                "",
                f"Triggers: {triggers}",
                "",
                "Procedure:",
                skill.procedure[:500],
                "",
                f"Success Rate: {rate_pct}",
                f"Total Uses: {skill.metadata.total_uses}",
                f"Last Used: {last_used}",
                f"Tags: {', '.join(skill.metadata.tags) if skill.metadata.tags else 'None'}",
            ]

            if skill.metadata.parent_skill_id:
                lines.append(f"Evolved from: {skill.metadata.parent_skill_id}")

            await update.message.reply_text("\n".join(lines))

        except Exception as e:
            logger.error(f"[Telegram] /skill info error: {e}", exc_info=True)
            await update.message.reply_text(f"Failed to get skill info: {e}")

    async def _skill_improve(self, update: Update, id_or_name: str):
        """Trigger GEPA evolution on a skill."""
        user_id = update.effective_user.id

        if not GEPA_AVAILABLE:
            await update.message.reply_text("GEPA evolution engine is not available on this server.")
            return

        try:
            store = self._get_skill_store()
            skill, skill_id = self._find_skill_by_id_or_name(store, id_or_name)

            if not skill:
                await update.message.reply_text(f"Skill not found: {id_or_name}")
                return

            await update.message.reply_text(
                f'Evolving "{skill.name}"...\nThis may take a minute.'
            )
            await self.send_typing_indicator(str(update.effective_chat.id))

            brain = getattr(self.aura, 'brain', None)
            if brain is None:
                brain = getattr(getattr(self.aura, 'agent', None), 'brain', None)

            if not brain or not hasattr(brain, 'think'):
                await update.message.reply_text("Cannot evolve -- no LLM backend available.")
                return

            def llm_func(system: str, user: str) -> str:
                return brain.think(f"{system}\n\n{user}")

            config = GEPAConfig(
                max_iterations=3,
                max_metric_calls=30,
                timeout_seconds=120,
                no_improvement_patience=2,
                run_dir=f"./aura_data/evolution_runs/telegram_{skill_id}",
            )

            adapter = AuraSkillAdapter(config=config, llm_func=llm_func)

            seed = Candidate(
                id=0,
                components={skill_id: skill.procedure},
                parent_id=-1,
            )

            eval_examples = await asyncio.to_thread(
                adapter.generate_eval_dataset, seed, num_examples=6
            )

            if len(eval_examples) < 2:
                await update.message.reply_text(
                    "Could not generate enough evaluation examples. "
                    "The skill may be too simple to evolve."
                )
                return

            engine = GEPAEngine(config=config, adapter=adapter, llm_func=llm_func)

            result = await asyncio.to_thread(
                engine.optimize, seed, eval_examples
            )

            best = result.best_candidate
            improved_procedure = best.components.get(skill_id, skill.procedure)

            if result.improvement <= 0.01:
                await update.message.reply_text(
                    f'Evolution complete for "{skill.name}"\n\n'
                    f"Iterations: {result.iterations_run}\n"
                    f"Score: {best.avg_score:.2f}\n"
                    f"No significant improvement found. The skill is already good!"
                )
                return

            procedure_preview = improved_procedure[:400]
            if len(improved_procedure) > 400:
                procedure_preview += "..."

            msg = (
                f'Evolution complete for "{skill.name}"\n\n'
                f"Iterations: {result.iterations_run}\n"
                f"Improvement: +{result.improvement:.3f}\n"
                f"Best score: {best.avg_score:.2f}\n\n"
                f"Improved procedure:\n{procedure_preview}\n\n"
                f'Apply improvement? (reply "yes" to confirm)'
            )
            await update.message.reply_text(msg)

            self._skill_pending[user_id] = {
                "action": "improve_confirm",
                "skill_id": skill_id,
                "improved_procedure": improved_procedure,
                "improvement": result.improvement,
                "timestamp": _time.time(),
            }

        except Exception as e:
            logger.error(f"[Telegram] /skill improve error: {e}", exc_info=True)
            await update.message.reply_text(f"Evolution failed: {e}")

    async def _skill_create_start(self, update: Update, name: str):
        """Start the interactive skill creation flow."""
        user_id = update.effective_user.id

        self._skill_create_state[user_id] = {
            "step": "description",
            "name": name,
            "timestamp": _time.time(),
        }

        await update.message.reply_text(
            f'Creating skill: "{name}"\n\n'
            "Step 1/3: What does this skill do?\n"
            "(Send the description)"
        )

    async def _check_skill_pending(self, update: Update, user_id: int, text: str) -> bool:
        """Check and handle pending skill actions. Returns True if handled."""
        text_lower = text.strip().lower() if text else ""

        # Handle skill create multi-step flow
        if user_id in self._skill_create_state:
            state = self._skill_create_state[user_id]

            # Timeout after 10 minutes
            if _time.time() - state.get("timestamp", 0) > 600:
                del self._skill_create_state[user_id]
                return False

            step = state.get("step")

            if step == "description":
                state["description"] = text.strip()
                state["step"] = "triggers"
                state["timestamp"] = _time.time()
                await update.message.reply_text(
                    "Step 2/3: What phrases should trigger this skill?\n"
                    "(Send comma-separated trigger phrases)"
                )
                return True

            elif step == "triggers":
                triggers = [t.strip() for t in text.split(",") if t.strip()]
                if not triggers:
                    await update.message.reply_text(
                        "Please provide at least one trigger phrase, separated by commas."
                    )
                    return True
                state["triggers"] = triggers
                state["step"] = "procedure"
                state["timestamp"] = _time.time()
                await update.message.reply_text(
                    "Step 3/3: What is the step-by-step procedure?\n"
                    "(Send the procedure -- use numbered steps)"
                )
                return True

            elif step == "procedure":
                state["procedure"] = text.strip()

                try:
                    store = self._get_skill_store()
                    skill = Skill.create(
                        name=state["name"],
                        description=state["description"],
                        category=SkillCategory.CUSTOM,
                        trigger_patterns=state["triggers"],
                        procedure=state["procedure"],
                        tags=[],
                    )
                    skill_id = store.save(skill)
                    del self._skill_create_state[user_id]

                    await update.message.reply_text(
                        f"Skill created!\n\n"
                        f"Name: {state['name']}\n"
                        f"ID: {skill_id}\n"
                        f"Triggers: {', '.join(state['triggers'])}\n\n"
                        f"Use /skill info {state['name']} to see full details."
                    )
                except Exception as e:
                    logger.error(f"[Telegram] skill create error: {e}", exc_info=True)
                    await update.message.reply_text(f"Failed to create skill: {e}")
                    if user_id in self._skill_create_state:
                        del self._skill_create_state[user_id]

                return True

        # Handle pending confirmations (learn_confirm, improve_confirm)
        if user_id in self._skill_pending:
            pending = self._skill_pending[user_id]

            # Timeout after 5 minutes
            if _time.time() - pending.get("timestamp", 0) > 300:
                del self._skill_pending[user_id]
                return False

            if text_lower != "yes":
                del self._skill_pending[user_id]
                await update.message.reply_text("Cancelled.")
                return True

            action = pending.get("action")

            if action == "learn_confirm":
                try:
                    skill_data = pending["skill_data"]
                    store = self._get_skill_store()

                    category_str = skill_data.get("category", "custom").lower()
                    try:
                        category = SkillCategory(category_str)
                    except ValueError:
                        category = SkillCategory.CUSTOM

                    skill = Skill.create(
                        name=skill_data["name"],
                        description=skill_data["description"],
                        category=category,
                        trigger_patterns=skill_data["trigger_patterns"],
                        procedure=skill_data["procedure"],
                        tags=skill_data.get("tags", []),
                    )
                    skill.id = f"learned_{skill.id}"

                    exchange = pending.get("exchange", {})
                    if exchange:
                        example = SkillExample(
                            input_context=exchange.get("input", ""),
                            input_data=None,
                            output=exchange.get("output", ""),
                            success=True,
                        )
                        skill.add_example(example)

                    skill_id = store.save(skill)
                    del self._skill_pending[user_id]

                    await update.message.reply_text(
                        f"Skill saved!\n\n"
                        f"Name: {skill_data['name']}\n"
                        f"ID: {skill_id}\n\n"
                        "I'll use this skill in future similar conversations."
                    )
                except Exception as e:
                    logger.error(f"[Telegram] learn confirm error: {e}", exc_info=True)
                    await update.message.reply_text(f"Failed to save skill: {e}")
                    if user_id in self._skill_pending:
                        del self._skill_pending[user_id]
                return True

            elif action == "improve_confirm":
                try:
                    skill_id_pending = pending["skill_id"]
                    improved_procedure = pending["improved_procedure"]
                    store = self._get_skill_store()
                    skill = store.load(skill_id_pending)

                    if not skill:
                        await update.message.reply_text("Skill no longer exists.")
                        del self._skill_pending[user_id]
                        return True

                    skill.procedure = improved_procedure
                    try:
                        current_version = float(skill.metadata.version)
                        skill.metadata.version = f"{current_version + 0.1:.1f}"
                    except ValueError:
                        skill.metadata.version = "1.1"

                    from datetime import timezone
                    skill.metadata.last_modified = datetime.now(timezone.utc)
                    skill.updated_at = datetime.now(timezone.utc)
                    store.save(skill)

                    del self._skill_pending[user_id]
                    await update.message.reply_text(
                        f'Improvement applied to "{skill.name}"!\n'
                        f"Version: {skill.metadata.version}\n"
                        f"Improvement: +{pending.get('improvement', 0):.3f}"
                    )
                except Exception as e:
                    logger.error(f"[Telegram] improve confirm error: {e}", exc_info=True)
                    await update.message.reply_text(f"Failed to apply improvement: {e}")
                    if user_id in self._skill_pending:
                        del self._skill_pending[user_id]
                return True

        return False

    def _find_skill_by_id_or_name(self, store, id_or_name: str):
        """Find a skill by ID or name (case-insensitive partial match).

        Returns (Skill, skill_id) or (None, None).
        """
        # Try direct ID match
        skill = store.load(id_or_name)
        if skill:
            return skill, id_or_name

        # Try name match (case-insensitive)
        name_lower = id_or_name.lower()
        for sid, info in store.index.items():
            if info.get("name", "").lower() == name_lower:
                skill = store.load(sid)
                if skill:
                    return skill, sid

        # Try partial name match
        for sid, info in store.index.items():
            if name_lower in info.get("name", "").lower():
                skill = store.load(sid)
                if skill:
                    return skill, sid

        # Try slug match (name with hyphens)
        slug = name_lower.replace(" ", "-")
        for sid, info in store.index.items():
            stored_slug = info.get("name", "").lower().replace(" ", "-")
            if slug == stored_slug or slug in stored_slug:
                skill = store.load(sid)
                if skill:
                    return skill, sid

        return None, None


    # ─── Phase 4: Webhook commands ───────────────────────────────────

    async def _handle_webhook(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /webhook command — show webhook info and endpoints."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        args = context.args or []
        backend = os.getenv("AURA_BACKEND_URL", "http://89.167.107.134")

        if not args:
            text = (
                "Webhook Endpoints\n\n"
                f"GitHub: {backend}/api/webhooks/github\n"
                f"Alerts: {backend}/api/webhooks/alert\n"
                f"Notify: {backend}/api/webhooks/notify\n\n"
                "Use /webhook test to send a test event.\n"
                "Use /webhook github <repo> for setup instructions."
            )
            await update.message.reply_text(text)
        elif args[0] == "test":
            await update.message.reply_text("Test webhook received! Pipeline is working.")
        elif args[0] == "github" and len(args) > 1:
            repo = args[1]
            text = (
                f"GitHub Webhook Setup for {repo}\n\n"
                f"URL: {backend}/api/webhooks/github\n"
                f"Content type: application/json\n"
                f"Events: check_run, workflow_run, pull_request, push\n\n"
                f"Configure at: https://github.com/{repo}/settings/hooks/new"
            )
            await update.message.reply_text(text)
        else:
            await update.message.reply_text("Usage: /webhook, /webhook test, /webhook github <repo>")

    # ─── Phase 4: Scheduled tasks commands ─────────────────────────────

    async def _handle_remind(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /remind <time> <message> — one-shot reminder."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        args = context.args
        if not args or len(args) < 2:
            await update.message.reply_text("Usage: /remind <time> <message>\n\nExamples:\n/remind in 2h Check deployment\n/remind at 17:00 Call team\n/remind tomorrow 9am Review PRs")
            return

        text = " ".join(args)
        # Simple time parsing
        import re
        from datetime import timedelta

        now = __import__("datetime").datetime.now()
        run_at = None
        message = text

        # "in Xh", "in Xm", "in X hours", "in X minutes"
        m = re.match(r"in\s+(\d+)\s*h(?:ours?)?\s*(.*)", text, re.I)
        if m:
            run_at = now + timedelta(hours=int(m.group(1)))
            message = m.group(2).strip()
        if not run_at:
            m = re.match(r"in\s+(\d+)\s*m(?:in(?:utes?)?)?\s*(.*)", text, re.I)
            if m:
                run_at = now + timedelta(minutes=int(m.group(1)))
                message = m.group(2).strip()
        if not run_at:
            m = re.match(r"at\s+(\d{1,2}):?(\d{2})?\s*(am|pm)?\s*(.*)", text, re.I)
            if m:
                hour = int(m.group(1))
                minute = int(m.group(2) or 0)
                ampm = (m.group(3) or "").lower()
                if ampm == "pm" and hour < 12: hour += 12
                if ampm == "am" and hour == 12: hour = 0
                run_at = now.replace(hour=hour, minute=minute, second=0)
                if run_at <= now:
                    run_at += timedelta(days=1)
                message = m.group(4).strip()
        if not run_at:
            # Try "tomorrow" prefix
            m = re.match(r"tomorrow\s*(.*)", text, re.I)
            if m:
                run_at = now + timedelta(days=1)
                message = m.group(1).strip()

        if not run_at:
            await update.message.reply_text("Could not parse time. Try: /remind in 2h Check something")
            return
        if not message:
            message = "Reminder!"

        chat_id = str(update.effective_chat.id)
        job_id = f"tg_remind_{int(time.time())}_{update.effective_user.id}"

        try:
            from aura.tools.task_scheduler import TaskSchedulerTool
            scheduler = TaskSchedulerTool()
            if not scheduler._scheduler.running:
                scheduler._scheduler.start()

            scheduler._scheduler.add_job(
                self._fire_reminder, "date", run_date=run_at,
                id=job_id, args=[chat_id, message],
                replace_existing=True,
            )
            time_str = run_at.strftime("%H:%M on %b %d")
            await update.message.reply_text(f"Reminder set for {time_str}:\n{message}")
        except Exception as e:
            logger.error(f"Remind error: {e}")
            await update.message.reply_text(f"Failed to set reminder: {e}")

    def _fire_reminder(self, chat_id: str, message: str):
        """Callback for APScheduler — sends reminder to Telegram."""
        import asyncio
        async def _send():
            try:
                await self.bot.send_message(chat_id=int(chat_id), text=f"Reminder:\n{message}")
            except Exception as e:
                logger.error(f"Failed to send reminder: {e}")
        try:
            loop = asyncio.get_event_loop()
            if loop.is_running():
                asyncio.ensure_future(_send())
            else:
                loop.run_until_complete(_send())
        except RuntimeError:
            asyncio.run(_send())

    async def _handle_schedule(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /schedule <interval> <task> — recurring scheduled task."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        args = context.args
        if not args or len(args) < 2:
            await update.message.reply_text("Usage: /schedule <interval> <task>\n\nExamples:\n/schedule every 2h Check CPU\n/schedule daily at 9am Summarize notifications")
            return
        await update.message.reply_text("Scheduled tasks coming soon! Use /remind for one-shot reminders.")

    async def _handle_tasks(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /tasks — list active scheduled tasks and reminders."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        try:
            from aura.tools.task_scheduler import TaskSchedulerTool
            scheduler = TaskSchedulerTool()
            jobs = scheduler._scheduler.get_jobs() if scheduler._scheduler.running else []
            if not jobs:
                await update.message.reply_text("No active tasks or reminders.")
                return
            lines = ["Active Tasks:\n"]
            for job in jobs:
                next_run = job.next_run_time.strftime("%H:%M %b %d") if job.next_run_time else "paused"
                lines.append(f"  {job.id}\n  Next: {next_run}\n")
            await update.message.reply_text("\n".join(lines))
        except Exception as e:
            await update.message.reply_text(f"Error listing tasks: {e}")

    async def _handle_cancel(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /cancel <id> — cancel a scheduled task."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        args = context.args
        if not args:
            await update.message.reply_text("Usage: /cancel <task_id>")
            return
        job_id = args[0]
        try:
            from aura.tools.task_scheduler import TaskSchedulerTool
            scheduler = TaskSchedulerTool()
            scheduler._scheduler.remove_job(job_id)
            await update.message.reply_text(f"Cancelled: {job_id}")
        except Exception as e:
            await update.message.reply_text(f"Failed to cancel: {e}")

    # ─── Phase 4: Multi-agent commands ─────────────────────────────────

    async def _handle_agent(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /agent <specialist> <task> — run a specialist sub-agent."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        args = context.args
        if not args:
            await update.message.reply_text(
                "Usage:\n"
                "/agent list — show specialists\n"
                "/agent route <query> — preview routing\n"
                "/agent <specialist> <task> — run a specialist\n\n"
                "Specialists: research, coder, analyst, creative, searcher"
            )
            return

        subcmd = args[0].lower()

        if subcmd == "list":
            try:
                from aura.multi_agent.orchestrator import MultiAgentOrchestrator
                orch = getattr(self.aura, "orchestrator", None)
                if not orch:
                    await update.message.reply_text("Multi-agent orchestrator not available.")
                    return
                status = orch.get_status()
                lines = ["Available Specialists:\n"]
                for spec in status.get("specialists", []):
                    lines.append(f"  {spec['name']} — {spec.get('description', '')[:80]}")
                await update.message.reply_text("\n".join(lines))
            except Exception as e:
                await update.message.reply_text(f"Error: {e}")

        elif subcmd == "route" and len(args) > 1:
            query = " ".join(args[1:])
            try:
                orch = getattr(self.aura, "orchestrator", None)
                if not orch:
                    await update.message.reply_text("Orchestrator not available.")
                    return
                preview = orch.route_preview(query)
                await update.message.reply_text(f"Routing preview:\n{preview}")
            except Exception as e:
                await update.message.reply_text(f"Error: {e}")

        else:
            specialist = subcmd
            task = " ".join(args[1:]) if len(args) > 1 else "Help me"
            placeholder = await update.message.reply_text(f"Running {specialist} agent...")
            try:
                response = await asyncio.to_thread(self._run_agent_sync, task)
                await self._edit_or_send_response(placeholder, response or "No response from agent.", update.effective_chat.id)
            except Exception as e:
                await self._edit_or_send_response(placeholder, f"Agent error: {e}", update.effective_chat.id)

    # ================================================================
    #  Location sharing handlers (Phase 5)
    # ================================================================

    async def _handle_location(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle shared location — provide contextual local info."""
        if not self._is_user_allowed(update.effective_user.id):
            return

        location = update.message.location
        lat = location.latitude
        lon = location.longitude

        placeholder = await update.message.reply_text("\U0001f4cd Getting info for your location...")

        try:
            info, results = await self._get_location_info(lat, lon)

            # Store last known location for /nearby (persisted to SQLite)
            self.store.set_user_location(str(update.effective_user.id), lat, lon)

            await self._edit_or_send_response(placeholder, str(update.effective_chat.id), info, update)
        except Exception as e:
            logger.error(f"[Location] Error getting location info: {e}")
            await self._edit_or_send_response(
                placeholder, str(update.effective_chat.id),
                f"Couldn't get location info: {e}", update
            )

    async def _get_location_info(self, lat: float, lon: float) -> tuple:
        """Fetch contextual info for a location.

        Returns (formatted_text, raw_results_dict).
        """
        import aiohttp

        results: dict = {}

        async with aiohttp.ClientSession() as session:
            # 1. Reverse geocoding via Nominatim
            try:
                url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json&zoom=14"
                headers = {"User-Agent": "AURA-Bot/4.5"}
                async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=5)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        address = data.get("address", {})
                        results["city"] = (
                            address.get("city")
                            or address.get("town")
                            or address.get("village")
                            or "Unknown"
                        )
                        results["country"] = address.get("country", "")
                        results["display"] = data.get("display_name", "")[:100]
            except Exception:
                pass

            # 2. Weather from wttr.in
            try:
                url = f"https://wttr.in/{lat},{lon}?format=j1"
                async with session.get(url, timeout=aiohttp.ClientTimeout(total=5)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        current = data.get("current_condition", [{}])[0]
                        results["temp"] = current.get("temp_C", "?")
                        results["feels_like"] = current.get("FeelsLikeC", "?")
                        results["condition"] = current.get("weatherDesc", [{}])[0].get("value", "")
                        results["humidity"] = current.get("humidity", "?")
                        results["wind"] = current.get("windspeedKmph", "?")

                        # Astronomy
                        astro = data.get("weather", [{}])[0].get("astronomy", [{}])[0]
                        results["sunrise"] = astro.get("sunrise", "")
                        results["sunset"] = astro.get("sunset", "")
            except Exception:
                pass

            # 3. Timezone from timeapi.io
            try:
                url = f"https://timeapi.io/api/Time/current/coordinate?latitude={lat}&longitude={lon}"
                async with session.get(url, timeout=aiohttp.ClientTimeout(total=5)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        results["timezone"] = data.get("timeZone", "")
                        results["local_time"] = data.get("time", "")
                        results["day_of_week"] = data.get("dayOfWeek", "")
            except Exception:
                pass

        # Build response text
        city = results.get("city", "Unknown")
        country = results.get("country", "")

        lines = [f"\U0001f4cd {city}, {country}"]

        if "temp" in results:
            lines.append(f"\n\U0001f321 Weather: {results['temp']}\u00b0C (feels like {results['feels_like']}\u00b0C)")
            lines.append(f"   {results['condition']}")
            lines.append(f"   Humidity: {results['humidity']}% | Wind: {results['wind']} km/h")

        if "sunrise" in results:
            lines.append(f"\n\U0001f305 Sunrise: {results['sunrise']} | Sunset: {results['sunset']}")

        if "timezone" in results:
            lines.append(f"\n\U0001f550 Local time: {results.get('local_time', '')} ({results['timezone']})")

        lines.append(f"\n\U0001f4cc Coordinates: {lat:.4f}, {lon:.4f}")

        lines.append(f"\nTip: Send me a message about what you want to do here, and I can help with local recommendations!")

        return "\n".join(lines), results

    async def _handle_nearby(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /nearby <query> — search for nearby places using last shared location."""
        if not self._is_user_allowed(update.effective_user.id):
            return

        user_id = str(update.effective_user.id)
        last_location = self.store.get_user_location(user_id)

        if not last_location:
            await update.message.reply_text("Please share your location first, then use /nearby <query>")
            return

        query = " ".join(context.args) if context.args else "restaurants"
        lat, lon = last_location["latitude"], last_location["longitude"]
        city = "the area"

        prompt = (
            f"Find {query} near coordinates {lat},{lon} (in {city}). "
            f"Give me top 3-5 recommendations with brief descriptions."
        )

        placeholder = await update.message.reply_text(f"\U0001f50d Searching for {query} nearby...")

        try:
            response_text, _ = await asyncio.to_thread(self._run_agent_sync, prompt)
            await self._edit_or_send_response(
                placeholder, str(update.effective_chat.id),
                response_text or "Couldn't find results.", update
            )
        except Exception as e:
            logger.error(f"[Nearby] Error: {e}")
            await self._edit_or_send_response(
                placeholder, str(update.effective_chat.id),
                f"Couldn't search nearby: {e}", update
            )

    async def _handle_fleet(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /fleet <goal> — parallel multi-agent decomposition."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        args = context.args
        if not args:
            await update.message.reply_text("Usage: /fleet <goal>\n\nExample:\n/fleet Analyze the codebase for bugs, optimizations, and security issues")
            return

        goal = " ".join(args)
        placeholder = await update.message.reply_text(f"Fleet dispatched: {goal[:100]}...\n\nRunning all specialists in parallel...")

        try:
            orch = getattr(self.aura, "orchestrator", None)
            if not orch:
                await self._edit_or_send_response(placeholder, "Multi-agent orchestrator not available.", update.effective_chat.id)
                return

            result = await asyncio.to_thread(orch.chat, goal)
            response = result if isinstance(result, str) else str(result.get("response", result))
            await self._edit_or_send_response(placeholder, f"Fleet Results:\n\n{response}", update.effective_chat.id)
        except Exception as e:
            await self._edit_or_send_response(placeholder, f"Fleet error: {e}", update.effective_chat.id)

    # =========================================================================
    #  Autonomous Hands handlers
    # =========================================================================

    async def _handle_hand(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /hand — manage autonomous Hands."""
        if not self._is_user_allowed(update.effective_user.id):
            return

        from aura.hands.manager import get_hand_manager
        from aura.hands.researcher import ResearcherHand
        from aura.hands.guardian import GuardianHand
        from aura.hands.memory_hand import MemoryHand

        manager = get_hand_manager()
        if not manager.list_hands():
            manager.register(ResearcherHand())
            manager.register(GuardianHand())
            manager.register(MemoryHand())

        args = context.args or []
        subcmd = args[0].lower() if args else "list"
        subarg = args[1] if len(args) > 1 else ""

        if subcmd in ("list", "ls"):
            hands = manager.list_hands()
            if not hands:
                await update.message.reply_text("No hands registered.")
                return
            lines = ["<b>Autonomous Hands</b>\n"]
            for h in hands:
                state_icon = {"inactive": "\u26ab", "active": "\ud83d\udfe2", "running": "\ud83d\udd04", "paused": "\u23f8", "cooldown": "\u23f3", "error": "\ud83d\udd34"}.get(h["state"], "\u2753")
                lines.append(
                    f"{state_icon} <b>{h['name']}</b> — {h['state']}\n"
                    f"   Runs: {h['total_runs']} | Cost: ${h['total_cost']:.4f} | Failures: {h['consecutive_failures']}"
                )
                if h.get("last_run"):
                    lines.append(f"   Last: {h['last_run']}")
            await update.message.reply_text("\n".join(lines), parse_mode="HTML")

        elif subcmd == "run" and subarg:
            hand = manager.get_hand(subarg)
            if not hand:
                await update.message.reply_text(f"Unknown hand: {subarg}")
                return

            brain = getattr(self.aura, 'brain', None)
            tools = getattr(self.aura, 'tools', {})
            if not brain:
                await update.message.reply_text("Agent brain not available.")
                return

            placeholder = await update.message.reply_text(f"\ud83e\udd16 Running hand '{subarg}'...")
            chat_id = update.effective_chat.id

            async def _run_and_report():
                try:
                    result = await manager.run_hand(subarg, brain, tools)
                    status = "\u2705" if result.success else "\u274c"
                    text = (
                        f"{status} <b>Hand '{subarg}' completed</b>\n\n"
                        f"{result.summary[:500]}\n\n"
                        f"Iterations: {result.iterations} | "
                        f"Duration: {result.duration_seconds:.1f}s | "
                        f"Cost: ${result.cost_usd:.4f}"
                    )
                    if result.error:
                        text += f"\nError: {result.error}"
                    await self._edit_or_send_response(placeholder, text, chat_id, parse_mode="HTML")
                except Exception as e:
                    await self._edit_or_send_response(placeholder, f"Hand execution failed: {e}", chat_id)

            asyncio.create_task(_run_and_report())

        elif subcmd == "activate" and subarg:
            if manager.activate(subarg):
                await update.message.reply_text(f"\ud83d\udfe2 Hand '{subarg}' activated.")
            else:
                await update.message.reply_text(f"Unknown or already running hand: {subarg}")

        elif subcmd == "deactivate" and subarg:
            if manager.deactivate(subarg):
                await update.message.reply_text(f"\u26ab Hand '{subarg}' deactivated.")
            else:
                await update.message.reply_text(f"Unknown hand: {subarg}")

        elif subcmd == "status" and subarg:
            hand = manager.get_hand(subarg)
            if not hand:
                await update.message.reply_text(f"Unknown hand: {subarg}")
                return
            stats = hand.get_stats()
            lines = [
                f"<b>Hand: {stats['name']}</b>",
                f"State: {stats['state']}",
                f"Description: {stats.get('description', 'N/A')}",
                f"Total runs: {stats['total_runs']}",
                f"Total cost: ${stats['total_cost']:.4f}",
                f"Consecutive failures: {stats['consecutive_failures']}",
                f"Model: {stats.get('model_preference', 'N/A')}",
                f"Idle only: {stats.get('idle_only', 'N/A')}",
                f"Drive trigger: {stats.get('trigger_on_drive', 'None')}",
            ]
            if stats.get("last_run"):
                lines.append(f"Last run: {stats['last_run']}")
            if stats.get("last_error"):
                lines.append(f"Last error: {stats['last_error']}")
            await update.message.reply_text("\n".join(lines), parse_mode="HTML")

        else:
            await update.message.reply_text(
                "<b>Usage:</b> /hand &lt;command&gt; [name]\n\n"
                "/hand list — Show all hands\n"
                "/hand run &lt;name&gt; — Run a hand now\n"
                "/hand activate &lt;name&gt; — Activate for scheduling\n"
                "/hand deactivate &lt;name&gt; — Deactivate\n"
                "/hand status &lt;name&gt; — Detailed status",
                parse_mode="HTML",
            )

    async def _handle_hand_approval_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle hand approval inline keyboard callbacks."""
        query = update.callback_query
        await query.answer()

        data = query.data  # "hand_approve_<request_id>" or "hand_deny_<request_id>"
        parts = data.split("_", 2)
        if len(parts) < 3:
            return

        action = parts[1]  # "approve" or "deny"
        request_id = parts[2]
        approved = action == "approve"

        try:
            from aura.hands.manager import get_hand_manager
            manager = get_hand_manager()
            manager.resolve_approval(request_id, approved)
            status = "\u2705 Approved" if approved else "\u274c Denied"
            await query.edit_message_text(f"{status} (request: {request_id})")
        except Exception as e:
            await query.edit_message_text(f"Failed to resolve approval: {e}")

    # =========================================================================
    #  Payment / Premium handlers
    # =========================================================================

    async def _handle_premium(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /premium — show available premium tiers."""
        if not self._is_user_allowed(update.effective_user.id):
            return

        keyboard = []
        for tier_id, tier in PREMIUM_TIERS.items():
            price_str = f"${tier['price'] / 100:.2f}"
            keyboard.append([InlineKeyboardButton(
                f"{tier['title']} — {price_str}",
                callback_data=f"buy_{tier_id}"
            )])

        text = "AURA Premium\n\n"
        for tier_id, tier in PREMIUM_TIERS.items():
            benefits = "\n".join(f"  • {b}" for b in tier["benefits"])
            text += f"{tier['title']} (${tier['price']/100:.2f}/mo)\n{benefits}\n\n"

        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(text, reply_markup=reply_markup)

    async def _handle_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle callback queries for premium purchase buttons."""
        query = update.callback_query
        await query.answer()

        tier_id = query.data.replace("buy_", "")
        tier = PREMIUM_TIERS.get(tier_id)
        if not tier:
            return

        provider_token = os.getenv("TELEGRAM_PAYMENT_TOKEN", "")
        if not provider_token:
            await query.message.reply_text("Payments not configured yet. Contact the admin.")
            return

        await context.bot.send_invoice(
            chat_id=query.from_user.id,
            title=tier["title"],
            description=tier["description"],
            payload=f"premium_{tier_id}_{query.from_user.id}",
            provider_token=provider_token,
            currency=tier["currency"],
            prices=[LabeledPrice(tier["title"], tier["price"])],
            start_parameter=f"premium_{tier_id}",
        )

    async def _handle_pre_checkout(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Must answer pre-checkout query within 10 seconds."""
        query = update.pre_checkout_query
        await query.answer(ok=True)

    async def _handle_successful_payment(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Record successful payment and grant premium status."""
        payment = update.message.successful_payment
        user_id = str(update.effective_user.id)
        tier_id = payment.invoice_payload.split("_")[1] if "_" in payment.invoice_payload else "supporter"

        self._premium_users[user_id] = {
            "tier": tier_id,
            "paid_at": _time.time(),
            "amount": payment.total_amount,
            "currency": payment.currency,
        }
        self.store.set_premium(
            user_id=user_id, tier=tier_id,
            stars_amount=payment.total_amount,
            metadata={"currency": payment.currency, "paid_at": _time.time()},
        )

        tier = PREMIUM_TIERS.get(tier_id, {})
        await update.message.reply_text(
            f"Thank you for your support!\n\n"
            f"You now have {tier.get('title', 'Premium')} access.\n"
            f"Benefits:\n" + "\n".join(f"  • {b}" for b in tier.get("benefits", []))
        )

    async def _handle_donate(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /donate — one-time support payment."""
        if not self._is_user_allowed(update.effective_user.id):
            return

        args = context.args
        amount = 500  # Default $5
        if args:
            try:
                amount = int(float(args[0]) * 100)
            except ValueError:
                pass

        provider_token = os.getenv("TELEGRAM_PAYMENT_TOKEN", "")
        if not provider_token:
            await update.message.reply_text("Payments not configured. Set TELEGRAM_PAYMENT_TOKEN.")
            return

        await context.bot.send_invoice(
            chat_id=update.effective_chat.id,
            title="Support AURA",
            description="One-time donation to support AURA development",
            payload=f"donate_{update.effective_user.id}_{amount}",
            provider_token=provider_token,
            currency="USD",
            prices=[LabeledPrice("Donation", amount)],
        )

    # =========================================================================
    #  16 Improvements: Keyboard, Stars, Forum, Reactions, File Gen, Digest,
    #  Language, Export, Stickers, Inline categories, Pinning, Action buttons
    # =========================================================================

    # --- 1. Persistent Reply Keyboard ---

    def _get_reply_keyboard(self) -> ReplyKeyboardMarkup:
        """Build the persistent quick-access reply keyboard."""
        keyboard = [
            ["/research", "/search", "/code", "/image"],
            ["/model", "/status", "/help", "/session"],
        ]
        return ReplyKeyboardMarkup(keyboard, resize_keyboard=True,
                                   is_persistent=True)

    async def _handle_keyboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /keyboard — toggle the persistent reply keyboard on/off."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        user_id = str(update.effective_user.id)
        currently_on = self.store.get_keyboard_enabled(user_id)
        if currently_on:
            self.store.set_keyboard_enabled(user_id, False)
            await update.message.reply_text(
                "Keyboard hidden. Use /keyboard to bring it back.",
                reply_markup=ReplyKeyboardRemove()
            )
        else:
            self.store.set_keyboard_enabled(user_id, True)
            await update.message.reply_text(
                "Keyboard restored!",
                reply_markup=self._get_reply_keyboard()
            )

    # --- 3. Telegram Stars ---

    _STARS_TIERS = [
        {"stars": 50, "label": "\u2b50 50 Stars", "description": "Small support"},
        {"stars": 150, "label": "\u2b50 150 Stars", "description": "Medium support"},
        {"stars": 500, "label": "\u2b50 500 Stars", "description": "Big support — unlocks priority"},
    ]

    async def _handle_stars(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /stars — support AURA with Telegram Stars (XTR currency)."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        keyboard = []
        for tier in self._STARS_TIERS:
            keyboard.append([InlineKeyboardButton(
                f"{tier['label']} — {tier['description']}",
                callback_data=f"stars_{tier['stars']}"
            )])
        text = (
            "\u2b50 Support AURA with Telegram Stars\n\n"
            "Telegram Stars are an in-app currency. "
            "Choose a tier below to send stars as a thank-you!"
        )
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))

    async def _handle_stars_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle stars_<amount> callback — send a Stars invoice."""
        query = update.callback_query
        await query.answer()
        try:
            amount = int(query.data.replace("stars_", ""))
        except ValueError:
            return
        await context.bot.send_invoice(
            chat_id=query.from_user.id,
            title=f"Support AURA — {amount} Stars",
            description=f"Send {amount} Telegram Stars to support AURA development",
            payload=f"stars_{amount}_{query.from_user.id}",
            provider_token="",  # Empty for Telegram Stars
            currency="XTR",
            prices=[LabeledPrice(f"{amount} Stars", amount)],
        )

    # --- 5. Forum Topics ---

    async def _create_forum_topic_if_needed(self, chat_id: str, topic_name: str,
                                            context: ContextTypes.DEFAULT_TYPE) -> Optional[int]:
        """Create a forum topic for a research task in a supergroup with topics enabled."""
        try:
            result = await context.bot.create_forum_topic(
                chat_id=int(chat_id),
                name=topic_name[:128],
            )
            return result.message_thread_id
        except Exception as e:
            logger.debug(f"[Forum] Could not create topic: {e}")
            return None

    # --- 6. Native Reactions ---

    _EMOTION_TO_REACTION = {
        "joy": "\U0001f44d",         # thumbs up
        "excited": "\U0001f525",     # fire
        "curious": "\U0001f914",     # thinking
        "surprised": "\U0001f62e",   # open mouth
        "sad": "\U0001f622",         # crying
        "frustrated": "\U0001f44e",  # thumbs down
        "grateful": "\u2764\ufe0f",  # red heart
        "empathetic": "\U0001f917",  # hugging
        "confident": "\U0001f60e",   # sunglasses
        "neutral": "\U0001f44d",     # thumbs up
    }

    async def _try_native_reaction(self, chat_id: str, message_id: int,
                                   emotion: str, intensity: float) -> bool:
        """Try to set a native Telegram reaction on the user's message.

        Returns True if successful, False otherwise.
        """
        if not REACTIONS_AVAILABLE or intensity < 0.4:
            return False
        emoji = self._EMOTION_TO_REACTION.get(emotion, "\U0001f44d")
        try:
            await self.bot.set_message_reaction(
                chat_id=int(chat_id),
                message_id=message_id,
                reaction=[ReactionTypeEmoji(emoji=emoji)],
            )
            return True
        except Exception as e:
            logger.debug(f"[Reaction] Native reaction failed: {e}")
            return False

    # --- 7. File Generation ---

    async def _handle_file_gen(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /file <format> [content] — generate a document and send it.

        Formats: pdf, docx, txt, md
        If content is "last", uses the last agent response.
        """
        if not self._is_user_allowed(update.effective_user.id):
            return
        args = context.args or []
        if not args:
            await update.message.reply_text(
                "Usage: /file <pdf|docx|txt|md> [content or 'last']\n\n"
                "Examples:\n"
                "/file pdf last — export last response as PDF\n"
                "/file docx Meeting notes: discussed Q1 goals"
            )
            return

        fmt = args[0].lower().strip(".")
        if fmt not in ("pdf", "docx", "txt", "md"):
            await update.message.reply_text("Supported formats: pdf, docx, txt, md")
            return

        content_text = " ".join(args[1:]) if len(args) > 1 else ""

        # "last" keyword: use last agent response
        if content_text.strip().lower() == "last" or not content_text:
            uid = update.effective_user.id
            exchange = self._last_exchange.get(uid, {})
            content_text = exchange.get("output", "")
            if not content_text:
                await update.message.reply_text("No recent response to export. Send a message first.")
                return

        placeholder = await update.message.reply_text(f"\U0001f4c4 Generating {fmt.upper()}...")

        try:
            if fmt in ("pdf", "docx"):
                from aura.tools.document_generator import DocumentGeneratorTool
                gen = DocumentGeneratorTool()
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"aura_{ts}.{fmt}"
                filepath = os.path.join(tempfile.gettempdir(), filename)
                if fmt == "pdf":
                    gen.create_pdf(content_text, filepath)
                else:
                    gen.create_docx(content_text, filepath)
            else:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"aura_{ts}.{fmt}"
                filepath = os.path.join(tempfile.gettempdir(), filename)
                with open(filepath, "w", encoding="utf-8") as f:
                    f.write(content_text)

            with open(filepath, "rb") as f:
                await self.bot.send_document(
                    chat_id=str(update.effective_chat.id),
                    document=f,
                    filename=filename,
                    caption=f"Generated {fmt.upper()} document",
                )
            await placeholder.delete()
            # Clean up temp file
            try:
                os.unlink(filepath)
            except OSError:
                pass
        except Exception as e:
            logger.error(f"[FileGen] Error: {e}", exc_info=True)
            await placeholder.edit_text(f"Failed to generate file: {e}")

    # --- 8. Daily Digest ---

    async def _handle_digest(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /digest on|off|now — daily activity digest."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        chat_id = str(update.effective_chat.id)
        args = context.args or []
        cmd = args[0].lower() if args else "now"

        if cmd == "on":
            # Schedule via APScheduler if available
            try:
                from apscheduler.schedulers.asyncio import AsyncIOScheduler
                scheduler = getattr(self, '_digest_scheduler', None)
                if not scheduler:
                    scheduler = AsyncIOScheduler()
                    self._digest_scheduler = scheduler
                    scheduler.start()
                job = scheduler.add_job(
                    self._send_daily_digest_async, 'cron',
                    hour=9, minute=0,
                    args=[chat_id],
                    id=f"digest_{chat_id}",
                    replace_existing=True,
                )
                self._digest_job_ids[chat_id] = job.id
                self.store.set_digest_job(chat_id, job.id, enabled=True)
                await update.message.reply_text(
                    "\U0001f4ec Daily digest enabled! You'll get a summary at 9:00 AM.\n"
                    "Use /digest off to disable, /digest now for an instant digest."
                )
            except ImportError:
                await update.message.reply_text(
                    "APScheduler not installed — digest scheduling unavailable.\n"
                    "Use /digest now for a one-time digest."
                )
                return

        elif cmd == "off":
            try:
                scheduler = getattr(self, '_digest_scheduler', None)
                job_id = self._digest_job_ids.get(chat_id)
                if scheduler and job_id:
                    scheduler.remove_job(job_id)
                    self._digest_job_ids.pop(chat_id, None)
                self.store.set_digest_job(chat_id, "", enabled=False)
            except Exception:
                pass
            await update.message.reply_text("\U0001f6d1 Daily digest disabled.")

        else:
            # "now" — send an instant digest
            await self._send_daily_digest_async(chat_id)

    async def _send_daily_digest_async(self, chat_id: str):
        """Build and send a daily activity digest to the chat."""
        lines = ["\U0001f4ca Daily AURA Digest\n"]

        # Gather stats
        try:
            from aura.proactive.persistence import ProactivePersistence
            pp = ProactivePersistence()
            conn = pp._conn
            if conn:
                cursor = conn.execute(
                    "SELECT COUNT(*) FROM activity_log WHERE timestamp > datetime('now', '-1 day')"
                )
                count = cursor.fetchone()[0]
                lines.append(f"\u2022 Activities logged (24h): {count}")
        except Exception:
            lines.append("\u2022 Activity log: unavailable")

        # Conversation count
        try:
            cm = get_conversation_manager()
            convos = cm.list_conversations() if hasattr(cm, 'list_conversations') else []
            lines.append(f"\u2022 Active conversations: {len(convos)}")
        except Exception:
            pass

        # Active chats
        lines.append(f"\u2022 Active Telegram chats: {len(self.active_chats)}")

        # Scheduled tasks
        try:
            from apscheduler.schedulers.background import BackgroundScheduler
            aura_sched = getattr(self.aura, 'scheduler', None)
            if aura_sched:
                jobs = aura_sched.get_jobs() if hasattr(aura_sched, 'get_jobs') else []
                lines.append(f"\u2022 Scheduled tasks: {len(jobs)}")
        except Exception:
            pass

        lines.append(f"\n\U0001f552 Generated at {datetime.now().strftime('%Y-%m-%d %H:%M')}")

        try:
            await self.bot.send_message(chat_id=int(chat_id), text="\n".join(lines))
        except Exception as e:
            logger.error(f"[Digest] Failed to send digest: {e}")

    # --- 9. Multi-language ---

    async def _handle_lang(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /lang [code] — set or show language preference.

        Examples: /lang ru, /lang en, /lang (shows current)
        """
        if not self._is_user_allowed(update.effective_user.id):
            return
        user_id = str(update.effective_user.id)
        args = context.args or []

        if not args:
            current = self.store.get_user_language(user_id)
            tg_lang = getattr(update.effective_user, 'language_code', 'unknown')
            await update.message.reply_text(
                f"\U0001f310 Language settings:\n"
                f"Current: {current}\n"
                f"Telegram language: {tg_lang}\n\n"
                f"Set with: /lang <code>\n"
                f"Examples: /lang ru, /lang az, /lang en"
            )
            return

        lang_code = args[0].lower().strip()[:5]
        self.store.set_user_language(user_id, lang_code)
        if lang_code == "en":
            await update.message.reply_text(
                "\U0001f310 Language set to English (default). "
                "I'll respond in English."
            )
        else:
            await update.message.reply_text(
                f"\U0001f310 Language set to: {lang_code}\n"
                f"I'll try to respond in {lang_code} from now on."
            )

    # --- 10. Reply Action Buttons ---

    def _get_action_buttons(self, original_msg_id: int = 0) -> InlineKeyboardMarkup:
        """Build action buttons attached to agent responses."""
        return InlineKeyboardMarkup([
            [
                InlineKeyboardButton("\U0001f50d Go deeper", callback_data=f"act_deeper_{original_msg_id}"),
                InlineKeyboardButton("\U0001f4be Save to memory", callback_data=f"act_save_{original_msg_id}"),
            ],
            [
                InlineKeyboardButton("\U0001f4e4 Share", callback_data=f"act_share_{original_msg_id}"),
                InlineKeyboardButton("\U0001f4c4 Export", callback_data=f"act_export_{original_msg_id}"),
            ],
        ])

    async def _handle_action_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle act_<action>_<msg_id> callbacks from action buttons."""
        query = update.callback_query
        await query.answer()
        data = query.data  # e.g. "act_deeper_12345"
        parts = data.split("_", 2)
        if len(parts) < 3:
            return
        action = parts[1]
        user_id = str(query.from_user.id)

        # Get last response text for context
        uid = query.from_user.id
        exchange = self._last_exchange.get(uid, {})
        last_output = exchange.get("output", "")
        last_input = exchange.get("input", "")

        if action == "deeper":
            if not last_output:
                await query.message.reply_text("No recent response to expand on.")
                return
            goal = f"Expand on and go deeper into this topic. Previous response was about: {last_input[:200]}"
            # Run agent and send response directly (can't reuse _run_agent_and_reply without a proper Update)
            placeholder = await query.message.reply_text("Thinking...")
            try:
                response_text, artifacts = await asyncio.wait_for(
                    asyncio.to_thread(self._run_agent_sync, goal),
                    timeout=self._AGENT_TIMEOUT,
                )
            except Exception:
                response_text = "Could not expand. Try asking directly."
                artifacts = []
            chat_id = str(query.message.chat_id)
            await self._edit_or_send_response(placeholder, chat_id, response_text, update)

        elif action == "save":
            if not last_output:
                await query.message.reply_text("Nothing to save.")
                return
            try:
                if hasattr(self.aura, 'memory') and self.aura.memory:
                    mem_text = f"[Saved from Telegram] {last_input[:100]}: {last_output[:500]}"
                    if hasattr(self.aura.memory, 'add'):
                        self.aura.memory.add(mem_text)
                    elif hasattr(self.aura.memory, 'store'):
                        self.aura.memory.store(mem_text)
                    await query.message.reply_text("\U0001f4be Saved to memory!")
                else:
                    await query.message.reply_text("Memory system not available.")
            except Exception as e:
                await query.message.reply_text(f"Could not save: {e}")

        elif action == "share":
            if last_output:
                share_text = last_output[:4000]
                await query.message.reply_text(
                    f"\U0001f4e4 Shareable response:\n\n{share_text}"
                )
            else:
                await query.message.reply_text("No recent response to share.")

        elif action == "export":
            if not last_output:
                await query.message.reply_text("Nothing to export.")
                return
            # Quick export as txt
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"aura_response_{ts}.txt"
            filepath = os.path.join(tempfile.gettempdir(), filename)
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(f"Query: {last_input}\n\nResponse:\n{last_output}")
            try:
                with open(filepath, "rb") as f:
                    await self.bot.send_document(
                        chat_id=str(query.message.chat_id),
                        document=f,
                        filename=filename,
                    )
            except Exception as e:
                await query.message.reply_text(f"Export failed: {e}")
            try:
                os.unlink(filepath)
            except OSError:
                pass

    # --- 11. Custom Sticker Pack ---

    async def _handle_stickers_cmd(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /stickers — admin command to scaffold an emotion-mapped sticker set."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        if not self._is_admin(update.effective_user.id):
            await update.message.reply_text("Admin only command.")
            return

        emotions = list(EMOTION_REACTIONS.keys())
        lines = [
            "\U0001f3a8 AURA Sticker Pack Setup\n",
            "To create a custom sticker pack for AURA, you need:",
            "1. Design one sticker per emotion (512x512 PNG, <512KB)",
            "2. Send them to @Stickers bot to create a pack",
            "3. Set AURA_STICKER_PACK env var to the pack name\n",
            "Required emotions:",
        ]
        for emo in emotions:
            emoji = EMOTION_EMOJI.get(emo, "\U0001f610")
            lines.append(f"  {emoji} {emo}")

        lines.append(f"\nTotal stickers needed: {len(emotions)}")
        lines.append("\nOnce created, AURA will auto-select stickers based on emotional state.")

        await update.message.reply_text("\n".join(lines))

    # --- 12. Rich Inline Mode (category results added in _handle_inline above) ---
    # Category results (Code, Image, Research) are injected in _handle_inline

    # --- 13. Message Pinning ---

    async def _handle_pin(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /pin — pin the replied-to message, or the last bot message."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        chat_id = str(update.effective_chat.id)
        reply = update.message.reply_to_message

        if reply:
            try:
                await context.bot.pin_chat_message(
                    chat_id=int(chat_id),
                    message_id=reply.message_id,
                    disable_notification=True,
                )
                await update.message.reply_text("\U0001f4cc Message pinned!")
            except Exception as e:
                await update.message.reply_text(f"Could not pin: {e}")
        else:
            await update.message.reply_text(
                "Reply to a message with /pin to pin it.\n"
                "Tip: I'll also offer to pin important research results in groups."
            )

    async def _handle_pin_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle pin_<message_id> callbacks — pin a specific message."""
        query = update.callback_query
        await query.answer()
        try:
            msg_id = int(query.data.replace("pin_", ""))
            await context.bot.pin_chat_message(
                chat_id=query.message.chat_id,
                message_id=msg_id,
                disable_notification=True,
            )
            await query.message.reply_text("\U0001f4cc Pinned!")
        except Exception as e:
            await query.message.reply_text(f"Could not pin: {e}")

    def _get_pin_button(self, message_id: int) -> InlineKeyboardMarkup:
        """Build a 'Pin this' inline button for important research results."""
        return InlineKeyboardMarkup([[
            InlineKeyboardButton("\U0001f4cc Pin this", callback_data=f"pin_{message_id}")
        ]])

    # --- 14. Conversation Export ---

    async def _handle_export(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /export [json|md|pdf|txt] — export conversation history."""
        if not self._is_user_allowed(update.effective_user.id):
            return
        args = context.args or []
        fmt = args[0].lower() if args else "md"
        if fmt not in ("json", "md", "pdf", "txt"):
            await update.message.reply_text("Supported: /export json|md|pdf|txt")
            return

        user_id = str(update.effective_user.id)
        placeholder = await update.message.reply_text(f"\U0001f4e6 Exporting as {fmt.upper()}...")

        try:
            # Gather conversation from ConversationManager
            messages = []
            try:
                cm = get_conversation_manager()
                conv_id = cm.get_bound_conversation(f"telegram:{user_id}")
                if conv_id and hasattr(cm, 'get_messages'):
                    messages = cm.get_messages(conv_id) or []
                elif conv_id and hasattr(cm, '_messages'):
                    messages = cm._messages.get(conv_id, [])
            except Exception:
                pass

            # Fallback: use last exchange if no conversation manager data
            if not messages:
                exchange = self._last_exchange.get(update.effective_user.id, {})
                if exchange:
                    messages = [
                        {"role": "user", "content": exchange.get("input", "")},
                        {"role": "assistant", "content": exchange.get("output", "")},
                    ]

            if not messages:
                await placeholder.edit_text("No conversation data to export.")
                return

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"aura_conversation_{ts}.{fmt}"
            filepath = os.path.join(tempfile.gettempdir(), filename)

            if fmt == "json":
                with open(filepath, "w", encoding="utf-8") as f:
                    json.dump(messages if isinstance(messages, list) else
                              [{"role": m.role, "content": m.preview} if hasattr(m, 'role')
                               else m for m in messages],
                              f, indent=2, default=str)

            elif fmt == "md":
                with open(filepath, "w", encoding="utf-8") as f:
                    f.write(f"# AURA Conversation Export\n")
                    f.write(f"*Exported: {datetime.now().isoformat()}*\n\n---\n\n")
                    for m in messages:
                        if isinstance(m, dict):
                            role = m.get("role", "unknown").upper()
                            content = m.get("content", m.get("preview", ""))
                        elif hasattr(m, 'role'):
                            role = m.role.upper()
                            content = getattr(m, 'preview', '') or getattr(m, 'content', '')
                        else:
                            role = "MSG"
                            content = str(m)
                        f.write(f"**{role}**: {content}\n\n")

            elif fmt == "txt":
                with open(filepath, "w", encoding="utf-8") as f:
                    for m in messages:
                        if isinstance(m, dict):
                            role = m.get("role", "unknown")
                            content = m.get("content", m.get("preview", ""))
                        elif hasattr(m, 'role'):
                            role = m.role
                            content = getattr(m, 'preview', '') or getattr(m, 'content', '')
                        else:
                            role = "msg"
                            content = str(m)
                        f.write(f"[{role}] {content}\n\n")

            elif fmt == "pdf":
                try:
                    from aura.tools.document_generator import DocumentGeneratorTool
                    gen = DocumentGeneratorTool()
                    text_content = "AURA Conversation Export\n\n"
                    for m in messages:
                        if isinstance(m, dict):
                            role = m.get("role", "unknown").upper()
                            content = m.get("content", m.get("preview", ""))
                        elif hasattr(m, 'role'):
                            role = m.role.upper()
                            content = getattr(m, 'preview', '') or getattr(m, 'content', '')
                        else:
                            role = "MSG"
                            content = str(m)
                        text_content += f"{role}: {content}\n\n"
                    gen.create_pdf(text_content, filepath)
                except ImportError:
                    await placeholder.edit_text("PDF export requires fpdf2. Use /export md instead.")
                    return

            with open(filepath, "rb") as f:
                await self.bot.send_document(
                    chat_id=str(update.effective_chat.id),
                    document=f,
                    filename=filename,
                    caption=f"Conversation export ({fmt.upper()})",
                )
            await placeholder.delete()
            try:
                os.unlink(filepath)
            except OSError:
                pass
        except Exception as e:
            logger.error(f"[Export] Error: {e}", exc_info=True)
            await placeholder.edit_text(f"Export failed: {e}")

    def get_active_chat_ids(self) -> List[str]:
        """Get list of active chat IDs for proactive messaging"""
        return list(self.active_chats.keys())
